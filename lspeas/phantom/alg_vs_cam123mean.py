""" Compare algortihm to mean of cam123 (=reference)

"""

from stentseg.utils.datahandling import select_dir
from lspeas.phantom.motion_pattern_error import readAnalysisExcel
import openpyxl # http://openpyxl.readthedocs.org/

def write_alg_vs_cam_excel(dir, ttCam, ttCamS, ppCam, ppCamS, errorsPos, errorsPosAbs, 
        errorsPosAbsMeanoverall, errorsPosAbsStdoverall, errorsPosAbsMeanOfLandmarks,
        errorsPosAbsStdOfLandmarks, MAE_errors, rmse_errors, ACam, Alandmarks, Aerrors):
            """
            """
            wb = openpyxl.Workbook()
            dest_filename = 'motion_pattern_error_out.xlsx'
            # wb = openpyxl.load_workbook(os.path.join(dir, dest_filename))
            ws = wb.active
            # errors_profile = np.mean(np.vstack(errors_periods), axis=0) # mean of periods in cam signal
            ws.cell(row=1, column=1).value = 'cam123 ref'
            ws.cell(row=2, column=1).value = 'time (s)'
            ws.cell(row=3, column=1).value = 'position (mm)'
            ws.cell(row=4, column=1).value = 'cam123 ref sampled'
            ws.cell(row=5, column=1).value = 'time (s)'
            ws.cell(row=6, column=1).value = 'position (mm)'
            for i, t in enumerate(ttCam):
                ws.cell(row=2, column=i+2).value = t
            for i, p in enumerate(ppCam):
                ws.cell(row=3, column=i+2).value = p
            for i, t in enumerate(ttCamS):
                ws.cell(row=5, column=i+2).value = t
            for i, p in enumerate(ppCamS):
                ws.cell(row=6, column=i+2).value = p   
            col = 2
            ws.cell(row=8, column=1).value = 'errorsPosition for points on ring analyzed'
            for i, phase in enumerate(['0%','10%','20%','30%','40%','50%','60%','70%','80%','90%','0%']):
                ws.cell(row=9+i, column=1).value = phase   
            for i, errorPos in enumerate(errorsPos):
                for j, e in enumerate(errorPos):
                    ws.cell(row=j+9, column=col+i).value = e
            ws.cell(row=j+9+2, column=1).value = 'abs of errorsPosition for points on ring analyzed'
            for i, errorPosAbs in enumerate(errorsPosAbs):
                for k, e in enumerate(errorPosAbs):
                    ws.cell(row=j+9+3+k, column=col+i).value = e
            ws.cell(row=j+9+3+k+2, column=1).value = 'errorsPosAbsMeanoverall+/-errorsPosAbsStdoverall'
            ws.cell(row=j+9+3+k+3, column=2).value = errorsPosAbsMeanoverall
            ws.cell(row=j+9+3+k+3, column=3).value = errorsPosAbsStdoverall
            
            ws.cell(row=j+9+3+k+5, column=2).value = 'errorsPosAbsMeanOfLandmarks'
            for i, e_per_pos_landmarks in enumerate(errorsPosAbsMeanOfLandmarks):
                ws.cell(row=j+9+3+k+6+i, column=2).value = e_per_pos_landmarks
            ws.cell(row=j+9+3+k+5, column=3).value = 'errorsPosAbsStdOfLandmarks'
            for i, estd_per_pos_landmarks in enumerate(errorsPosAbsStdOfLandmarks):
                ws.cell(row=j+9+3+k+6+i, column=3).value = estd_per_pos_landmarks
            
            ws.cell(row=j+9+3+k+6+i+2, column=1).value = 'MAE_errors for each landmark'    
            ws.cell(row=j+9+3+k+6+i+4, column=1).value = 'rmse_errors for each landmark'
            for l, MAE in enumerate(MAE_errors):
                ws.cell(row=j+9+3+k+6+i+3, column=l+2).value = MAE
            for l, rmse_error in enumerate(rmse_errors):
                ws.cell(row=j+9+3+k+6+i+5, column=l+2).value = rmse_error
            # amplitude diffs
            ws.cell(row=j+9+3+k+6+i+7, column=1).value = 'Amplitude cam123 ref'
            ws.cell(row=j+9+3+k+6+i+8, column=2).value = ACam
            ws.cell(row=j+9+3+k+6+i+9, column=1).value = 'Amplitude for each landmark'
            for m, A in enumerate(Alandmarks):
                ws.cell(row=j+9+3+k+6+i+10, column=m+2).value = A
            ws.cell(row=j+9+3+k+6+i+12, column=1).value = 'Amplitude error for each landmark'
            for m, Ae in enumerate(Aerrors):
                ws.cell(row=j+9+3+k+6+i+13, column=m+2).value = Ae
                ws.cell(row=j+9+3+k+6+i+15, column=m+2).value = abs(Ae)
            
            # save excel
            wb.save(os.path.join(dir, dest_filename))
            


# preset dirs
dirsave =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
exceldir = select_dir(r'C:\Users\Maaike\Dropbox\UTdrive\LSPEAS\Analysis\Validation robot', 
                r'D:\Profiles\koenradesma\Dropbox\UTdrive\LSPEAS\Analysis\Validation robot')
workbookAlg = '20160624 DATA Toshiba.xlsx'

profile = 'B4'
saveFig = True
saveErrorsExcel = True

# cam data from camera_error.py
ttCam = ttperiodmeanC123
ttCamRep2 = ttperiodmeanC123rep
ppCam = pperiodsC123bestCutMean
# ppCamRep = pperiodsC123bestCutMeanRep
ppCamStd = pperiodsC123bestCutStd
# ppCamStdRep = pperiodsC123bestCutStdRep

# shift ppCam, have ppCam period end when position is 0, since alg is 0 at t=T
zero  = 0.002 # noise limit 0.02
for i in [-2,-1]: # used lag 2 so check for last 2 points
    if ppCam[i] < zero: # move to start of signal
        if i == -2:
            ppCam = np.insert(ppCam, 0, ppCam[-1])
            ppCam = np.insert(ppCam, 0, (ppCam[-2]+ppCam[-3])/2)
            ppCam = ppCam[:i]
            ppCamStd = np.insert(ppCamStd, 0, ppCamStd[-1])
            ppCamStd = np.insert(ppCamStd, 0, (ppCamStd[-2]+ppCamStd[-3])/2)
            ppCamStd = ppCamStd[:-2]
            break
        if i == -1:
            ppCam = np.insert(ppCam, 0, ppCam[-1])
            ppCam = ppCam[:i]
            ppCamStd = np.insert(ppCamStd, 0, (ppCamStd[-1]+ppCamStd[-1])/2)
            ppCamStd = ppCamStd[:-1]
            break

# resample cam signal to match algorithm point interval (n=11; 10 phases)
samplepoints = 11
ttCamS, ppCamS = resample(ttCam,ppCam, num=samplepoints)
# repeat cam pattern
ttCamSrep, ppCamSrep = repeatCamPeriod(ttCamS, ppCamS, correct0=False)
ttCamRep, ppCamRep, ppCamStdRep  = repeatCamPeriod(ttCam, ppCam, ppCamStd, correct0=False)

# read algorithm data of ring-stent points that were analyed
ppall = readAnalysisExcel(exceldir, workbookAlg, sheetProfile) # reads 8x10x3

pzall = []
for i_point in range(len(ppall)):
    pz = ppall[i_point][:,2] # for a point, z-axis
    pz = np.append(pz, pz[0]) # make signal continuous, full period
    pzall.append(pz)

# get average of all pp periods of the ring-stent points
pzMean =  np.mean(pzall, axis=0)
pzMedian =  np.median(pzall, axis=0)
pzMax = np.max(pzall, axis=0)
pzMin = np.min(pzall, axis=0)
pzStd = np.std(pzall, axis=0)
pz25 = np.percentile(pzall, 25, axis=0)
pz75 = np.percentile(pzall, 75, axis=0)

# repeat alg
pzMean = np.asarray(list(pzMean) * 3)
pzMin = np.asarray(list(pzMin) * 3)
pzMax = np.asarray(list(pzMax) * 3)
pzStd = np.asarray(list(pzStd) * 3)

# plot
fignum = 5
xlim = 2.1 # 1.5, 2.1, 1.1
ylim = (ppall[:,:,2]).max() + 0.2 
f1 = plt.figure(figsize=(9,5.5), num=fignum); plt.clf()
ax4 = f1.add_subplot(111)

# plot camera data from camera_error.py
ax4.plot(ttCamRep, ppCamRep, 'k.-', label='reference (camera)')
ax4.fill_between(ttCamRep, ppCamRep-ppCamStdRep,     
            ppCamRep+ppCamStdRep, color='k', alpha=0.2)

# plot cam ref sampled
ax4.plot(ttCamSrep, ppCamSrep, 'ks', alpha=0.6)

# plot algorithm ring-stent points that were analyzed
# for i, pp in enumerate(pzall):
#     if i == 0:
#         ax4.plot(ttCamS, pp, 'gs-', alpha=0.5, label='algorithm')
#     else:
#          ax4.plot(ttCamS, pp, 'gs-', alpha=0.5)

ax4.plot(ttCamSrep, pzMean, 'rs-', alpha=0.6, label='algorithm') #mean of ring-stent points
ax4.plot(ttCamSrep, pzMax, 'r--') # dotted line for min and max
ax4.plot(ttCamSrep, pzMin, 'r--')
ax4.fill_between(ttCamSrep, pzMean-pzStd,     
            pzMean+pzStd, color='r', alpha=0.2)

_initaxis([ax4], legend='upper right', xlabel='time (s)', ylabel='position (mm)',
           legendtitle=profile)
major_ticks = np.arange(0, xlim, 0.2)
ax4.set_ylim((0, ylim))
ax4.set_xlim(-0.02,xlim)
ax4.set_xticks(major_ticks)

# store fig
if saveFig:
    name = 'alg_cam123mean_{}.pdf'.format(sheetProfile)
    f1.savefig(os.path.join(dirsave, name), papertype='a0', dpi=300)

# ============================================
# errors for pointpositions alg vs cam ref

# exclude last timepoint  which is again t=0
pzall090 = [ppz[:-1] for ppz in pzall]

errorsPos = pzall090 - ppCamS[:-1] # Pos = position
errorsPosAbs = abs(errorsPos)
errorsPosAbsMeanoverall = np.mean(errorsPosAbs) # meanMAEprofile
errorsPosAbsStdoverall = np.std(errorsPosAbs)
errorsPosAbsMeanOfLandmarks = np.mean(errorsPosAbs, axis=0) # MEA positions, n = 11, ring points averaged
errorsPosAbsStdOfLandmarks = np.std(errorsPosAbs, axis=0) # n = 11
# errorPosAbs = np.mean(errorsPosAbsMeanOfLandmarks) # from n=11 pointpositions
# errorPosAbsstd = np.std(errorsPosAbsMeanOfLandmarks) # from n=11 pointpositions

rmse_errors = []
MAE_errors = []
errorsStd = []
for i, error in enumerate(errorsPos):
    print(abs(error).max())
    # MeanAbsError MAE
    MAE = np.mean(abs(error))
    errorStd = np.std(abs(error))
    MAE_errors.append(MAE)
    errorsStd.append(errorStd)
    # root mean squared error
    rmse_error = rmse(ppCamS[:-1], pzall090[i])
    rmse_errors.append(rmse_error)

# positions errors of analyzed points
meanRMSEprofile = np.mean(rmse_errors) # from n analyzed points 
maxRMSEprofile = np.max(rmse_errors) 
meanMAEprofile = np.mean(MAE_errors) # from n analyzed points
maxMAEprofile =  np.max(MAE_errors) # ring point with largest MAE error

print('rmse of motion pattern for all points={}'.format(meanRMSEprofile))
print('mean abs error of motion pattern for all points={}/{}'.format(meanMAEprofile, errorsPosAbsMeanoverall))
print('max abs error for position of points analyzed: ', abs(errorsPos).max())

# plot
fignum = 6
ylim2 = 0.5
f2 = plt.figure(figsize=(9,11), num=fignum); plt.clf()
ax5 = f2.add_subplot(211)
for i, ee in enumerate(errorsPosAbs):
    if i == 0:
        ax5.plot(ttCamS[:-1], ee, 'gs-', alpha=0.5, label='errors per stent point')
    else:
         ax5.plot(ttCamS[:-1], ee, 'gs-', alpha=0.5)

ax5.plot(ttCamS[:-1], errorsPosAbsMeanOfLandmarks, 'rs-', label='errors mean of ring-stent points')
ax5.fill_between(ttCamS[:-1], errorsPosAbsMeanOfLandmarks-errorsPosAbsStdOfLandmarks,     
            errorsPosAbsMeanOfLandmarks+errorsPosAbsStdOfLandmarks, color='r', alpha=0.2)

_initaxis([ax5], legend='upper right', xlabel='time (s)', ylabel='abs error (mm)')
ax5.set_ylim((0, ylim2))
ax5.set_xlim(-0.02,xlim)
ax5.set_xticks(major_ticks)

# plot points vs error
ax6 = f2.add_subplot(212)
ax6.plot(range(len(MAE_errors)), MAE_errors, 'rs-', label='MAE mean of pointpositions')
ax6.fill_between(range(1,(len(MAE_errors)+1)), np.asarray(MAE_errors)-np.asarray(errorsStd),     
            np.asarray(MAE_errors)+np.asarray(errorsStd), color='y', alpha=0.2)
_initaxis([ax6], legend='upper right', xlabel='ring points', ylabel='MAE (mm)')


# ============================================
# errors for amplitude alg vs cam ref
ACam = max(ppCam)
Alandmarks = [max(ppz) for ppz in pzall]
Aerrors = Alandmarks - ACam

fignum = 7
f3 = plt.figure(figsize=(9,5.5), num=fignum); plt.clf()
ax7 = f3.add_subplot(111)
ax7.plot(range(1,(len(Aerrors)+1)), Aerrors, 'rs-', label='amplitude error per stent point')
_initaxis([ax7], legend='upper right', xlabel='ring points', ylabel='error amplitude (mm)')


if saveErrorsExcel:
    write_alg_vs_cam_excel(dirsave, ttCam, ttCamS, ppCam, ppCamS, errorsPos, 
        errorsPosAbs, errorsPosAbsMeanoverall, errorsPosAbsStdoverall, 
        errorsPosAbsMeanOfLandmarks, errorsPosAbsStdOfLandmarks, MAE_errors,
        rmse_errors, ACam, Alandmarks, Aerrors)
    print('******************************')

    
    