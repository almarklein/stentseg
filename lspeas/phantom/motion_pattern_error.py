# Code to analyze error between motion pattern found by algortihm and camera ground truth
 
# Created 4/7/2016
# Maaike Koenrades

import os
import openpyxl # http://openpyxl.readthedocs.org/
import matplotlib.pyplot as plt
import matplotlib as mpl
import prettyplotlib as ppl
from prettyplotlib import brewer2mpl # colormaps
import numpy as np
import scipy
import string
from peakdetection import peakdet


def readCameraExcel(exceldir, workbookCam, sheetProfile, colSt='B'):
    """ Read camera patterns. Start at colSt column.
    """
    wb = openpyxl.load_workbook(os.path.join(exceldir, workbookCam))
    sheet = wb.get_sheet_by_name(sheetProfile)
    start = col2num(colSt)-1
    r1 = sheet.rows[0][start:]
    r2 = sheet.rows[1][start:]
    time = [obj.value for obj in r1] 
    positions = [obj.value for obj in r2]
    
    return time, positions

def getCameraPeriod(time, positions, T):
    """ get one period of camera signal from time[0] based on T of period
    """
    t0 = time[0]
    tend = t0 + T
    tdif = [t-tend for t in time if t-tend<0]
    end = len(tdif)+1 # get 1 timepoint up to
    time = time[:end]
    positions = positions[:end]
    # positions = np.asarray(positions)-min(positions) # so that positions have value 0
    
    return time, positions

def getFreqCamera(t, signal):
    import scipy
    import scipy.fftpack
    import pylab
    from scipy import pi
    FFT = abs(scipy.fft(signal))
    freqs = scipy.fftpack.fftfreq(len(signal), t[1]-t[0])
    
    pylab.subplot(211)
    pylab.plot(t, signal)
    pylab.subplot(212)
    pylab.plot(freqs,20*scipy.log10(FFT),'x') # 20log10 provides conversion for a magnitude spectrum
    pylab.show()
    return freqs, 20*scipy.log10(FFT)

def readAnalysisExcel(exceldir, workbookAlg, sheetProfile, cols=['G','H','I'], startRows=[18,31,55,68,92,105,129,142]):
    """ To read 'Data Toshiba/Siemens' with algorithm analysis. 
    Return nx10x3 for n points starting at every startRow
    """
    import numpy as np
    
    wb = openpyxl.load_workbook(os.path.join(exceldir, workbookAlg), data_only=True)
    sheet = wb.get_sheet_by_name(sheetProfile)
    for startRow in startRows:
        for col in cols: # x,y,z
            p1 = tuple(sheet[col+str(startRow):col+str(startRow+9) ])
            p1 = [obj[0].value for obj in p1]
            p1 = np.vstack(p1)
            if col == cols[0]:
                p1_d = p1
            else:
                p1_d = np.hstack( [ p1_d, p1] )
        if startRow == startRows[0]:
            pp = p1_d
        else: 
            pp = np.append(pp, p1_d, axis=0) 
        
    return pp.reshape(len(startRows),10,len(cols))


def cor_timeshift(cor_seq, y, x):
    """ obtain physical offsets per lag position
    """
    #Generate an x axis
    xcorr = np.arange(cor_seq.size)
    #Convert this into lag units, not physical yet
    lags = xcorr - (y.size-1)
    distancePerLag = (x[-1] - x[0])/float(x.size-1)  # timestep in data
    #Convert lags into physical units
    offsets = -lags*distancePerLag
    
    return offsets, lags

def resample(x,y, num=50, Tnew=None):
    """ Use the univariate interpolators in scipy.interpolate to resample x,y
    Tnew must be =< x[-1]
    """
    from scipy import interpolate
    f = interpolate.interp1d(x, y) # default kind=‘linear’
    xx = np.linspace(x[0], x[-1], num) # sampled equidistant
    if Tnew:
        xx = np.linspace(x[0], x[0]+Tnew, num) # sampled equidistant & to new T
    yy = f(xx)
    
    return xx, yy

def col2num(col):
    """ convert column in excel to number.
    Input = 'K'; output = 11
    """
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

def rmse(predictions, targets):
    """ root mean squared error
    returns a single number that represents, on average, the distance between 
    every value of list1 to it's corresponding element value of list2
    """
    return np.sqrt(((predictions - targets) ** 2).mean())

def write_errors_excel(dir, errors_periods, rmse_val_periods, col=0):
    """ write errors column wise
    """
    # wb = openpyxl.Workbook()
    dest_filename = 'motion_pattern_error_out.xlsx'
    wb = openpyxl.load_workbook(os.path.join(dir, dest_filename))
    ws = wb.active
    errors_profile = np.mean(np.vstack(errors_periods), axis=0) # mean of periods in cam signal
    mean_abs_error_profile = np.mean(abs(errors_profile))
    rmse_profile = np.mean(rmse_val_periods)
    ws.cell(row=1, column=1).value = 'rmse_profile'
    ws.cell(row=2, column=1).value = 'mean_abs_error_profile'
    ws.cell(row=1, column=2+col).value = rmse_profile
    ws.cell(row=2, column=2+col).value = mean_abs_error_profile
    for i, error in enumerate(errors_profile):
        ws.cell(row=i+5, column=2+col).value = abs(error) # store abs value of error    
    # save excel
    wb.save(os.path.join(dir, dest_filename))


if __name__ == '__main__':
    
    from stentseg.utils.datahandling import select_dir
    
    # preset dirs
    dirsave =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
    exceldir = select_dir(r'C:\Users\Maaike\Dropbox\UTdrive\LSPEAS\Analysis\Validation robot', 
                  r'D:\Profiles\koenradesma\Dropbox\UTdrive\LSPEAS\Analysis\Validation robot')
    workbookCam = '20160215 GRAFIEKEN van camera systeem uit matlab in excel.xlsx'
    workbookAlg = '20160624 DATA Toshiba.xlsx'
    sheetProfile = 'ZB1'
    colSt = 'P'
    n_samplepoints = 11
    visf2 = False
    ylim = 0.45
    xlim = 7 # 3.5
    
    # read camera data
    time_cam_all, pos_cam_all = readCameraExcel(exceldir, workbookCam, sheetProfile, colSt)
    
    # get local minima as starting points for camera periods
    peakmax, peakmin = peakdet(pos_cam_all, 0.05)
    # peakmin = peakmin[::2]# only for B5 and B6 with extra gauss peak
    tt = [time_cam_all[int(peak)] for peak in peakmin[:,0]] # get time in s
    T = (tt[-1]-tt[0])/(len(peakmin)-1) # period of signal
    
    # read algorithm data
    pp = readAnalysisExcel(exceldir, workbookAlg, sheetProfile) # reads 8x10x3
    
    for i_point in range(len(pp)):
        pz = pp[i_point][:,2] # for a point, z-axis
        pz = np.append(pz, pz[0]) 
        time_pp = np.linspace(0,T,11) # scale phases to time domain
        time_pp, pz = resample(time_pp,pz, num=n_samplepoints)
        
        f1 = plt.figure(figsize=(18,5.25))
        ax0 = f1.add_subplot(111)
        ax0.plot(time_cam_all, pos_cam_all, 'r.-', alpha=0.5, label='camera reference')
        
        # get errors after resampling signals
        errors_periods = []
        rmse_val_periods = []
        time_cam_periods = []
        pos_cam_periods = []
        for peak in peakmin[:-1,0]: # we do not include last peak
            if visf2:
                f2, (ax1, ax2) = plt.subplots(2, 1, sharex=False, figsize=(17,11))
            rmse_val = 10000
            for i in range(-3,3): # analyse for 6 cam start points
                istart = int(peak)+i
                tc, pc = time_cam_all[istart:], pos_cam_all[istart:]
                time_cam, pos_cam = getCameraPeriod(tc, pc, T)
                
                # downsample camera to 10 positions
                time_cam_s, pos_cam_s = resample(time_cam,pos_cam, num=11)
                # down sample camera to be same length as algorithm period
                time_cam_sT, pos_cam_sT = resample(time_cam,pos_cam, num=n_samplepoints,Tnew=T)
                pos_offset = min(pos_cam_sT) # not always zero as min value
                pos_cam_sT = np.asarray(pos_cam_sT)- pos_offset # so that positions cam have value 0
                
                # error by subtracting camera from found locations by algorithm
                errors = pz - pos_cam_sT
                
                # root mean squared error
                rmse_val_new = rmse(pos_cam_sT, pz)
                # print('rms error for lag', i, 'is:', str(rmse_val_new))
                rmse_val = min(rmse_val, rmse_val_new) # keep smallest, better overlay with algorithm
                if rmse_val == rmse_val_new:
                    pos_cam_period = pos_cam_sT + pos_offset
                    time_cam_period = time_cam_sT
                    errors_period = errors
                    best_i = i
                
                # vis
                if visf2:
                    time_cam_sT = np.array(time_cam_sT)-time_cam_sT[0] # start at To=0, equal to time_pp
                    ax1.plot(time_cam_sT, pos_cam_sT, 's-', label='camera sampled scaled')
                    ax2.plot(time_pp,errors, 'o--', label='error (alg-cam) (mm)')
            print('period was read from index: ', best_i)
            print('--------------')
            
            # store signal with smallest rmse for each peak/period
            pos_cam_periods.append(pos_cam_period)
            time_cam_periods.append(time_cam_period)
            rmse_val_periods.append(rmse_val)
            errors_periods.append(errors_period)
            
            # vis
            # ax0.plot(time_cam_period, pos_cam_period, linewidth=3, alpha=0.5)
            ax0.plot(time_cam_period, pos_cam_period, 'rs', alpha=0.5)
            if peak == peakmin[-2,0]: # plot legend ones, not looped 
                ax0.plot(time_pp+time_cam_period[0], pz+min(pos_cam_period),'o:',
                    color='k',linewidth=3,alpha=0.5,label='algorithm')
            else:
                ax0.plot(time_pp+time_cam_period[0], pz+min(pos_cam_period),'o:',
                    color='k',linewidth=3,alpha=0.5)
            if visf2:
                ax1.plot(time_pp,pz, 'o:', label='algorithm') # plot algorithm
                ax1.legend()
                ax1.set_xlabel('time (s)')
                ax1.set_ylabel('position (mm)')
                ax2.legend()
                ax2.set_xlabel('time (s)')
                ax2.set_ylabel('error (mm)')
                ax2.axhline(y=0, color='k')
            
        ax0.legend(loc='best')
        ax0.spines["top"].set_visible(False)  
        ax0.spines["right"].set_visible(False)
        ax0.get_xaxis().tick_bottom()  
        ax0.get_yaxis().tick_left()
        ax0.set_xlabel('time (s)',fontsize=16)
        ax0.set_ylabel('position (mm)',fontsize=16)
        for label in (ax0.get_xticklabels() + ax0.get_yticklabels()):
            label.set_fontsize(15)
        plt.xlim(0.0,xlim) # 0.2,xlim
        plt.ylim(0,ylim)
        
        # calc errors
        rmse_profile = np.mean(rmse_val_periods)
        mean_abs_error_periods = [np.mean(abs(e)) for e in errors_periods]
        mean_abs_error_profile = np.mean(mean_abs_error_periods)
        print('rmse of profile=', rmse_profile)
        print('mean abs error of profile=', mean_abs_error_profile)
        
        ## visualize
        # https://github.com/olgabot/prettyplotlib/wiki/Examples-with-code#fill_between-area-between-two-lines%22
        # colormap = brewer2mpl.get_map('YlGnBu', 'sequential', 5).mpl_colormap
        
        # vis signal camera with peaks
        f3 = plt.figure()  
        ax3 = f3.add_subplot(111)
        ax3.plot(time_cam_all, pos_cam_all, 'r.-', alpha=0.5, label='camera')
        t = [time_cam_all[int(peak)] for peak in peakmin[:,0]] # get time in s
        ax3.scatter(t, np.array(peakmin)[:,1], color='green')
        ax3.plot(time_pp,pz, 'o:', label='algorithm')
        ax3.set_xlabel('time (s)')
        ax3.set_ylabel('position (mm)')
        ax3.legend()
        
        ## write excel
        if True:
            dir =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
            write_errors_excel(dir, errors_periods, rmse_val_periods, col=i_point)
            print('******************************')
    
    # save f1
    f1.savefig(os.path.join(dirsave, 'patterncamalg.pdf'), papertype='a0', dpi=300)