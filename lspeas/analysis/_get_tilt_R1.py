""" LSPEAS: Script to analyze and plot tilt of ring with respect to center lumen line
excelsheet1: dist_peaks_valleys_cll.xlsx
excelsheet2: LSPEAS_pulsatility_expansion_avgreg_subp_v15.6.xlsx
Author: M.A. Koenrades. Created August 2018.
__|      |__
__        __
  |_ _ _ |
  | \)   |
  |  \   |
  |   \  |
"""

from lspeas.analysis.utils_analysis import readRingExcel, _initaxis, cols2num
from stentseg.utils import PointSet
import openpyxl
from stentseg.utils.datahandling import select_dir
import sys, os
import math
import numpy as np
import matplotlib.pyplot as plt
import itertools
from matplotlib import gridspec
from lspeas.utils import normality_statistics

exceldir = select_dir(r'C:\Users\Maaike\SURFdrive\UTdrive\LSPEAS\Analysis', 
                r'D:\Profiles\koenradesma\SURFdrive\UTdrive\LSPEAS\Analysis')
workbook_stent = 'LSPEAS_pulsatility_expansion_avgreg_subp_v2.1.xlsx'
workbook_renal_cll = 'Peak valley displacement\\dist_peaks_valleys_cll.xlsx' # data with distances over CLL
# workbook_renal_cll = 'Peak valley displacement\\dist_peaks_valleys_cll_interobserver_JS_fillnans.xlsx'
dirsaveIm =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')


def read_dist_to_renal(exceldir, workbook_renal_cll, ctcode, obs='obsMK'):
    """Read distance from peaks and valleys of R1 to lowest renal 
    (along CLL by Terarecon) for all patients
    return lists
    """
    wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_renal_cll), data_only=True)
    sheet = wb.get_sheet_by_name(obs) # patient order as neck dilation sheet! ...21,25,15,22
    
    colsStart = ['D','I','N','S','X'] # timepoints D tot 24M
    colsStart = cols2num(colsStart)
    
    if ctcode == 'D':
        col = colsStart[0]
    elif ctcode == '1M':
        col = colsStart[1]
    elif ctcode == '6M':
        col = colsStart[2]
    elif ctcode == '12M':
        col = colsStart[3]
    elif ctcode == '24M':
        col = colsStart[4]
    else:
        print('ctcode not known')
        ValueError
    
    rowstart = 7
    numpts = 15
    
    VR = sheet.columns[col][rowstart:rowstart+numpts] # valley right
    PA = sheet.columns[col+1][rowstart:rowstart+numpts] # peak anterior
    VL = sheet.columns[col+2][rowstart:rowstart+numpts]
    PP = sheet.columns[col+3][rowstart:rowstart+numpts]
    
    VR = [obj.value for obj in VR] # list with distance per patient
    PA = [obj.value for obj in PA]
    VL = [obj.value for obj in VL]
    PP = [obj.value for obj in PP]
    
    # patient order as neck dilation sheet! ...21,25,15,22
    return VR, PA, VL, PP # lists with distance per patient for ctcode provided
    

def read_distance_peaks_and_valleys(exceldir, workbook_stent, rowsmm=[143,157]):
    """Read ring distance between peaks and ring distance between valleys 
    of R1 from excel LSPEAS_pulsatility_expansion....xlsx
    """
    wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
    # wbvars = openpyxl.load_workbook(os.path.join(exceldir, workbook_vars), data_only=True)
    sheet = wb.get_sheet_by_name('Deployment') # read from this summery sheet
    # patient order is ... 21,22,19,25 
    
    # define cols and rows
    colsStart = ['E', 'S'] # R1 discharge -> peaks and valleys
    colsStart = cols2num(colsStart)
    rowStart = rowsmm[0]
    rowEnd = rowsmm[1]
        
    xlabels = ['D', '1M', '6M', '12M', '24M']
    
    # get arrays with distances peaks and valleys all patients
    PPVVarray_mm = []
    for i in range(len(xlabels)):
        tPP = sheet.columns[colsStart[0]+i][rowStart:rowEnd+1] 
        tPP = [obj.value for obj in tPP]
        tVV = sheet.columns[colsStart[1]+i][rowStart:rowEnd+1] 
        tVV = [obj.value for obj in tVV]
        PPVVarray_mm.append([tPP, tVV])
    
    PPVVdistances = { 'D':PPVVarray_mm[0], # D with PP and VV for all 15 patients (rowsmm)
            '1M':PPVVarray_mm[1],
            '6M':PPVVarray_mm[2],
            '12M':PPVVarray_mm[3],
            '24M':PPVVarray_mm[4]
        }
    
    # patient order is ... 21,22,19,25
    return PPVVdistances # dict with list per scan with dists per patient
   
    
def tilt_peaks_valleys_centerline(PPVVdistances, ctcode, VR, PA, VL, PP):
    """Estimate tilt of ring by distances to renal of both peaks or both valleys
    and by ring diameter distance. 
    Sin(theta) = (dist_peak1 - dist_peak2) / d_peaks
    dist_peak1 = PA; dist_peak2 = PP; d_peaks = PPVVdistances[ctcode][0]
    """
    # distance between peaks and between valleys
    d_peaks = PPVVdistances[ctcode][0] # list with dists per patient
    d_valleys = PPVVdistances[ctcode][1]
    
    # diff in distance to renal for peaks vs valleys; replace 24m missing values with nan
    diffDistToRenal_peaks = [abs(PA[i]-PP[i]) if not isinstance(PA[i], str) else 
                            np.nan for i, el in enumerate(PA)]
    diffDistToRenal_valleys = [abs(VR[i]-VL[i]) if not isinstance(VR[i], str) else 
                            np.nan for i, el in enumerate(VR)]
    
    # reorder to have same patient order
    diffDistToRenal_peaks = [diffDistToRenal_peaks[i] for i in patient_order(type=23)]
    
    diffDistToRenal_valleys = [diffDistToRenal_valleys[i] for i in patient_order(type=23)]
    
    d_peaks = [d_peaks[i] for i in patient_order(type=13)]
    d_valleys = [d_valleys[i] for i in patient_order(type=13)]
    
    #to array
    diffDistToRenal_peaks = np.asarray(diffDistToRenal_peaks, dtype=np.float)
    diffDistToRenal_valleys = np.asarray(diffDistToRenal_valleys, dtype=np.float)
    d_peaks = np.asarray(d_peaks, dtype=np.float)
    d_valleys = np.asarray(d_valleys, dtype=np.float)
    
    # estimate angle with line perpendicular to center lumen line
    tiltAnglePeaks = [math.degrees(math.asin(diffDistToRenal_peaks[i]/d_peaks[i])) 
                        for i in range(len(d_peaks))] 
    tiltAngleValleys = [math.degrees(math.asin(diffDistToRenal_valleys[i]/d_valleys[i]))
                        for i in range(len(d_peaks))]
    
    return np.vstack(tiltAnglePeaks), np.vstack(tiltAngleValleys)


def patient_order(type=23):
    """ Get index order to rearrange list from the excelsheets
    """
    #type1:ring deployment sheet pt order
    IDorder_ring = [1,2,3,5,8,9,11,15,17,18,20,21,22,19,25]
        
    #type2:renal distance sheet pt order
    IDorder_renal = [1,2,3,5,8,9,11,17,18,19,20,21,25,15,22]
        
    #type3: plot order conform ring evolution paper; chronologic
    IDorder_plot = [1,2,3,5,8,9,11,15,17,18,19,20,21,22,25]
    
    if type == 13: # from 1 to 3
        # [1,2,3].index(2) # => 1
        neworder = [IDorder_ring.index(el) for el in IDorder_plot]
                
    if type == 23: # from 2 to 3
        neworder = [IDorder_renal.index(el) for el in IDorder_plot]
    
    elif type == 32:
        neworder = [IDorder_plot.index(el) for el in IDorder_renal]
                    
    return neworder
        
    
def plot_tilt_lines(exceldir, workbook_stent, workbook_renal_cll, obs='obsMK',
                    ylim=[0,50], saveFig=False):
    """ Plot tilt over time for all patients
    uses the functions above to get data for each time point
    """
    ctcodes=['D', '1M', '6M', '12M', '24M']
    
    # data ring dimension
    PPVVdistances = read_distance_peaks_and_valleys(exceldir, workbook_stent)
    
    tiltPeaks_all = [] # all scans
    tiltValleys_all = []
    for ctcode in ctcodes:
        try:
            # data centerline renal
            VR, PA, VL, PP = read_dist_to_renal(exceldir, workbook_renal_cll, ctcode, obs=obs)
            # tilt
            tiltPeaks, tiltValleys = tilt_peaks_valleys_centerline(PPVVdistances, 
                                        ctcode, VR, PA, VL, PP)
        except TypeError: # when excel cells are Nonetype (empty/not scored)
            tiltPeaks = np.empty((15, 1,)) * np.nan
            tiltValleys = np.empty((15, 1,)) * np.nan
        tiltPeaks_all.append(tiltPeaks)
        tiltValleys_all.append(tiltValleys)
    
    # init figure
    f1 = plt.figure(num=1, figsize=(11.7, 5)) # 11.6,4.6 or 5
    xlabels = ['D', '1M', '6M', '12M', '24M']
    xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
    
    # init axis
    gs = gridspec.GridSpec(1, 2, width_ratios=[1, 1.3]) # 6.4/5.2
    ax1 = plt.subplot(gs[0])
    plt.xticks(xrange, xlabels)
    ax2 = plt.subplot(gs[1])
    plt.xticks(xrange, xlabels)
    
    # Set the font name and size for axis tick labels
    fontName = "Arial" # most similar to Helvetica (not available), which Matlab uses
    ax1.set_ylabel('Tilt peaks ($^\circ$)', fontsize=15, fontname=fontName)
    ax2.set_ylabel('Tilt valleys ($^\circ$)', fontsize=15, fontname=fontName)
    ax1.set_ylim(ylim)
    ax2.set_ylim(ylim)
    ax1.set_yticks(np.arange(0,ylim[1], 5)) # steps of 5
    ax2.set_yticks(np.arange(0,ylim[1], 5))
    ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
    ax2.set_xlim([0.8, len(xlabels)+1.4]) # longer for legend
    
    # lines and colors; 12-class Paired (ring evolution paper)
    colors = itertools.cycle(['#a6cee3','#1f78b4','#b2df8a','#33a02c',
    '#fb9a99','#e31a1c','#fdbf6f','#ff7f00','#cab2d6','#6a3d9a','#ffff99','#b15928'])
    # lStyles = ['-', '--']
    markers = ['D', 'o', '^', 's', '*'] # for device size
    lw = 1
    alpha = 1
    ls = '-'
    ls2 = '--'
    
    IDlegend =['01', '02','03', '05', '08', '09','11','15','17',	
                '18', '19', '20', '21', '22', '25']
    ptdevicesizes = [25.5, 30.5, 30.5, 28, 28, 30.5, 28, 28, 32, 30.5, 30.5, 34,
                        30.5, 28, 34]
    
    # to array
    tiltPeaks_all = np.stack(tiltPeaks_all, axis=1) # 15x(5x1) array # np.stack from 1.10.0
    tiltValleys_all = np.stack(tiltValleys_all, axis=1)
    
    tiltPeaks_mean = np.nanmean(tiltPeaks_all, axis=0)
    tiltValleys_mean = np.nanmean(tiltValleys_all, axis=0)
    
    tiltPeaks_median = np.nanmedian(tiltPeaks_all, axis=0)
    tiltValleys_median = np.nanmedian(tiltValleys_all, axis=0)
    
    # plot per patient the data of all CT scans
    for i, tiltPeak in enumerate(tiltPeaks_all):
        
        patient = IDlegend[i]
        devicesize = ptdevicesizes[i]
        color = next(colors)
        if devicesize == 25.5:
            marker = markers[0]
            olb = 'OLB25' 
        elif devicesize == 28:
            marker = markers[1]
            olb = 'OLB28' 
        elif devicesize == 30.5:
            marker = markers[2]
            olb = 'OLB30'
        elif devicesize == 32:
            marker = markers[3]
            olb = 'OLB32'
        else:
            marker = markers[4]
            olb = 'OLB34'
        
        # plot tilt
        ax1.plot(xrange, tiltPeaks_all[i], ls=ls, lw=lw, marker=marker, color=color, 
        label='%s:%s' % (patient[-2:], olb), alpha=alpha)
        ax2.plot(xrange, tiltValleys_all[i], ls=ls, lw=lw, marker=marker, color=color, 
        label='%s:%s' % (patient[-2:], olb), alpha=alpha)
    # plot mean (median?)
    ax1.plot(xrange, tiltPeaks_median, ls=ls2, lw=2, marker='p', color='k', 
    label='Median', alpha=alpha)
    ax2.plot(xrange, tiltValleys_median, ls=ls2, lw=2, marker='p', color='k', 
    label='Median', alpha=alpha)
    
    # Set the font name and for axis tick labels
    for tick in ax1.get_xticklabels():
        tick.set_fontname(fontName)
    for tick in ax2.get_xticklabels():
        tick.set_fontname(fontName)
    for tick in ax1.get_yticklabels():
        tick.set_fontname(fontName)
    for tick in ax2.get_yticklabels():
        tick.set_fontname(fontName)
    
    ax2.legend(loc='upper right', fontsize=10, numpoints=1)
    _initaxis([ax1, ax2], axsize=14) # sets also xtick ytick fontsize
    
    if saveFig:
        plt.savefig(os.path.join(dirsaveIm, 
        'plot_pp_vv_tilt.png'), papertype='a0', dpi=600)
    
    return f1, tiltPeaks_all, tiltValleys_all


if __name__ == '__main__':
    
    obs = 'obsMK'
    # obs = 'obs2'
    
    if False:
        ctcode = '24M'
        # data ring
        PPVVdistances = read_distance_peaks_and_valleys(exceldir, workbook_stent)
        # data centerline renal
        VR, PA, VL, PP = read_dist_to_renal(exceldir, workbook_renal_cll, ctcode, obs=obs)
        # tilt calculation
        tiltPeaks, tiltValleys = tilt_peaks_valleys_centerline(PPVVdistances, 
            ctcode, VR, PA, VL, PP) # in chronologic order of pt id to plot
    
    # plot tilt all CT's
    f1, tiltPeaks_all, tiltValleys_all = plot_tilt_lines(exceldir, workbook_stent, 
                    workbook_renal_cll, obs=obs,
                    ylim=[-0.3,46], saveFig=True)
    
    #normality check
    # normality_check(tiltPeaks_all[:,0,:]) # per ctcode
    
    # Get median
    if False:
        # pt 21 separate
        tiltPeaks_all[12,:,:] = np.nan
        tiltValleys_all[12,:,:] = np.nan
    # peaks
    mediantiltp = np.nanmedian(tiltPeaks_all, axis=0)
    q1p = np.nanpercentile(tiltPeaks_all, 25, axis=0)
    q3p = np.nanpercentile(tiltPeaks_all, 75, axis=0)
    mintiltp = np.nanmin(tiltPeaks_all, axis=0)
    maxtiltp = np.nanmax(tiltPeaks_all, axis=0)
    # valleys
    mediantiltv = np.nanmedian(tiltValleys_all, axis=0)
    q1v = np.nanpercentile(tiltValleys_all, 25, axis=0)
    q3v = np.nanpercentile(tiltValleys_all, 75, axis=0)
    mintiltv = np.nanmin(tiltValleys_all, axis=0)
    maxtiltv = np.nanmax(tiltValleys_all, axis=0)
    # for paper from spss; different values for quartiles
    
    # get change from discharge
    tiltPeaks_all_change = [tiltpeaks - tiltpeaks[0] for tiltpeaks in tiltPeaks_all]
    tiltValleys_all_change = [tiltvalleys - tiltvalleys[0] for tiltvalleys in tiltValleys_all]
    
    # save to mat var to copy to spss
    import scipy.io
    if True:
        # order as renal distance sheet order
        tiltPeaks_all = [tiltPeaks_all[i] for i in patient_order(type=32)]
        tiltValleys_all = [tiltValleys_all[i] for i in patient_order(type=32)]
        tiltPeaks_all_change = [tiltPeaks_all_change[i] for i in patient_order(type=32)]
        tiltValleys_all_change = [tiltValleys_all_change[i] for i in patient_order(type=32)]
        
        storemat = os.path.join(dirsaveIm, 'tilt_peaks_valleys_from_python.mat')
        storevar = dict()
        storevar['tiltPeaks_all'] = tiltPeaks_all
        storevar['tiltValleys_all'] = tiltValleys_all
        storevar['tiltPeaks_all_change'] = tiltPeaks_all_change
        storevar['tiltValleys_all_change'] = tiltValleys_all_change
        
        storevar['workbook_renal_cll'] = workbook_renal_cll
        storevar['obs'] = obs
        storevar['workbook_stent'] = workbook_stent
        scipy.io.savemat(storemat,storevar)
        print('')
        print('tilt peaks valleys and change was stored as.mat to {}'.format(storemat))
        print('')
        
    
    
    