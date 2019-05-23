""" LSPEAS RING DYNAMICS: Script to collect ring dynamics data from
automated analysis in ring_dynamics.py (excel sheet per patient)
Copyright 2019, Maaike A. Koenrades
"""

import openpyxl
from openpyxl.utils import column_index_from_string
from stentseg.utils.datahandling import select_dir
import sys, os
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np 
import itertools
from matplotlib import gridspec
import visvis as vv
import scipy
from scipy import io
from lspeas.analysis.utils_analysis import _initaxis

class ExcelAnalysisRingDynamics():
    """ Create graphs from excel file per patient created by ring_dynamics.py
    """
    
    exceldir = select_dir(r'C:\Users\Maaike\SURFdrive\UTdrive\LSPEAS\Analysis\Ring motion\ringdynamics', 
                    r'D:\Profiles\koenradesma\SURFdrive\UTdrive\LSPEAS\Analysis\Ring motion\ringdynamics')
    dirsaveIm =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
    
    
    def __init__(self):
        self.exceldir =  ExcelAnalysisRingDynamics.exceldir
        self.dirsaveIm = ExcelAnalysisRingDynamics.dirsaveIm
        self.patients =['LSPEAS_001', 
                        'LSPEAS_002',	
                        'LSPEAS_003', 
                        'LSPEAS_005',	
                        'LSPEAS_008', 
                        'LSPEAS_009',	
                        'LSPEAS_011', 
                        'LSPEAS_015',	
                        'LSPEAS_017',	
                        'LSPEAS_018',
                        'LSPEAS_019', 
                        'LSPEAS_020', 
                        'LSPEAS_021', 
                        'LSPEAS_022', 
                        'LSPEAS_025', 
                        ]
        
        self.fontsize1 = 16 
        self.fontsize2 = 16 # ylabel
        self.fontsize4 = 10 # legend contents
        
    def get_curvature_change(self, patients=None, analysis=['R1','Q1R1','Q2R1','Q3R1','Q4R1'], 
                curvetype='pointchange', storemat=False):
        """ Read curvature change for the ring and/or ring quadrant
        analysis: R1 and/or Q1R1 etc for ring or part of ring
        curvetype: 'pointchange'        --> curvature change per point
                   'peakcurvature'      --> max diff between peak angle over phases
                   'meancurvature'      --> max diff between mean curvature over phases
        """
        
        if patients == None:
            patients = self.patients
        
        ctcodes = ['discharge', '1month', '6months', '12months', '24months']
        
        # read workbooks per patient (this might take a while)
        for a in analysis:
            # init to collect over time all pts
            curvechangeOverTime_pts = [] # cm-1; 5 values for each patient
            curvechangePOverTime_pts = [] # %
            curvechangeRelLocOverTime_pts = [] # %
            avgcurvechangeOverTime_pts = [] # cm-1
            avgcurvechangePOverTime_pts = [] # %
            avgcurveMinOverTime_pts = []
            avgcurveMaxOverTime_pts = []
            
            for patient in patients:
                # init to collect parameter for all timepoints for this patient
                curvechangeOverTime = [] # cm-1
                curvechangePOverTime = [] # %
                curvechangeRelLocOverTime = [] # %
                avgcurvechangeOverTime = []
                avgcurvechangePOverTime = []
                avgcurveMinOverTime = []
                avgcurveMaxOverTime = []
                 
                for i in range(len(ctcodes)):
                    filename = '{}_ringdynamics_{}.xlsx'.format(patient, ctcodes[i])
                    workbook_stent = os.path.join(self.exceldir, filename)
                    # read workbook
                    try:
                        wb = openpyxl.load_workbook(workbook_stent, data_only=True)
                    except FileNotFoundError: # handle missing scans
                        if curvetype == 'pointchange':
                            # collect
                            curvechangeOverTime.append(np.nan)
                            curvechangePOverTime.append(np.nan)
                            curvechangeRelLocOverTime.append(np.nan)
                            avgcurvechangeOverTime.append(np.nan)
                            avgcurvechangePOverTime.append(np.nan)
                            avgcurveMinOverTime.append(np.nan)
                            avgcurveMaxOverTime.append(np.nan)
                            
                        continue # no scan, next
                    
                    # get sheet
                    sheetname = 'Curvature{}'.format(a)
                    sheet = wb.get_sheet_by_name(sheetname)
                    # set row for type of curvature analysis
                    if curvetype == 'pointchange':
                        rowstart = 5 # as in excel; max change of a point
                        rowstart2 = 9 # as in excel; rel location
                        rowstart3 = 13 # as in excel; mean curvature change all points
                        rowstart4 = 26 # as in excel; mean curvature min and max value
                        colstart = 1 # 1 = B
                        # read change
                        curvechange = sheet.rows[rowstart-1][colstart].value
                        curvechangeP = sheet.rows[rowstart-1][colstart+1].value
                        curvechangeRelLoc = sheet.rows[rowstart2-1][colstart].value 
                        avgcurvechange = sheet.rows[rowstart3-1][colstart].value
                        avgcurvechangeP = sheet.rows[rowstart3-1][colstart+1].value
                        avgcurvemin = sheet.rows[rowstart4-1][colstart].value
                        avgcurvemax = sheet.rows[rowstart4-1][colstart+1].value
                        # collect
                        curvechangeOverTime.append(curvechange)
                        curvechangePOverTime.append(curvechangeP)
                        curvechangeRelLocOverTime.append(curvechangeRelLoc)
                        avgcurvechangeOverTime.append(avgcurvechange)
                        avgcurvechangePOverTime.append(avgcurvechangeP)
                        avgcurveMinOverTime.append(avgcurvemin)
                        avgcurveMaxOverTime.append(avgcurvemax)
                        
                # collect for pts
                curvechangeOverTime_pts.append(np.asarray(curvechangeOverTime))
                curvechangePOverTime_pts.append(np.asarray(curvechangePOverTime))
                curvechangeRelLocOverTime_pts.append(np.asarray(curvechangeRelLocOverTime))        
                avgcurvechangeOverTime_pts.append(np.asarray(avgcurvechangeOverTime))
                avgcurvechangePOverTime_pts.append(np.asarray(avgcurvechangePOverTime))
                avgcurveMinOverTime_pts.append(np.asarray(avgcurveMinOverTime))
                avgcurveMaxOverTime_pts.append(np.asarray(avgcurveMaxOverTime))
                
            # store in self
            self.curvechangeOverTime_pts = np.asarray(curvechangeOverTime_pts)
            self.curvechangePOverTime_pts = np.asarray(curvechangePOverTime_pts)
            self.curvechangeRelLocOverTime_pts = np.asarray(curvechangeRelLocOverTime_pts)
            self.avgcurvechangeOverTime_pts = np.asarray(avgcurvechangeOverTime_pts)
            self.avgcurvechangePOverTime_pts = np.asarray(avgcurvechangePOverTime_pts)
            self.avgcurveMinOverTime_pts = np.asarray(avgcurveMinOverTime_pts)
            self.avgcurveMaxOverTime_pts = np.asarray(avgcurveMaxOverTime_pts)
            
            print_stats_var_over_time_(self.curvechangeOverTime_pts, varname='Curvature{}_maxchange_pts'.format(a), dec=3)
            print_stats_var_over_time_(self.curvechangePOverTime_pts, varname='Curvature{}_maxchangeP_pts'.format(a))
            print_stats_var_over_time_(self.curvechangeRelLocOverTime_pts, varname='Curvature{}_maxchangeRelLoc_pts'.format(a))
            print_stats_var_over_time_(self.avgcurvechangeOverTime_pts, varname='Curvature{}_meanchange_pts'.format(a), dec=3)
            print_stats_var_over_time_(self.avgcurvechangePOverTime_pts, varname='Curvature{}_meanchangeP_pts'.format(a) )
            print_stats_var_over_time_(self.avgcurveMinOverTime_pts, varname='Curvature{}_meanMin_pts'.format(a), dec=3 )
            print_stats_var_over_time_(self.avgcurveMaxOverTime_pts, varname='Curvature{}_meanMax_pts'.format(a), dec=3 )
            
            # Store to .mat per analysis
            if storemat:
                if curvetype == 'pointchange':
                    self.store_var_to_mat(np.asarray(curvechangeOverTime_pts), varname='Curvature{}_maxchange_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(curvechangePOverTime_pts), varname='Curvature{}_maxchangeP_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(curvechangeRelLocOverTime_pts), varname='Curvature{}_maxchangeRelLoc_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(avgcurvechangeOverTime_pts), varname='Curvature{}_meanchange_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(avgcurvechangePOverTime_pts), varname='Curvature{}_meanchangeP_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(avgcurveMinOverTime_pts), varname='Curvature{}_meanMin_pts'.format(a) )
                    self.store_var_to_mat(np.asarray(avgcurveMaxOverTime_pts), varname='Curvature{}_meanMax_pts'.format(a) )
                    
    
    def get_displacement_cycle(self, patients=None, analysis=['R1','Q1R1','Q2R1','Q3R1','Q4R1'], 
                storemat=False):
        """ Read displacement during the cycle for the ring and/or ring quadrant
        analysis: R1 and/or Q1R1 etc for ring or part of ring
        Obtain mean displacement in 3d and x,y,z directions
        """
        
        if patients == None:
            patients = self.patients
        
        ctcodes = ['discharge', '1month', '6months', '12months', '24months']
        
        # read workbooks per patient (this might take a while)
        for a in analysis:
            # init to collect over time all pts
            displxyzOverTime_pts = [] #  xyz mm; 5 values for each patient
            curvechangePOverTime_pts = [] # x mm
            curvechangeRelLocOverTime_pts = [] # y mm
            avgcurvechangeOverTime_pts = [] # z mm
            for patient in patients:
                # init to collect parameter for all timepoints for this patient
                displxyzOverTime = [] 
                displxOverTime = [] 
                displyOverTime = [] 
                displzOverTime = []
                for i in range(len(ctcodes)):
                    filename = '{}_ringdynamics_{}.xlsx'.format(patient, ctcodes[i])
                    workbook_stent = os.path.join(self.exceldir, filename)
                    # read workbook
                    try:
                        wb = openpyxl.load_workbook(workbook_stent, data_only=True)
                    except FileNotFoundError: # handle missing scans
                        # collect
                        displxyzOverTime.append(np.nan)
                        displxOverTime.append(np.nan)
                        displyOverTime.append(np.nan)
                        displzOverTime.append(np.nan)
                        continue # no scan, next
                    
                    # get sheet
                    sheetname = 'Motion{}'.format(a)
                    sheet = wb.get_sheet_by_name(sheetname)
                    # set row for 3d,x,y,z
                    rowstart = 6 # as in excel; xyz/3d
                    rowstart2 = 7 # as in excel; x
                    rowstart3 = 8 # as in excel; y
                    rowstart4 = 9 # as in excel; z
                    
                    colstart = 1 # 1 = B
                    # read change
                    displxyz = sheet.rows[rowstart-1][colstart].value
                    displx = sheet.rows[rowstart2-1][colstart].value 
                    disply = sheet.rows[rowstart3-1][colstart].value
                    displz = sheet.rows[rowstart4-1][colstart].value 
                    # collect
                    displxyzOverTime.append(displxyz)
                    displxOverTime.append(displx)
                    displyOverTime.append(disply)
                    displzOverTime.append(displz)
                    
                # collect for pts
                displxyzOverTime_pts.append(np.asarray(displxyzOverTime))
                curvechangePOverTime_pts.append(np.asarray(displxOverTime))
                curvechangeRelLocOverTime_pts.append(np.asarray(displyOverTime))        
                avgcurvechangeOverTime_pts.append(np.asarray(displzOverTime))
            
            self.displxyzOverTime_pts = np.asarray(displxyzOverTime_pts)
            self.displxOverTime_pts = np.asarray(curvechangePOverTime_pts)
            self.displyOverTime_pts = np.asarray(curvechangeRelLocOverTime_pts)
            self.displzOverTime_pts = np.asarray(avgcurvechangeOverTime_pts)
            
            print_stats_var_over_time_(self.displxOverTime_pts, varname='Displacement{}_x_pts'.format(a))
            print_stats_var_over_time_(self.displyOverTime_pts, varname='Displacement{}_y_pts'.format(a))
            print_stats_var_over_time_(self.displzOverTime_pts, varname='Displacement{}_z_pts'.format(a))
            print_stats_var_over_time_(self.displxyzOverTime_pts, varname='Displacement{}_3d_pts'.format(a))
                
            # Store to .mat per analysis
            if storemat:
                self.store_var_to_mat(np.asarray(displxyzOverTime_pts), varname='Displacement{}_3d_pts'.format(a) )
                self.store_var_to_mat(np.asarray(curvechangePOverTime_pts), varname='Displacement{}_x_pts'.format(a) )
                self.store_var_to_mat(np.asarray(curvechangeRelLocOverTime_pts), varname='Displacement{}_y_pts'.format(a) )
                self.store_var_to_mat(np.asarray(avgcurvechangeOverTime_pts), varname='Displacement{}_z_pts'.format(a) )
    
    
    def get_distance_change_between_rings(self, patients=None, analysis=['R1R2','Q1R1R2','Q2R1R2','Q3R1R2','Q4R1R2'], 
                storemat=False):
        """ Read distance change during the cycle for the ring and/or ring quadrant
        analysis: R1 and/or Q1R1 etc for ring or part of ring
        Obtain mean and max distance change
        """
        
        if patients == None:
            patients = self.patients
        
        ctcodes = ['discharge', '1month', '6months', '12months', '24months']
        
        # read workbooks per patient (this might take a while)
        for a in analysis:
            # init to collect over time all pts
            meanOverTime_pts = [] # mm; 5 values for each patient
            maxOverTime_pts = [] # mm
            for patient in patients:
                # init to collect parameter for all timepoints for this patient
                meanOverTime = [] 
                maxOverTime = [] 
                for i in range(len(ctcodes)):
                    filename = '{}_ringdynamics_{}.xlsx'.format(patient, ctcodes[i])
                    workbook_stent = os.path.join(self.exceldir, filename)
                    # read workbook
                    try:
                        wb = openpyxl.load_workbook(workbook_stent, data_only=True)
                    except FileNotFoundError: # handle missing scans
                        # collect
                        meanOverTime.append(np.nan)
                        maxOverTime.append(np.nan)
                        continue # no scan, next
                    
                    # get sheet
                    sheetname = 'Distances{}'.format(a)
                    sheet = wb.get_sheet_by_name(sheetname)
                    # set row for 3d,x,y,z
                    rowstart = 16 # as in excel; mean
                    rowstart2 = 6 # as in excel; max
                    colstart = 1 # 1 = B
                    # read change
                    distmean = sheet.rows[rowstart-1][colstart].value
                    distmax = sheet.rows[rowstart2-1][colstart].value 
                    # collect
                    meanOverTime.append(distmean)
                    maxOverTime.append(distmax)
                    
                # collect for pts
                meanOverTime_pts.append(np.asarray(meanOverTime))
                maxOverTime_pts.append(np.asarray(maxOverTime))
            
            self.meanOverTime_pts = np.asarray(meanOverTime_pts)
            self.maxOverTime_pts = np.asarray(maxOverTime_pts)
            
            print_stats_var_over_time_(self.meanOverTime_pts, varname='Distances{}_mean_pts'.format(a) )
            print_stats_var_over_time_(self.maxOverTime_pts, varname='Distances{}_max_pts'.format(a)   )
                
            # Store to .mat per analysis
            if storemat:
                self.store_var_to_mat(np.asarray(meanOverTime_pts), varname='Distances{}_mean_pts'.format(a) )
                self.store_var_to_mat(np.asarray(maxOverTime_pts), varname='Distances{}_max_pts'.format(a) )
                
    
    def store_var_to_mat(self, variable, varname=None, storematdir=None):
        """ Save as .mat to easy copy to spss
        """
        if storematdir is None:
            storematdir = os.path.join(self.dirsaveIm, 'python_to_mat_output_ringdynamics')
        if varname is None:
            varname = 'variable_from_python'
        storemat = os.path.join(storematdir, varname+'.mat')
        
        storevar = dict()
        # check if variable has multiple vars in list
        if isinstance(variable, list):
            for i, var in enumerate(variable):
                name = varname+'{}'.format(i)
                storevar[name] = var
        else:
            storevar[varname] = variable
        
        storevar['workbooks_ringdynamics'] = '_ringdynamics_'
        storevar['patients'] = self.patients
        
        io.savemat(storemat,storevar)
        print('')
        print('variable {} was stored as.mat to {}'.format(varname, storemat))
                    
    
    def plot_curvature_during_cycle(self, patients=['LSPEAS_001'], analysis=[''], 
                ctcode = 'discharge', ring='R1', ylim=[0, 2], ylimRel=[-0.1,0.1], saveFig=False):
        """ plot change in curvature during the cardiac cycle for point with max curvature
        change at ring or ring quadrants at a certain time point (CT scan)
        * ctcode = discharge, 1month, 6 months, ... 24months
        * analysis = '' for ring or Q1, Q2, Q3, Q4
        """
        self.curveAll = {}
        self.curveRelAll = {}
        
        # init figure
        self.f1 = plt.figure(figsize=(11.5, 4.2)) 
        xlabels = ['0', '10', '20', '30', '40', '50', '60', '70','80','90']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        fontsize1 = self.fontsize1
        fontsize2 = self.fontsize2
        
        # init axis
        factor = 1.33 # 1.36
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, factor]) # plot right wider
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        ax1.set_ylabel('Curvature (cm$^{-1}$)', fontsize=fontsize2) # absolute curvature values
        ax1.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax2.set_ylabel('Relative curvature change (cm$^{-1}$)', fontsize=fontsize2) # relative dist from avgreg
        ax2.set_ylim(ylimRel)
        ax2.set_xlim([0.8, len(xlabels)*factor+0.2]) # xlim margins 0.2; # longer for legend
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        # plot init
        lw = 1
        alpha = 0.9
        ls = '-'
        markers = ['o','D', '^', 'o', 's']
        colorsdirections = [           # from phantom paper in vivo plots
                                'k',
                                '#d73027', # 1 red
                                '#fc8d59', # orange
                                '#91bfdb', # blue
                                '#4575b4' # 5
                                ]
        Qnames = ['Full ring', 'QP','QA','QL','QR']
        Qs =['','Q1','Q3','Q2','Q4']
        
        for patient in patients:
            for a in analysis:
                # read workbook
                filename = '{}_ringdynamics_{}.xlsx'.format(patient, ctcode)
                workbook_stent = os.path.join(self.exceldir, filename)
                wb = openpyxl.load_workbook(workbook_stent, data_only=True)
                sheetname = 'Curvature{}{}'.format(a, ring)
                sheet = wb.get_sheet_by_name(sheetname)
                # read cuvrature of point with max change
                rowstart = 12 # as in excel; during cycle
                colstart = 1 # 1 = B
                # read change
                curvatures = sheet.rows[rowstart-1][colstart:colstart+10]
                curvatures = [obj.value for obj in curvatures]
                # convert to array
                curvatures = np.asarray(curvatures)
                # get curvature at mid heart cycle
                avgcurvature = np.nanmean(curvatures) 
                # relative curvatures from avgreg
                curveRel = curvatures - avgcurvature
                
                self.curveAll['{}_{}'.format(patient, a)] = curvatures
                self.curveRelAll['{}_{}'.format(patient, a)] = curveRel
                
                # plotting
                # color1 = color_per_patient(patient)
                color1 = colorsdirections[Qs.index(a)]
                marker = markers[Qs.index(a)]
                label = Qnames[Qs.index(a)]
                ax1.plot(xrange, curvatures, ls=ls, lw=lw, marker=marker, color=color1, 
                        label=label, alpha=alpha)
                ax2.plot(xrange, curveRel, ls=ls, lw=lw, marker=marker, color=color1, 
                        label=label, alpha=alpha)
                
        ax2.legend(loc='upper right', fontsize=self.fontsize4, numpoints=1, title='Legend:')
        _initaxis([ax1,ax2], axsize=fontsize1)
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_curvature_over_cycle_{}.png'.format(analysis)), papertype='a0', dpi=600)
    
def print_stats_var_over_time_(varOverTime_pts, varname='varOverTime_pts', dec=1, showvar=True):
    """ To print stats for paper
    varOverTime_pts is an array of size 15x5 (pts x timepoints)
    dec to set number of decimals
    """
    timepoints = ['D', '1M', '6M', '12M', '24M']
    if showvar:
        print(varOverTime_pts)
    print(varname)
    for time in range(len(timepoints)): # 5 time points
        print(timepoints[time]+':')
        if dec == 3:
            print('Average±std, min, max: {:.3f} ± {:.3f} ({:.3f}-{:.3f})'.format(
                                                np.nanmean(varOverTime_pts[:,time]),
                                                np.nanstd(varOverTime_pts[:,time]),
                                                np.nanmin(varOverTime_pts[:,time]),
                                                np.nanmax(varOverTime_pts[:,time])
                                                ))
            #change compared to discharge
            print('CHANGE: Average±std, min, max: {:.3f} ± {:.3f} ({:.3f}-{:.3f})'.format(
                                                np.nanmean((varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanstd( (varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanmin( (varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanmax( (varOverTime_pts[:,time]-varOverTime_pts[:,0]))
                                                )) 
        else:
            print('Average±std, min, max: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.nanmean(varOverTime_pts[:,time]),
                                                np.nanstd(varOverTime_pts[:,time]),
                                                np.nanmin(varOverTime_pts[:,time]),
                                                np.nanmax(varOverTime_pts[:,time])
                                                ))
            #change compared to discharge
            print('CHANGE: Average±std, min, max: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.nanmean((varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanstd( (varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanmin( (varOverTime_pts[:,time]-varOverTime_pts[:,0])),
                                                np.nanmax( (varOverTime_pts[:,time]-varOverTime_pts[:,0]))
                                                )) 
            
        print()


if __name__ == '__main__':
    
    foo = ExcelAnalysisRingDynamics()
    
    ## Ring motion manuscript 
    if False:
        # Get and store curvature change rings and quadrants post left ant right
        print('       R1      ')
        foo.get_curvature_change(patients=None, analysis=['R1','Q1R1','Q2R1','Q3R1','Q4R1'], 
                curvetype='pointchange', storemat=True)
        print('       R2      ')
        foo.get_curvature_change(patients=None, analysis=['R2','Q1R2','Q2R2','Q3R2','Q4R2'], 
                curvetype='pointchange', storemat=True)
        
        
        # Get and store displacement rings and quadrants
        print('       R1      ')
        foo.get_displacement_cycle(patients=None, analysis=['R1','Q1R1','Q2R1','Q3R1','Q4R1'], 
                storemat=False)
        print('       R2      ')
        foo.get_displacement_cycle(patients=None, analysis=['R2','Q1R2','Q2R2','Q3R2','Q4R2'], 
                storemat=False)
        
        # Get and store distance between rings
        foo.get_distance_change_between_rings(patients=None, analysis=['R1R2','Q1R1R2','Q2R1R2','Q3R1R2','Q4R1R2'], 
                    storemat=False)
    
    # Plot curvature during the cycle (same as distance change)
    if True:
        foo.plot_curvature_during_cycle(patients=['LSPEAS_002'], analysis=['Q1','Q3','Q2','Q4'], 
            ctcode = '12months', ring='R1', ylim=[0, 2], ylimRel=[-0.1,0.1], saveFig=True)
    