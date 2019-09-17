""" Module to analyze vessel pulsatility during the heart cycle in ecg-gated CT
radius change - area change - volume change

"""

import os
import sys
import time

import openpyxl
import pirt
import numpy as np
import visvis as vv

from stentseg.utils.datahandling import select_dir, loadvol, loadmodel, loadmesh
from stentseg.stentdirect.stentgraph import create_mesh
from stentseg.motion.vis import create_mesh_with_abs_displacement
from stentseg.utils import PointSet, fitting

from lspeas.utils.vis import showModelsStatic
from lspeas.utils.deforminfo import DeformInfo
from lspeas.utils.curvature import measure_curvature
from lspeas.utils import meshlib

assert openpyxl.__version__ < "2.4", "Do pip install openpyxl==2.3.5"


#todo: move function to lspeas utils?
def load_excel_centerline(basedirCenterline, vol, ptcode, ctcode, filename=None):
    """ Load centerline data from excel
    Centerline exported from Terarecon; X Y Z coordinates in colums
    """
    if filename is None:
        filename = '{}_{}_centerline.xlsx'.format(ptcode,ctcode)
    excel_location = os.path.join(basedirCenterline,ptcode)
    #read sheet
    try:
        wb = openpyxl.load_workbook(os.path.join(excel_location,filename),read_only=True)
    except FileNotFoundError:
        wb = openpyxl.load_workbook(os.path.join(basedirCenterline,filename),read_only=True)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0]) # data on first sheet
    colStart = 2 # col C
    rowStart = 1 # row 2 excel
    coordx = sheet.columns[colStart][rowStart:]
    coordy = sheet.columns[colStart+1][rowStart:]
    coordz = sheet.columns[colStart+2][rowStart:]
    #convert to values
    coordx = [obj.value for obj in coordx]
    coordy = [obj.value for obj in coordy]
    coordz = [obj.value for obj in coordz]
    #from list to array
    centerlineX = np.asarray(coordx, dtype=np.float32)
    centerlineY = np.asarray(coordy, dtype=np.float32)
    centerlineZ = np.asarray(coordz, dtype=np.float32)
    centerlineZ = np.flip(centerlineZ, axis=0) # z of volume is also flipped

    # convert centerline coordinates to world coordinates (mm)
    origin = vol1.origin # z,y,x
    sampling = vol1.sampling # z,y,x

    centerlineX = centerlineX*sampling[2] +origin[2]
    centerlineY = centerlineY*sampling[1] +origin[1]
    centerlineZ = (centerlineZ-0.5*vol1.shape[0])*sampling[0] + origin[0]

    return centerlineX, centerlineY, centerlineZ


# Select the ssdf basedir
basedir = select_dir(
    os.getenv('LSPEAS_BASEDIR', ''),
    r'D:\LSPEAS\LSPEAS_ssdf',
    r'F:\LSPEAS_ssdf_backup',
    r'F:\LSPEAS_ssdf_BACKUP')

basedirMesh = select_dir(
    r'D:\Profiles\koenradesma\SURFdrive\UTdrive\MedDataMimics\LSPEAS_Mimics',
    r'C:\Users\Maaike\SURFdrive\UTdrive\MedDataMimics\LSPEAS_Mimics',
    r"C:\stack\data\lspeas\vaatwand")

basedirCenterline = select_dir(
    r"C:\stack\data\lspeas\vaatwand",
    r'D:\Profiles\koenradesma\SURFdrive\UTdrive\LSPEAS_centerlines_terarecon',
    r'C:\Users\Maaike\SURFdrive\UTdrive\LSPEAS_centerlines_terarecon',
    r"C:\stack\data\lspeas\vaatwand")

# Select dataset
ptcode = 'LSPEAS_003'
ctcode1 = 'discharge'
cropname = 'ring'
modelname = 'modelavgreg'
cropvol = 'stent'

drawModelLines = False  # True or False
drawRingMesh, ringMeshDisplacement = True, False
meshColor = [(1,1,0,1)]
removeStent = True # for iso visualization
dimensions = 'xyz'
showAxis = False
showVol  = 'ISO'  # MIP or ISO or 2D or None
showvol2D = False
drawVessel = True

clim = (0,2500)
clim2D = -200,500 # MPR
clim2 = (0,2)
isoTh = 180 # 250


## Load data

# Load CT image data for reference, and deform data to measure motion
try:
    # If we run this script without restart, we can re-use volume and deforms
    vol1
    deforms
except NameError:
    vol1 = loadvol(basedir, ptcode, ctcode1, cropvol, 'avgreg').vol
    s_deforms = loadvol(basedir, ptcode, ctcode1, cropvol, 'deforms')
    deforms = [s_deforms[key] for key in dir(s_deforms) if key.startswith('deform')]
    deforms = [pirt.DeformationFieldBackward(*fields) for fields in deforms]

# Load vessel mesh (mimics)
# We make sure that it is a mesh without faces, which makes our sampling easier
try:
    ppvessel
except NameError:
    # Load mesh with visvis, then put in our meshlib.Mesh() and let it ensure that
    # the mesh is closed, check the winding, etc. so that we can cut it with planes,
    # and reliably calculate volume.
    filename = '{}_{}_neck.stl'.format(ptcode,ctcode1)
    vesselMesh = loadmesh(basedirMesh, ptcode[-3:], filename) #inverts Z
    vv.processing.unwindFaces(vesselMesh)
    vesselMesh = meshlib.Mesh(vesselMesh._vertices)
    vesselMesh.ensure_closed()
    ppvessel = PointSet(vesselMesh.get_flat_vertices())  # Must be flat!


# Load ring model
try:
    modelmesh1
except NameError:
    s1 = loadmodel(basedir, ptcode, ctcode1, cropname, modelname)
    if drawRingMesh:
        if not ringMeshDisplacement:
            modelmesh1 = create_mesh(s1.model, 0.7)  # Param is thickness
        else:
            modelmesh1 = create_mesh_with_abs_displacement(s1.model, radius = 0.7, dim=dimensions)

# Load vessel centerline (excel terarecon) (is very fast)
centerline = PointSet(np.column_stack(
    load_excel_centerline(basedirCenterline, vol1, ptcode, ctcode1, filename=None)))


## Setup visualization

# Show ctvolume, vessel mesh, ring model - this uses figure 1 and clears it
axes1, cbars = showModelsStatic(ptcode, ctcode1, [vol1], [s1], [modelmesh1],
              [vv.BaseMesh(*vesselMesh.get_vertices_and_faces())],
              showVol, clim, isoTh, clim2, clim2D, drawRingMesh,
              ringMeshDisplacement, drawModelLines, showvol2D, showAxis,
              drawVessel, vesselType=1,
              climEditor=True, removeStent=removeStent, meshColor=meshColor)
axes1 = axes1[0]
axes1.position = 0, 0, 0.6, 1

# Show or hide the volume (showing is nice, but also slows things down)
tex3d = axes1.wobjects[1]
tex3d.visible = False

# VesselMeshes
vesselVisMesh1 = axes1.wobjects[4]
vesselVisMesh1.cullFaces = "front"  # Show the back
# vesselVisMesh2 = vv.Mesh(axes1, *vesselMesh.get_vertices_and_faces())
vesselVisMesh2 = vv.Mesh(axes1, np.zeros((6, 3), np.float32), np.zeros((3, 3), np.int32))
vesselVisMesh2.cullFaces = "back"
vesselVisMesh2.faceColor = "red"


# Show the centerline
vv.plot(centerline, ms='.', ls='', mw=8, mc='b', alpha=0.5)

# Initialize 2D view
axes2 = vv.Axes(vv.gcf())
axes2.position = 0.65, 0.05, 0.3, 0.4
axes2.daspectAuto = False
axes2.camera = '2d'
axes2.axis.showGrid = True
axes2.axis.axisColor = 'k'

# Initialize axes to put widgets and labels in
container = vv.Wibject(vv.gcf())
container.position = 0.65, 0.5, 0.3, 0.5

# Create labels to show measurements
labelpool = []
for i in range(16):
    label = vv.Label(container)
    label.fontSize = 11
    label.position = 10, 100 + 25 * i, -20, 25
    labelpool.append(label)

# Initialize sliders and buttons
slider_ref = vv.Slider(container, fullRange=(1, len(centerline)-2), value=10)
slider_ves = vv.Slider(container, fullRange=(1, len(centerline)-2), value=10)
button_go = vv.PushButton(container, "Take all measurements (incl. volume)")
slider_ref.position = 10, 5, -20, 25
slider_ves.position = 10, 40, -20, 25
button_go.position = 10, 70, -20, 25
button_go.bgcolor = slider_ref.bgcolor = slider_ves.bgcolor = 0.8, 0.8, 1.0

# Initialize line objects for showing the plane orthogonal to centerline
slider_ref.line_plane = vv.plot([], [], [], axes=axes1, ls='-', lw=3, lc='w', alpha = 0.9)
slider_ves.line_plane = vv.plot([], [], [], axes=axes1, ls='-', lw=3, lc='y', alpha = 0.9)
# Initialize line objects for showing selected points close to that plane
slider_ref.line_3d = vv.plot([], [], [], axes=axes1, ms='.', ls='', mw=8, mc='w', alpha = 0.9)
slider_ves.line_3d = vv.plot([], [], [], axes=axes1, ms='.', ls='', mw=8, mc='y', alpha = 0.9)

# Initialize line objects for showing selected points and ellipse in 2D
line_2d = vv.plot([], [], axes=axes2,  ms='.', ls='', mw=8, mc='y')
line_ellipse1 = vv.plot([], [], axes=axes2,  ms='', ls='-', lw=2, lc='b')
line_ellipse2 = vv.plot([], [], axes=axes2,  ms='', ls='+', lw=2, lc='b')


## Functions to update visualization and do measurements


def get_plane_points_from_centerline_index(i):
    """ Get a set of points that lie on the plane orthogonal to the centerline
    at the given index. The points are such that they can be drawn as a line for
    visualization purposes. The plane equation can be obtained via a plane-fit.
    """

    if True:
        # Cubic fit of the centerline

        i = max(1.1, min(i, centerline.shape[0] - 2.11))
        # Sample center point and two points right below/above, using
        # "cardinal" interpolating (C1-continuous), or "basic" approximating (C2-continious).
        pp = []
        for j in [i - 0.1, i, i + 0.1]:
            index = int(j)
            t = j - index
            coefs = pirt.interp.get_cubic_spline_coefs(t, "basic")
            samples = centerline[index - 1], centerline[index], centerline[index + 1], centerline[index + 2]
            pp.append(samples[0] * coefs[0] + samples[1] * coefs[1] + samples[2] * coefs[2] + samples[3] * coefs[3])
        # Get center point and vector pointing down the centerline
        p = pp[1]
        vec1 = (pp[2] - pp[1]).normalize()

    else:
        # Linear fit of the centerline

        i = max(0, min(i, centerline.shape[0] - 2))
        index = int(i)
        t = i - index
        # Sample two points of interest
        pa, pb = centerline[index], centerline[index + 1]
        # Get center point and vector pointing down the centerline
        p = t * pb + (1 - t) * pa
        vec1 = (pb - pa).normalize()

    # Get two orthogonal vectors that define the plane that is orthogonal
    # to the above vector. We can use an arbitrary vector to get the first,
    # but there is a tiiiiiny chance that it is equal to vec1 so that the
    # normal collapese.
    vec2 = vec1.cross([0, 1, 0])
    if vec2.norm() == 0:
        vec2 = vec1.cross((1, 0, 0))
    vec3 = vec1.cross(vec2)

    # Sample some point on the plane and get the plane's equation
    pp = PointSet(3)
    radius = 6
    pp.append(p)
    for t in np.linspace(0, 2 * np.pi, 12):
        pp.append(p + np.sin(t) * radius * vec2 + np.cos(t) * radius * vec3)
    return pp


def get_vessel_points_from_plane_points(pp):
    """ Select points from the vessel points that are very close to the plane
    defined by the given plane points. Returns a 2D and a 3D point set.
    """
    abcd = fitting.fit_plane(pp)

    # Get 2d and 3d coordinates of points that lie (almost) on the plane
    # pp2 = fitting.project_to_plane(ppvessel, abcd)
    # pp3 = fitting.project_from_plane(pp2, abcd)
    signed_distances = fitting.signed_distance_to_plane(ppvessel, abcd)
    distances = np.abs(signed_distances)

    # Select points to consider. This is just to reduce the search space somewhat.
    selection = np.where(distances < 5)[0]

    # We assume that each tree points in ppvessel makes up a triangle (face)
    # We make sure of that when we load the mesh.
    # Select first index of each face (every 3 vertices is 1 face), and remove duplicates
    selection_faces = set(3 * (selection // 3))

    # Now iterate over the faces (triangles), and check each edge. If the two
    # points are on different sides of the plane, then we interpolate on the
    # edge to get the exact spot where the edge intersects the plane.
    sampled_pp3 = PointSet(3)
    visited_edges = set()
    for fi in selection_faces:  # for each face index
        for edge in [(fi + 0, fi + 1), (fi + 0, fi + 2), (fi + 1, fi + 2)]:
            if signed_distances[edge[0]] * signed_distances[edge[1]] < 0:
                if edge not in visited_edges:
                    visited_edges.add(edge)
                    d1, d2 = distances[edge[0]], distances[edge[1]]
                    w1, w2 = d2 / (d1 + d2), d1 / (d1 + d2)
                    p = w1 * ppvessel[edge[0]] + w2 * ppvessel[edge[1]]
                    sampled_pp3.append(p)

    return fitting.project_to_plane(sampled_pp3, abcd), sampled_pp3


def get_distance_along_centerline():
    """ Get the distance along the centerline between the two reference points,
    (using linear interpolation at the ends).
    """
    i1 = slider_ref.value
    i2 = slider_ves.value

    index1 = int(np.ceil(i1))
    index2 = int(np.floor(i2))
    t1 = i1 - index1  # -1 < t1 <= 0
    t2 = i2 - index2  # 0 <= t2 < 1

    dist = 0
    dist += -t1 * (centerline[index1] - centerline[index1 - 1]).norm()
    dist += +t2 * (centerline[index2] - centerline[index2 + 1]).norm()

    for index in range(index1, index2):
        dist += (centerline[index + 1] - centerline[index]).norm()

    return float(dist)


def triangle_area(p1, p2, p3):
    """ Calcualate triangle area based on its three vertices.
    """
    # Use Heron's formula to calculate a triangle's area
    # https://www.mathsisfun.com/geometry/herons-formula.html
    a = p1.distance(p2)
    b = p2.distance(p3)
    c = p3.distance(p1)
    s = (a + b + c) / 2
    return (s * (s - a) * (s - b) * (s - c)) ** 0.5


def deform_points_2d(pp2, plane):
    """ Given a 2D pointset (and the plane that they are on),
    return a list with the deformed versions of that pointset.
    """
    pp3 = fitting.project_from_plane(pp2, plane)
    deformed = []
    for phase in range(len(deforms)):
        deform = deforms[phase]
        dx = deform.get_field_in_points(pp3, 0) #todo: shouldn't this be z=0 y=1 x=2! see dynamic.py; adapt in all functions?!
        dy = deform.get_field_in_points(pp3, 1)
        dz = deform.get_field_in_points(pp3, 2)
        deform_vectors = PointSet(np.stack([dx, dy, dz], 1))
        pp3_deformed = pp3 + deform_vectors
        deformed.append(fitting.project_to_plane(pp3_deformed, plane))
    return deformed


def measure_centerline_strain():
    """ Measure the centerline strain.
    """
    i1 = slider_ref.value
    i2 = slider_ves.value

    # Get the centerline section of interest
    index1 = int(np.ceil(i1))
    index2 = int(np.floor(i2))
    section = centerline[index1:index2 + 1]

    # get this section of the centerline for each phase
    sections = []
    for phase in range(len(deforms)):
        deform = deforms[phase]
        dx = deform.get_field_in_points(section, 0)
        dy = deform.get_field_in_points(section, 1)
        dz = deform.get_field_in_points(section, 2)
        deform_vectors = PointSet(np.stack([dx, dy, dz], 1))
        sections.append(section + deform_vectors)

    # Measure the strain of the full section, by measuring the total length in each phase.
    lengths = []
    for phase in range(len(deforms)):
        section = sections[phase]
        length = sum(float(section[i].distance(section[i + 1]))
                     for i in range(len(section) - 1))
        lengths.append(length)

    if min(lengths) == 0:
        return 0
    else:
        # Strain as delta-length divided by initial length
        return (max(lengths) - min(lengths)) / min(lengths)
        # ... or as what Wikipedia calls "stretch ratio">
        # return max(lengths) / min(lengths)


def take_measurements(measure_volume_change):
    """ This gets called when the slider is releases. We take measurements and
    update the corresponding texts and visualizations.
    """

    # Get points that form the contour of the vessel in 2D
    pp = get_plane_points_from_centerline_index(slider_ves.value)
    pp2, pp3 = get_vessel_points_from_plane_points(pp)
    plane = pp2.plane

    # Collect measurements in a dict. That way we can process it in one step at the end
    measurements = {}

    # Store slider positions, so we can reproduce this measurement later
    measurements["centerline indices"] = slider_ref.value, slider_ves.value

    # Early exit?
    if len(pp2) == 0:
        line_2d.SetPoints(pp2)
        line_ellipse1.SetPoints(pp2)
        line_ellipse2.SetPoints(pp2)
        vesselVisMesh2.SetFaces(np.zeros((3, 3), np.int32))
        vesselVisMesh2.SetNormals(None)
        process_measurements(measurements)
        return

    # Measure length of selected part of the centerline and the strain in that section
    measurements["centerline distance"] = get_distance_along_centerline()
    measurements["centerline strain"] = measure_centerline_strain()

    # Measure centerline curvature
    curvature_mean, curvature_max, curvature_max_pos, curvature_max_change = measure_curvature(centerline, deforms)
    measurements["curvature mean"] = DeformInfo(curvature_mean)
    measurements["curvature max"] = DeformInfo(curvature_max)
    measurements["curvature max pos"] = DeformInfo(curvature_max_pos)
    measurements["curvature max change"] = curvature_max_change

    # Get ellipse and its center point
    ellipse = fitting.fit_ellipse(pp2)
    p0 = PointSet([ellipse[0], ellipse[1]])

    # Sample ellipse to calculate its area
    pp_ellipse = fitting.sample_ellipse(ellipse, 256)  # results in N + 1 points
    area = 0
    for i in range(len(pp_ellipse)-1):
        area += triangle_area(p0, pp_ellipse[i], pp_ellipse[i + 1])
    # measurements["reference area"] = float(area)
    # Do a quick check to be sure that this triangle-approximation is close enough
    assert abs(area - fitting.area(ellipse)) < 2, "area mismatch"  # mm2  typically ~ 0.1 mm2

    # Measure ellipse area (and how it changes)
    measurements["ellipse area"] = DeformInfo(unit="mm2")
    for pp_ellipse_def in deform_points_2d(pp_ellipse, plane):
        area = 0
        for i in range(len(pp_ellipse_def)-1):
            area += triangle_area(p0, pp_ellipse_def[i], pp_ellipse_def[i + 1])
        measurements["ellipse area"].append(area)

    # # Measure expansion of ellipse in 256 locations?
    # # Measure distances from center to ellipse edge. We first get the distances
    # # in each face, for each point. Then we aggregate these distances to
    # # expansion measures. So in the end we have 256 expansion measures.
    # distances_per_point = [[] for i in range(len(pp_ellipse))]
    # for pp_ellipse_def in deform_points_2d(pp_ellipse, plane):
    #     # todo: Should distance be measured to p0 or to p0 in that phase?
    #     for i, d in enumerate(pp_ellipse_def.distance(p0)):
    #         distances_per_point[i].append(float(d))
    # distances_per_point = distances_per_point[:-1]  # Because pp_ellipse[-1] == pp_ellipse[0]
    # #
    # measurements["expansions"] = DeformInfo()  # 256 values, not 10
    # for i in range(len(distances_per_point)):
    #     distances = distances_per_point[i]
    #     measurements["expansions"].append((max(distances) - min(distances)) / min(distances))

    # Measure radii of ellipse major and minor axis (and how it changes)
    pp_ellipse4 = fitting.sample_ellipse(ellipse, 4)  # major, minor, major, minor
    measurements["ellipse expansion major1"] = DeformInfo(unit="mm")
    measurements["ellipse expansion minor1"] = DeformInfo(unit="mm")
    measurements["ellipse expansion major2"] = DeformInfo(unit="mm")
    measurements["ellipse expansion minor2"] = DeformInfo(unit="mm")
    for pp_ellipse4_def in deform_points_2d(pp_ellipse4, plane):
        measurements["ellipse expansion major1"].append(float( pp_ellipse4_def[0].distance(p0) ))
        measurements["ellipse expansion minor1"].append(float( pp_ellipse4_def[1].distance(p0) ))
        measurements["ellipse expansion major2"].append(float( pp_ellipse4_def[2].distance(p0) ))
        measurements["ellipse expansion minor2"].append(float( pp_ellipse4_def[3].distance(p0) ))

    # Measure how the volume changes - THIS BIT IS COMPUTATIONALLY EXPENSIVE
    submesh = meshlib.Mesh(np.zeros((3, 3)))
    if measure_volume_change:
        # Update the submesh
        plane1 = fitting.fit_plane(get_plane_points_from_centerline_index(slider_ref.value))
        plane2 = fitting.fit_plane(get_plane_points_from_centerline_index(slider_ves.value))
        plane2 = [-x for x in plane2]  # flip the plane upside doen
        submesh = vesselMesh.cut_plane(plane1).cut_plane(plane2)
        # Measure its motion
        measurements["volume"] = DeformInfo(unit="mm3")
        submesh._ori_vertices = submesh._vertices.copy()
        for phase in range(len(deforms)):
            deform = deforms[phase]
            submesh._vertices = submesh._ori_vertices.copy()
            dx = deform.get_field_in_points(submesh._vertices, 0)
            dy = deform.get_field_in_points(submesh._vertices, 1)
            dz = deform.get_field_in_points(submesh._vertices, 2)
            submesh._vertices[:, 0] += dx
            submesh._vertices[:, 1] += dy
            submesh._vertices[:, 2] += dz
            measurements["volume"].append(submesh.volume())

    # Show measurements
    process_measurements(measurements)

    # Update line objects
    line_2d.SetPoints(pp2)
    line_ellipse1.SetPoints(fitting.sample_ellipse(ellipse))
    major_minor = PointSet(2)
    for p in [p0, pp_ellipse4[0], p0, pp_ellipse4[2], p0, pp_ellipse4[1], p0, pp_ellipse4[3]]:
        major_minor.append(p)
    line_ellipse2.SetPoints(major_minor)
    axes2.SetLimits(margin=0.12)

    # Update submesh object
    vertices, faces = submesh.get_vertices_and_faces()
    vesselVisMesh2.SetVertices(vertices)
    vesselVisMesh2.SetFaces(np.zeros((3, 3), np.int32) if len(faces) == 0 else faces)
    vesselVisMesh2.SetNormals(None)


# Global value that will be a dictionary with measurements
mm = {}

def process_measurements(measurements):
    """ Show measurements. Now the results are shown in a label object, but we could do anything here ...
    """
    # Store in global for further processing
    mm.clear()
    mm.update(measurements)

    # Print in shell
    print("Measurements:")
    for key, val in measurements.items():
        val = val.summary if isinstance(val, DeformInfo) else val
        print(key.rjust(16) + ": " + str(val))

    # Show in labels
    index = 0
    for key, val in measurements.items():
        val = val.summary if isinstance(val, DeformInfo) else val
        val = "{:0.4g}".format(val) if isinstance(val, float) else val
        labelpool[index].text = key + ": " + str(val)
        index += 1
    # Clean remaining labels
    for index in range(index, len(labelpool)):
        labelpool[index].text = ""


def on_sliding(e):
    """ When the slider is moved, update the centerline position indicator.
    """
    slider = e.owner
    pp = get_plane_points_from_centerline_index(slider.value)
    slider.line_plane.SetPoints(pp)


def on_sliding_done(e):
    """ When the slider is released, update the whole thing.
    """
    slider = e.owner
    pp = get_plane_points_from_centerline_index(slider.value)
    slider.line_plane.SetPoints(pp)
    take_measurements(False)  # dont do volume


def on_button_press(e):
    """ When the button is pressed, take measurements.
    """
    take_measurements(True)


def set_sliders(value_ref, value_ves):
    """ Set the sliders to a specific position, e.g. to reproduce a measurement.
    """
    slider_ref.value = value_ref
    slider_ves.value = value_ves


# Connect!
slider_ref.eventSliding.Bind(on_sliding)
slider_ves.eventSliding.Bind(on_sliding)
slider_ref.eventSliderChanged.Bind(on_sliding_done)
slider_ves.eventSliderChanged.Bind(on_sliding_done)
button_go.eventMouseDown.Bind(on_button_press)

#todo: measure (change of) curvature of stent rings
#todo: visualize mesh with motion and use colors to represent radius change
