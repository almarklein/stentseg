""" Functionality for stent analysis LSPEAS
Author: Maaike Koenrades
"""

from stentseg.utils import PointSet
import openpyxl
from stentseg.utils.datahandling import select_dir
import sys, os
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np 
import itertools
from matplotlib import gridspec


def point_in_pointcloud_closest_to_p(pp, point):
    """ Find point in PointSet which is closest to a point
    Returns a PointSet with point found in PointSet and point 
    """
    vecs = pp-point
    dists_to_pp = ( (vecs[:,0]**2 + vecs[:,1]**2 + vecs[:,2]**2)**0.5 ).reshape(-1,1)
    pp_index =  list(dists_to_pp).index(dists_to_pp.min() ) # index on path
    pp_point = pp[pp_index]
    p_in_pp_and_point = PointSet(3, dtype='float32')
    [p_in_pp_and_point.append(*p) for p in (pp_point, point)]
    
    return p_in_pp_and_point


def _initaxis(axis, legend=None, xlabel=None, ylabel=None, labelsize=16, 
              axsize=15, legendtitle=None):
    """ Set axis for nice visualization
    axis is list such as [ax] or [ax1, ax2]
    legend = None or provide location 'upper right'
    xlabel = 'time (s)'
    """
    for ax in axis:
        ax.spines["top"].set_visible(False)  
        ax.spines["right"].set_visible(False)
        ax.get_xaxis().tick_bottom()  
        ax.get_yaxis().tick_left()
        if not legend is None:
            if not legendtitle is None:
                ax.legend(loc=legend, title=legendtitle)
            else:
                ax.legend(loc=legend)
        if not xlabel is None:
            ax.set_xlabel(xlabel, fontsize=labelsize)
        if not ylabel is None:
            ax.set_ylabel(ylabel, fontsize=labelsize)
        # set fontsize axis numbers
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(axsize)
    plt.tight_layout() # so that labels are not cut off


def grouped_boxplot_2subgroups(data, group_names=['A', 'B', 'C'], 
            subgroup_names=['Apples', 'Oranges'], ax=None, 
            subgroup_colors=['blue', 'red'], drawMean = True,
            box_width=0.6, box_spacing=1.0, ylim=[0,30], legendloc='upper right'):
    """ Draws a grouped boxplot for two subgroups.
        data: dict as:
            data = { 'A':[np.random.randn(100), np.random.randn(100) + 5],
            'B':[np.random.randn(100)+1, np.random.randn(100) + 9],
            'C':[np.random.randn(100)-3, np.random.randn(100) -5]
            }
        group_names: list of strings
        
        based on http://stackoverflow.com/questions/16592222/matplotlib-group-boxplots
        works with matplotlib 1.3.1, newer error on flyers?
    """ 
    
    from matplotlib.pyplot import plot, show, savefig, xlim, figure, \
                    hold, legend, boxplot, setp, axes
    
    # function for setting the colors of the box plots pairs
    def setBoxColors(bp):
        setp(bp['boxes'][0], color=subgroup_colors[0])
        setp(bp['caps'][0], color=subgroup_colors[0])
        setp(bp['caps'][1], color=subgroup_colors[0])
        setp(bp['whiskers'][0], color=subgroup_colors[0], linestyle='-')
        setp(bp['whiskers'][1], color=subgroup_colors[0], linestyle='-')
        setp(bp['fliers'][0], color=subgroup_colors[0])
        setp(bp['fliers'][1], color=subgroup_colors[0])
        setp(bp['medians'][0], color=subgroup_colors[0])
    
        setp(bp['boxes'][1], color=subgroup_colors[1])
        setp(bp['caps'][2], color=subgroup_colors[1])
        setp(bp['caps'][3], color=subgroup_colors[1])
        setp(bp['whiskers'][2], color=subgroup_colors[1], linestyle='-')
        setp(bp['whiskers'][3], color=subgroup_colors[1], linestyle='-')
        setp(bp['fliers'][2], color=subgroup_colors[1])
        setp(bp['fliers'][3], color=subgroup_colors[1])
        setp(bp['medians'][1], color=subgroup_colors[1])
    
    if ax is None:
        fig = figure()
        ax = fig.add_subplot(111)
    hold(True)
    
    spos = 1
    xtickpositions = []
    for i, group in enumerate(group_names):
        # per boxplot pair
        if not i == 0:
            spos+= 2 + box_spacing # 2 subgroups
        positions = [spos,spos+1]
        datagroup = data[group] # e.g. data['A'] or data['1M'], holds data R1 and R2, 2 lists of n elements
        # remove missing datapoints
        datagroupori = datagroup.copy()
        datagroup[0] = [el for el in datagroup[0] if el is not None]
        datagroup[1] = [el for el in datagroup[1] if el is not None]
        if not datagroup == datagroupori:
            print("warning: some data elements are missing in group {}".format(group)) 
        bp = boxplot(datagroup, positions = positions, widths = box_width)
        setBoxColors(bp)
        if drawMean:
            # draw a black diamond for the mean
            plt.plot(positions, [np.mean(datagroup[0]), np.mean(datagroup[1])], 
            color='k', marker='D', markeredgecolor='k', markersize=2, linestyle='')
        xtickpositions.append(spos+0.5)
        
    # set axes limits and labels
    xlim(0,spos+2)
    ax.set_ylim(ylim)
    ax.set_xticklabels(group_names)
    # ax.set_xticks([1.5, 4.5, 7.5])
    ax.set_xticks(xtickpositions)
    
    # draw temporary red and blue lines and use them to create a legend
    hB, = plot([1,1], linestyle='-', color=subgroup_colors[0])
    hR, = plot([1,1], linestyle='-', color=subgroup_colors[1])
    legend((hB, hR),(subgroup_names[0], subgroup_names[1]), loc=legendloc)
    hB.set_visible(False)
    hR.set_visible(False)
    
    # savefig('boxcompare.png')
    show()


class ExcelAnalysis():
    """ Create graphs from excel data
    """
    
    exceldir = select_dir(r'C:\Users\Maaike\Dropbox\UTdrive\LSPEAS\Analysis', 
                    r'D:\Profiles\koenradesma\Dropbox\UTdrive\LSPEAS\Analysis')
    dirsaveIm =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
    
    def __init__(self):
        self.exceldir =  ExcelAnalysis.exceldir
        self.dirsaveIm = ExcelAnalysis.dirsaveIm
        self.workbook_stent = 'LSPEAS_pulsatility_expansion_avgreg_subp_v15.3.xlsx'
        self.workbook_renal = 'postop_measures_renal_ring.xlsx'
        self.workbook_variables = 'LSPEAS_Variables.xlsx'
        self.workbook_variables_presheet = 'preCTA_bone_align'
        self.sheet_valley = 'valley locations'
        self.patients =['LSPEAS_001', 'LSPEAS_002',	'LSPEAS_003', 'LSPEAS_005',	
                        'LSPEAS_008', 'LSPEAS_009',	'LSPEAS_011', 'LSPEAS_015',	'LSPEAS_017',	
                        'LSPEAS_018', 'LSPEAS_019',	'LSPEAS_020', 'LSPEAS_021',	'LSPEAS_022',
                        'LSPEAS_025', 'LSPEAS_023', 'LSPEAS_024', 'LSPEAS_004']   


    def readRingExcel(self, ptcode, ctcode, ring='R1'):
        """ To read peak and valley locations, R1 or R2, up to 24M
        """
        
        exceldir = self.exceldir
        workbook = self.workbook_stent
        sheet = 'Output_'+ ptcode[-3:]
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook), data_only=True)
        sheet = wb.get_sheet_by_name(sheet)
        
        if ring == 'R1':
            colsStart = [1,5,9,13,17]
        elif ring == 'R2':
            colsStart = [21,25,29,33,37]
        if ctcode == 'discharge':
            R1_ant = sheet.rows[16][colsStart[0]:colsStart[0]+3]
            R1_ant = [obj.value for obj in R1_ant]
            R1_post = sheet.rows[29][colsStart[0]:colsStart[0]+3]
            R1_post = [obj.value for obj in R1_post] 
            R1_left = sheet.rows[53][colsStart[0]:colsStart[0]+3]
            R1_left = [obj.value for obj in R1_left]
            R1_right = sheet.rows[66][colsStart[0]:colsStart[0]+3]
            R1_right = [obj.value for obj in R1_right]
        elif ctcode == '1month':
            R1_ant = sheet.rows[16][colsStart[1]:colsStart[1]+3]
            R1_ant = [obj.value for obj in R1_ant]
            R1_post = sheet.rows[29][colsStart[1]:colsStart[1]+3]
            R1_post = [obj.value for obj in R1_post] 
            R1_left = sheet.rows[53][colsStart[1]:colsStart[1]+3]
            R1_left = [obj.value for obj in R1_left]
            R1_right = sheet.rows[66][colsStart[1]:colsStart[1]+3]
            R1_right = [obj.value for obj in R1_right]
        elif ctcode == '6months':
            R1_ant = sheet.rows[16][colsStart[2]:colsStart[2]+3]
            R1_ant = [obj.value for obj in R1_ant]
            R1_post = sheet.rows[29][colsStart[2]:colsStart[2]+3]
            R1_post = [obj.value for obj in R1_post] 
            R1_left = sheet.rows[53][colsStart[2]:colsStart[2]+3]
            R1_left = [obj.value for obj in R1_left]
            R1_right = sheet.rows[66][colsStart[2]:colsStart[2]+3]
            R1_right = [obj.value for obj in R1_right]
        elif ctcode == '12months':
            R1_ant = sheet.rows[16][colsStart[3]:colsStart[3]+3]
            R1_ant = [obj.value for obj in R1_ant]
            R1_post = sheet.rows[29][colsStart[3]:colsStart[3]+3]
            R1_post = [obj.value for obj in R1_post] 
            R1_left = sheet.rows[53][colsStart[3]:colsStart[3]+3]
            R1_left = [obj.value for obj in R1_left]
            R1_right = sheet.rows[66][colsStart[3]:colsStart[3]+3]
            R1_right = [obj.value for obj in R1_right]
        elif ctcode == '24months':
            R1_ant = sheet.rows[16][colsStart[4]:colsStart[4]+3]
            R1_ant = [obj.value for obj in R1_ant]
            R1_post = sheet.rows[29][colsStart[4]:colsStart[4]+3]
            R1_post = [obj.value for obj in R1_post] 
            R1_left = sheet.rows[53][colsStart[4]:colsStart[4]+3]
            R1_left = [obj.value for obj in R1_left]
            R1_right = sheet.rows[66][colsStart[4]:colsStart[4]+3]
            R1_right = [obj.value for obj in R1_right]
        else:
            print('ctcode not known')
            ValueError
            
        return R1_ant, R1_post, R1_left, R1_right


    def writeRingPatientsExcel(self, ptcodes='anaconda', ctcodes='all'):
        """ Write valley locations from multiple patients
        ptcodes/ctcodes can be a list with ptcode/ctcode strings
        overwrites excel peak valleys locations!
        """
        
        sheetwrite = 'peak valley locations'
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        workbook_renal = self.workbook_renal
        sheetwrite = self.sheet_valley
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_renal), data_only=True)
        sheet_to_write = wb.get_sheet_by_name(sheetwrite)
        
        patients = ['LSPEAS_001', 'LSPEAS_002',	'LSPEAS_003', 'LSPEAS_005',	
                    'LSPEAS_008', 'LSPEAS_009',	'LSPEAS_011', 'LSPEAS_015',	'LSPEAS_017',	
                    'LSPEAS_018', 'LSPEAS_019',	'LSPEAS_020', 'LSPEAS_021',	'LSPEAS_022',
                    'LSPEAS_025']#, 'LSPEAS_023', 'LSPEAS_024', 'LSPEAS_004']
        codes = ['discharge', '1month', '6months', '12months', '24months']
        if ptcodes == 'anaconda':
            ptcodes = patients
        if ctcodes == 'all':
            ctcodes = codes
        rowStart = [8,31] # left, right; for writing
        colStart = [4,8,12,16,20] # for writing ctcodes; 4 = D
        for ctcode in ctcodes:
            col = colStart[ctcodes.index(ctcode)]
            for ptcode in ptcodes:
                R_ant, R_post, R_left, R_right = self.readRingExcel(ptcode, ctcode, ring='R2')
                # check which is left and right (origin is right anterior)
                if not None in (R_left or R_right): # datapoint is missing
                    if R_left[0] > R_right[0]: # x from right to left
                        R_LR = [R_left, R_right]
                    elif R_left[0] == R_right[0]:
                        print("Error for {} {}, ring position in x is the same for left and right".format(ptcode, ctcode)) 
                    else:
                        R_LR = [R_right, R_left]
                else:
                    R_LR = [R_left, R_right]
                    print("None was encountered in ring position for {} {}".format(ptcode, ctcode))
                # write positions
                for i, position in enumerate(R_LR):
                    row = rowStart[i] + patients.index(ptcode)
                    for j, coordinate in enumerate(position):
                        sheet_to_write.cell(row=row, column=col+j).value = coordinate
        wb.save(os.path.join(exceldir, workbook_renal))
        print('worksheet "%s" overwritten in workbook "%s"' % (sheetwrite,workbook_renal) )
    


    def readRenalsExcel(self, sheet, ptcode, ctcode, zPositive=True):
        """ To read renal locations. zPositive True will return z as non-negative
        use sheet to specify sheet in excel for observer
        """
        
        exceldir = self.exceldir
        workbook_renal = self.workbook_renal
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_renal), data_only=True)
        sheet = wb.get_sheet_by_name(sheet)
        
        colsStart = [5,9,13,17,21]
        patients = self.patients
        
        if ctcode == 'discharge':
            col = colsStart[0]
        elif ctcode == '1month':
            col = colsStart[1]
        elif ctcode == '6months':
            col = colsStart[2]
        elif ctcode == '12months':
            col = colsStart[3]
        elif ctcode == '24months':
            col = colsStart[4]
        else:
            print('ctcode not known')
            ValueError
        
        rowstartleft = 7 # left renal
        rowstartright = 30 # right renal
        
        rowleft = rowstartleft + patients.index(ptcode)
        rowright = rowstartright + patients.index(ptcode)
        
        renal_left = sheet.rows[rowleft][col:col+3]
        renal_left = [obj.value for obj in renal_left]
        renal_right = sheet.rows[rowright][col:col+3]
        renal_right = [obj.value for obj in renal_right] 
        
        if zPositive == True:
            for renal in [renal_left, renal_right]:
                if renal[-1] < 0:
                    renal[-1] *= -1
        
        renal_left = PointSet(renal_left)
        renal_right = PointSet(renal_right)
        
        return renal_left, renal_right


    def plot_distance_to_renal(self, sheet, ptcode):
        """ Plot distance renal from peak valleys for 1 patient
        e.g. sheet = 'distances to renal obs1'
        """
        import prettyplotlib as ppl
        from prettyplotlib import brewer2mpl # colormaps, http://colorbrewer2.org/#type=sequential&scheme=YlGnBu&n=5
        from scipy.interpolate import interp1d #, spline, splrep, splev, UnivariateSpline
        
        colormap = brewer2mpl.get_map('YlGnBu', 'sequential', 5).mpl_colormap # ppl...(fig,ax,...,cmap = )
        exceldir = self.exceldir
        workbook_renal = self.workbook_renal
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_renal), data_only=True)
        sheet = wb.get_sheet_by_name(sheet)
        
        patients = self.patients
        
        rowstartleft = 7 # left renal
        rowstartright = 30 # right renal
        
        colsStart = ['D', 'I', 'N', 'S']#, 'X'] # ctcodes
        colsStart = [(openpyxl.cell.column_index_from_string(char)-1) for char in colsStart]
        
        rowleft = rowstartleft + patients.index(ptcode)
        rowright = rowstartright + patients.index(ptcode)
        
        # read distances to renal for every ctcode and 4 peak valleys positions
        zDistances_ctcodes = [] # VR PA VL PP
        for col in colsStart:
            zDistance = sheet.rows[rowleft][col:col+4]
            zDistance = [obj.value for obj in zDistance]
            zDistances_ctcodes.append(zDistance) 
        
        # plot
        f1 = plt.figure(num=1, figsize=(7.6, 5))
        ax1 = f1.add_subplot(111)
        _initaxis([ax1])
        ax1.set_xlabel('location on ring-stent', fontsize=14)
        ax1.set_ylabel('distance to left renal (mm)', fontsize=14)
        plt.ylim(-10,12)
        xlabels = sheet.rows[rowstartleft-1][col:col+4] # VR PA VL PP
        xlabels = [obj.value for obj in xlabels]
        legend = ['discharge', '1month', '6months', '12months']#, '24months'] first adjust excel
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        for i, zDistances in enumerate(zDistances_ctcodes):
            ax1.plot(xrange, zDistances, linestyle='--', marker='o', label=legend[i])
            # # plot smooth with spline interpolation
            # xnew = np.linspace(1,4,25)
            # f2 = interp1d(xrange, zDistances, kind='quadratic')
            # plt.plot(xnew,f2(xnew))
        plt.xticks(xrange, xlabels, fontsize = 14)
        plt.xlim(0.8,len(xlabels)+0.2) # xlim margins 0.2
        ax1.legend(loc='best')
        for label in (ax1.get_xticklabels() + ax1.get_yticklabels()):
            label.set_fontsize(14)


    def plot_pp_vv_deployment(self, ring=1, patients=None, saveFig=True):
        """ Plot multipanel pp and vv deployment residu per patient; 
        show (a)symmetry; ring=1 or 2 for R1 or R2
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        workbook_vars = self.workbook_variables
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        wbvars = openpyxl.load_workbook(os.path.join(exceldir, workbook_vars), data_only=True) 
        
        # init figure
        f1 = plt.figure(num=2, figsize=(17, 12))
        # plt.xlim(18,34)
        # ax1.plot([0,30],[0,30], ls='--', color='dimgrey')
        xlabels = ['D', '1M', '6M', '12M', '24M']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        
        # read data
        colStart = [1, 6] # B, G
        rowStart = 83 # pp
        if patients is None:
            patients = self.patients
        
        for i, patient in enumerate(patients):
            # init axis
            ax1 = f1.add_subplot(5,3,i+1)
            _initaxis([ax1])
            # ax1.set_xlabel('PP distance (mm)', fontsize=14)
            ax1.set_ylabel('RDC (%)', fontsize=14) # ring deployment capacity
            plt.ylim(-10,47)
            
            sheet = wb.get_sheet_by_name(patient)
            
            if ring == 1 or ring == 12:
                # read R1
                ppR1 = sheet.rows[rowStart][colStart[0]:colStart[0]+5] # +4 is read until 12M
                ppR1 = [obj.value for obj in ppR1]
                vvR1 = sheet.rows[rowStart+1][colStart[0]:colStart[0]+5]
                vvR1 = [obj.value for obj in vvR1]
                # read R2
            if ring == 2 or ring == 12:
                ppR2 = sheet.rows[rowStart][colStart[1]:colStart[1]+5]
                ppR2 = [obj.value for obj in ppR2]
                vvR2 = sheet.rows[rowStart+1][colStart[1]:colStart[1]+5]
                vvR2 = [obj.value for obj in vvR2]
            
            # plot R1
            if ring == 1:
                ax1.plot(xrange, ppR1, ls='-', marker='o', color='#ef8a62', label='PP - R1')
                ax1.plot(xrange, vvR1, ls='-', marker='o', color='#67a9cf', label='VV - R1')
            if ring == 2:
                ax1.plot(xrange, ppR2, ls='-', marker='o', color='#ef8a62', label='PP - R2')
                ax1.plot(xrange, vvR2, ls='-', marker='o', color='#67a9cf', label='VV - R2')
            if ring == 12:
                ax1.plot(xrange, ppR1, ls='-', marker='o', color='#ef8a62', label='PP - R1')
                ax1.plot(xrange, vvR1, ls='-', marker='o', color='#67a9cf', label='VV - R1')
                ax1.plot(xrange, ppR2, ls='--', marker='^', color='#ef8a62', label='PP - R2')
                ax1.plot(xrange, vvR2, ls='--', marker='^', color='#67a9cf', label='VV - R2')
                
            
            plt.xticks(xrange, xlabels, fontsize = 14)
            plt.xlim(0.8,len(xlabels)+0.2) # xlim margins 0.2
            ax1.legend(loc='upper right', fontsize=8, title=('%i: ID %s' % (i+1, patient[-3:]) )  )
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_pp_vv_deployment_R{}.png'.format(ring)), papertype='a0', dpi=300)
    
    
    def plot_ring_deployment_lines(self, rowsrd=[76,90], rowsmm=[120,134], 
                ylim=[0, 42], ylim_mm=[18, 34], 
                subgroup_colors=['#D02D2E', 'blue'], saveFig=True):
        """
        plot mean with std and colors to show spread
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        # workbook_vars = self.workbook_variables
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        # wbvars = openpyxl.load_workbook(os.path.join(exceldir, workbook_vars), data_only=True)
        sheet = wb.get_sheet_by_name('Summary')
        
        # init figure
        f1 = plt.figure(num=3, figsize=(11.6, 4.6)) # 4.6
        xlabels = ['D', '1M', '6M', '12M', '24M']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        
        # init axis
        ax1 = f1.add_subplot(1,2,1)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax2 = f1.add_subplot(1,2,2)
        plt.xticks(xrange, xlabels, fontsize = 14)
        
        ax1.set_ylabel('Residual deployment capacity (%)', fontsize=15) # ring deployment capacity
        ax1.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        # plots in mm
        ax2.set_ylabel('Diameter (mm)', fontsize=15) # mean distance pp vv
        ax2.set_ylim(ylim_mm)
        ax2.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        
        # read excel
        colStart = ['E', 'J'] # R1 R2
        colStart = [(openpyxl.cell.column_index_from_string(char)-1) for char in colStart]
        rowStart = rowsrd[0]
        rowEnd = rowsrd[1]
        
        # get arrays with rdc R1 and R2 all patients in rows
        ttRarray_rd = [] # to collect data elements over time
        for i in range(len(xlabels)):
            tR1 = sheet.columns[colStart[0]+i][rowStart:rowEnd+1] 
            tR1 = [obj.value for obj in tR1]
            tR2 = sheet.columns[colStart[1]+i][rowStart:rowEnd+1] 
            tR2 = [obj.value for obj in tR2]
            ttRarray_rd.append([tR1, tR2])
        
        data_rd = { 'D':ttRarray_rd[0], # D met R1 en R2
                '1M':ttRarray_rd[1],
                '6M':ttRarray_rd[2],
                '12M':ttRarray_rd[3],
                '24M':ttRarray_rd[4]
            }
        
        rowStart = rowsmm[0]
        rowEnd = rowsmm[1]
        
        # get arrays with distances R1 and R2 all patients in rows
        ttRarray_mm = []
        for i in range(len(xlabels)):
            tR1 = sheet.columns[colStart[0]+i][rowStart:rowEnd+1] 
            tR1 = [obj.value for obj in tR1]
            tR2 = sheet.columns[colStart[1]+i][rowStart:rowEnd+1] 
            tR2 = [obj.value for obj in tR2]
            ttRarray_mm.append([tR1, tR2])
        
        data_mm = { 'D':ttRarray_mm[0], # D met R1 en R2
                '1M':ttRarray_mm[1],
                '6M':ttRarray_mm[2],
                '12M':ttRarray_mm[3],
                '24M':ttRarray_mm[4]
            }
        
        # get means and std for rdc
        means_rd = [] # n x 2 for n timepoints R1, R2
        stds_rd = []
        mins_rd = []
        maxs_rd = [] 
        for timepoint in xlabels:
            # remove missing datapoints
            datagroup = data_rd[timepoint]
            datagroupori = datagroup.copy()
            datagroup[0] = [el for el in datagroup[0] if el is not None] #R1
            datagroup[1] = [el for el in datagroup[1] if el is not None] #R2
            if not datagroup == datagroupori:
                print("warning: some data elements are missing in group {}".format(timepoint))
            
            data_rd_mean = [np.mean(l) for l in datagroup] # mean of R1 and R2
            data_rd_std = [np.std(l) for l in datagroup]
            data_rd_max = [np.max(l) for l in datagroup]
            data_rd_min = [np.min(l) for l in datagroup]
            means_rd.append(data_rd_mean)
            stds_rd.append(data_rd_std)
            maxs_rd.append(data_rd_max)
            mins_rd.append(data_rd_min)
        
        # get means and std for distance in mm
        means_mm = [] # n x 2 for n timepoints R1, R2
        stds_mm = []
        mins_mm = []
        maxs_mm = [] 
        for timepoint in xlabels:
            # remove missing datapoints
            datagroup = data_mm[timepoint]
            datagroupori = datagroup.copy()
            datagroup[0] = [el for el in datagroup[0] if el is not None] #R1
            datagroup[1] = [el for el in datagroup[1] if el is not None] #R2
            if not datagroup == datagroupori:
                print("warning: some data elements are missing in group {}".format(timepoint))
            
            data_mm_mean = [np.mean(l) for l in datagroup] # mean of R1 and R2
            data_mm_std = [np.std(l) for l in datagroup]
            data_mm_max = [np.max(l) for l in datagroup]
            data_mm_min = [np.min(l) for l in datagroup]
            means_mm.append(data_mm_mean)
            stds_mm.append(data_mm_std)
            maxs_mm.append(data_mm_max)
            mins_mm.append(data_mm_min)
        
        marker = 'o'
        markersize = 4
        # plot rdc
        # plot data R1
        ax1.plot(xrange, np.asarray(means_rd)[:,0], ls='-', marker='o', color=subgroup_colors[0]) 
        ax1.errorbar(xrange, np.asarray(means_rd)[:,0], 
            yerr = np.asarray(stds_rd)[:,0], fmt=None, ecolor=subgroup_colors[0], capsize=8)
        # ax1.fill_between(xrange, np.asarray(mins_rd)[:,0], np.asarray(maxs_rd)[:,0], color=subgroup_colors[0], alpha=0.2)
        ax1.plot(xrange, np.asarray(mins_rd)[:,0], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[0], markersize=markersize)
        ax1.plot(xrange, np.asarray(maxs_rd)[:,0], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[0], markersize=markersize) 
        # plot data R2
        ax1.plot(xrange, np.asarray(means_rd)[:,1], ls='-', marker='o', color=subgroup_colors[1]) 
        ax1.errorbar(xrange, np.asarray(means_rd)[:,1], 
            yerr = np.asarray(stds_rd)[:,1], fmt=None, ecolor=subgroup_colors[1], capsize=8)
        # ax1.fill_between(xrange, np.asarray(mins_rd)[:,1], np.asarray(maxs_rd)[:,1], color=subgroup_colors[1], alpha=0.2)
        ax1.plot(xrange, np.asarray(mins_rd)[:,1], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[1], markersize=markersize)
        ax1.plot(xrange, np.asarray(maxs_rd)[:,1], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[1], markersize=markersize) 
        
        # plot mm
        # plot data R1
        ax2.plot(xrange, np.asarray(means_mm)[:,0], ls='-', marker='o', color=subgroup_colors[0]) 
        ax2.errorbar(xrange, np.asarray(means_mm)[:,0], 
            yerr = np.asarray(stds_mm)[:,0], fmt=None, ecolor=subgroup_colors[0], capsize=8)
        # ax2.fill_between(xrange, np.asarray(mins_mm)[:,0], np.asarray(maxs_mm)[:,0], color=subgroup_colors[0], alpha=0.2)
        ax2.plot(xrange, np.asarray(mins_mm)[:,0], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[0], markersize=markersize)
        ax2.plot(xrange, np.asarray(maxs_mm)[:,0], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[0], markersize=markersize) 
        # plot data R2
        ax2.plot(xrange, np.asarray(means_mm)[:,1], ls='-', marker='o', color=subgroup_colors[1]) 
        ax2.errorbar(xrange, np.asarray(means_mm)[:,1], 
            yerr = np.asarray(stds_mm)[:,1], fmt=None, ecolor=subgroup_colors[1], capsize=8)
        # ax2.fill_between(xrange, np.asarray(mins_mm)[:,1], np.asarray(maxs_mm)[:,1], color=subgroup_colors[1], alpha=0.2)
        ax2.plot(xrange, np.asarray(mins_mm)[:,1], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[1], markersize=markersize)
        ax2.plot(xrange, np.asarray(maxs_mm)[:,1], ls='', marker=marker , color='w', 
            markeredgecolor=subgroup_colors[1], markersize=markersize) 
        
        
        _initaxis([ax1, ax2])
        
        
        
    
    def plot_ring_deployment(self, patients=None, preop=False, ylim=[0, 30], ylim_mm=[20,33.5], saveFig=True):
        """ Plot residual deployment capacity ring individul patients lines
        OUTER, mean peak and valley diameters
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        workbook_vars = self.workbook_variables
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        wbvars = openpyxl.load_workbook(os.path.join(exceldir, workbook_vars), data_only=True)
        
        # init figure
        f1 = plt.figure(num=3, figsize=(11.6, 9.2)) # 4.6
        if preop:
            xlabels = ['Pre','D', '1M', '6M', '12M', '24M']
        else:
            xlabels = ['D', '1M', '6M', '12M', '24M']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        
        # init axis
        ax1 = f1.add_subplot(2,2,1)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax2 = f1.add_subplot(2,2,2)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax3 = f1.add_subplot(2,2,3)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax4 = f1.add_subplot(2,2,4)
        plt.xticks(xrange, xlabels, fontsize = 14)
        
        ax1.set_ylabel('Residual deployment capacity R1 (%)', fontsize=15) # ring deployment capacity
        ax2.set_ylabel('Residual deployment capacity R2 (%)', fontsize=15) # ring deployment capacity
        ax1.set_ylim(ylim)
        ax2.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax2.set_xlim([0.8, len(xlabels)+0.2])
        # plots in mm
        ax3.set_ylabel('Diameter R1 (mm)', fontsize=15) # mean distance pp vv
        ax4.set_ylabel('Diameter R2 (mm)', fontsize=15) # mean distance pp vv
        ax3.set_ylim(ylim_mm)
        ax4.set_ylim(ylim_mm)
        ax3.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax4.set_xlim([0.8, len(xlabels)+0.2])
        
        # lines and colors; 12-class Paired
        colors = itertools.cycle(['#a6cee3','#1f78b4','#b2df8a','#33a02c',
        '#fb9a99','#e31a1c','#fdbf6f','#ff7f00','#cab2d6','#6a3d9a','#ffff99','#b15928'])
        # lStyles = ['-', '--']
        # mStyles = itertools.cycle(['^', '^'])#'D', 's', '+'])
        # marker = 'o'
        markers = ['D', 'o', '^', 's', '*']
        lw = 1
        
        # read data
        colStart = [1, 6] # B, G
        rowStart = 97 # 82 = mean pp vv middle bundle, row 83 excel
        rowStartmm = 96 # 81 = mean pp vv distance middle bundle, row 82 excel
        if patients is None:
            patients = self.patients
        # loop through patient sheets
        for i, patient in enumerate(patients):
            sheet = wb.get_sheet_by_name(patient)
            # read R1/R2
            if patient == ('LSPEAS_023' or 'LSPEAS_024'):
                R1 = sheet.rows[21][1:6] # D to 24 M; prox part of freeflow ring
                R2 = sheet.rows[22][1:6] # D to 24 M; dist part of freeflow ring
            else:    
                R1 = sheet.rows[rowStart][colStart[0]:colStart[0]+5] # +4 is read until 12M
                R2 = sheet.rows[rowStart][colStart[1]:colStart[1]+5] # +4 is read until 12M
                R1mm = sheet.rows[rowStartmm][colStart[0]:colStart[0]+5] # +4 is read until 12M
                R2mm = sheet.rows[rowStartmm][colStart[1]:colStart[1]+5] # +4 is read until 12M
            R1 = [obj.value for obj in R1]
            R2 = [obj.value for obj in R2]
            R1mm = [obj.value for obj in R1mm]
            R2mm = [obj.value for obj in R2mm]
            # read preop applied oversize
            sheet_preop = wbvars.get_sheet_by_name(self.workbook_variables_presheet)
            rowPre = 8 # row 9 in excel
            for j in range(18): # read sheet column with patients
                pt = sheet_preop.rows[rowPre+j][1]
                pt = pt.value
                if pt == patient[-3:]:
                    preR1 = sheet_preop.rows[rowPre+j][19] # 19 = col T
                    preR1 = preR1.value
                    preR2 = sheet_preop.rows[rowPre+j][20] # 20 = col U
                    preR2 = preR2.value
                    devicesize = sheet_preop.rows[rowPre+j][4].value # 4 = col E
                    preR1mm = sheet_preop.rows[rowPre+j][21]
                    preR1mm = preR1mm.value
                    preR2mm = sheet_preop.rows[rowPre+j][22]
                    preR2mm = preR2mm.value
                    break
            # plot
            ls = '-'
            color = next(colors)
            # if i > 11: # through 12 colors
            #     marker = next(mStyles)
            if devicesize == 25.5:
                marker = markers[0]
            elif devicesize == 28:
                marker = markers[1]
            elif devicesize == 30.5:
                marker = markers[2]
            elif devicesize == 32:
                marker = markers[3]
            else:
                marker = markers[4]
            if patient == 'LSPEAS_004': # FEVAR
                color = 'k'
                ls = ':'
            elif patient == 'LSPEAS_023': # endurant
                color = 'k'
                ls = '-.'
            
            # when scans are not scored in excel do not plot '#DIV/0!'
            R1 = [el if not isinstance(el, str) else None for el in R1]
            R2 = [el if not isinstance(el, str) else None for el in R2]
            R1mm = [el if not isinstance(el, str) else None for el in R1mm]
            R2mm = [el if not isinstance(el, str) else None for el in R2mm]
            alpha = 1
            if preop:
                xaxis = xrange[1:]
            else:
                xaxis = xrange
            # plot postop rdc
            ax1.plot(xaxis, R1, ls=ls, lw=lw, marker=marker, color=color, 
            label='%i: ID %s' % (i+1, patient[-3:]), alpha=alpha)
            ax2.plot(xaxis, R2, ls=ls, lw=lw, marker=marker, color=color, 
            label='%i: ID %s' % (i+1, patient[-3:]), alpha=alpha)
            if preop:
                # plot preop rdc
                if not isinstance(preR1, str): # not yet scored in excel so '#DIV/0!'
                    ax1.plot(xrange[:2], [preR1,R1[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
                if not isinstance(preR2, str):
                    ax2.plot(xrange[:2], [preR2,R2[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
            
            # plot in mm postop
            ax3.plot(xaxis, R1mm, ls=ls, lw=lw, marker=marker, color=color, 
            label='%i: ID %s' % (i+1, patient[-3:]), alpha=alpha)
            ax4.plot(xaxis, R2mm, ls=ls, lw=lw, marker=marker, color=color, 
            label='%i: ID %s' % (i+1, patient[-3:]), alpha=alpha)
            if preop:
                # plot in mm preop
                if not isinstance(preR1mm, str): # not yet scored in excel so '#DIV/0!'
                    ax3.plot(xrange[:2], [preR1mm,R1mm[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
                if not isinstance(preR2mm, str):
                    ax4.plot(xrange[:2], [preR2mm,R2mm[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
            
        ax2.legend(loc='upper right', fontsize=8, numpoints=1, title='Patients')
        _initaxis([ax1, ax2, ax3, ax4])
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_ring_deployment.png'), papertype='a0', dpi=300)
        
   
    def box_ring_deployment(self, rows=[76,90] , ylim=[0, 42], saveFig=True):
        """ Boxplot residual deployment capacity ring OUTER, mean peak and valley diameters
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        sheet = wb.get_sheet_by_name('Summary')
        
        # read excel
        colStart = ['E', 'J'] # R1 R2
        colStart = [(openpyxl.cell.column_index_from_string(char)-1) for char in colStart]
        rowStart = rows[0]
        rowEnd = rows[1]
        
        groups = ['D', '1M', '6M', '12M', '24M'] # xlabels
        # get arrays with rdc R1 and R2 all patients in rows
        ttRarray = []
        for i in range(len(groups)):
            tR1 = sheet.columns[colStart[0]+i][rowStart:rowEnd+1] 
            tR1 = [obj.value for obj in tR1]
            tR2 = sheet.columns[colStart[1]+i][rowStart:rowEnd+1] 
            tR2 = [obj.value for obj in tR2]
            ttRarray.append([tR1, tR2])
        
        f1 = plt.figure(num=1, figsize=(7.6, 5))
        ax1 = f1.add_subplot(111)
        
        data = { 'D':ttRarray[0], # D met R1 en R2
                '1M':ttRarray[1],
                '6M':ttRarray[2],
                '12M':ttRarray[3],
                '24M':ttRarray[4]
            }
        
        grouped_boxplot_2subgroups(data, group_names=groups, ax=ax1,
            subgroup_names=['R1', 'R2'], 
            subgroup_colors=['#D02D2E', 'blue'],
            box_width=0.6, box_spacing=1.0)
        plt.show()
        
        # set axis
        ax1.set_ylabel('Residual deployment capacity (%)', fontsize=15) # ring deployment capacity
        _initaxis([ax1])
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'box_ring_deploymentR1R2.png'), papertype='a0', dpi=300)
        
    
    def box_ring_distances(self, rows=[120,134], ylim=[18, 34], saveFig=True):
        """ Boxplot distances ring OUTER, mean peak and valley diameters
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        sheet = wb.get_sheet_by_name('Summary')
        
        # read excel
        colStart = ['E', 'J'] # R1 R2
        colStart = [(openpyxl.cell.column_index_from_string(char)-1) for char in colStart]
        rowStart = rows[0]
        rowEnd = rows[1]
        
        groups = ['D', '1M', '6M', '12M', '24M'] # xlabels
        # get arrays with rdc R1 and R2 all patients in rows
        ttRarray = []
        for i in range(len(groups)):
            tR1 = sheet.columns[colStart[0]+i][rowStart:rowEnd+1] 
            tR1 = [obj.value for obj in tR1]
            tR2 = sheet.columns[colStart[1]+i][rowStart:rowEnd+1] 
            tR2 = [obj.value for obj in tR2]
            ttRarray.append([tR1, tR2])
        
        f1 = plt.figure(num=1, figsize=(7.6, 5))
        ax1 = f1.add_subplot(111)
        
        data = { 'D':ttRarray[0], # D met R1 en R2
                '1M':ttRarray[1],
                '6M':ttRarray[2],
                '12M':ttRarray[3],
                '24M':ttRarray[4]
            }
        
        grouped_boxplot_2subgroups(data, group_names=groups, ax=ax1,
            subgroup_names=['R1', 'R2'], 
            subgroup_colors=['#D02D2E', 'blue'],
            box_width=0.6, box_spacing=1.0, ylim=ylim, legendloc='lower right')
        plt.show()
        
        # set axis
        ax1.set_ylabel('Diameter (%)', fontsize=15) # diam peaks valleys
        _initaxis([ax1])
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'box_ring_diametersR1R2.png'), papertype='a0', dpi=300)
    
    
    def change_in_rdc_D_12(self, rowStart = 53, rowEnd = 67):
        """ Do peaks expand more than valleys?
        """
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        sheet = wb.get_sheet_by_name('Summary')
        
        # read data
        colStart = ['L', 'M', 'N', 'O']#, 'X'] # R1 pp vv R2 pp vv
        colStart = [(openpyxl.cell.column_index_from_string(char)-1) for char in colStart]
        rowStart = rowStart # pt 001 summery sheet
        rowEnd = rowEnd
        nrows = rowEnd - rowStart + 1 
        
        # get arrays with change in rdc all patients
        R1pp = sheet.columns[colStart[0]][rowStart:rowStart+nrows] # 
        R1pp = [obj.value for obj in R1pp]
        R1pp004 = R1pp.pop(3) # pt 004 is in between, remove from list
        R1vv = sheet.columns[colStart[1]][rowStart:rowStart+nrows] # 
        R1vv = [obj.value for obj in R1vv]
        R1vv004 = R1vv.pop(3) # pt 004 is in between, remove from list
        R2pp = sheet.columns[colStart[2]][rowStart:rowStart+nrows] # 
        R2pp = [obj.value for obj in R2pp]
        R2pp004 = R2pp.pop(3) # pt 004 is in between, remove from list
        R2vv = sheet.columns[colStart[3]][rowStart:rowStart+nrows] # 
        R2vv = [obj.value for obj in R2vv]
        R2vv004 = R2vv.pop(3) # pt 004 is in between, remove from list
        
        # boxplot
        data = [R1pp, R1vv, R2pp, R2vv]
        labels = ['R1 peaks', 'R1 valleys', 'R2 peaks', 'R2 valleys']
        
        import plotly
        from plotly.offline import plot
        import plotly.graph_objs as go
        # https://plot.ly/python/box-plots/
        # https://plot.ly/python/axes/
        
        x = ['R1'] * len(R1pp) + ['R2'] * len(R2pp)
        
        trace0 = go.Box(
            y= R1pp + R2pp,
            x = x,
            name = 'peaks',
            boxmean=True,
            fillcolor= 'white', # default is half opacity of line/marker color
            line = dict(
                    color='#ef8a62' # orange/red
            ),
            marker=dict(
                    color='#ef8a62' # orange/red
            ) 
        )
        trace1 = go.Box(
            y= R1vv + R2vv,
            x = x,
            name = 'valleys',
            boxmean=True,
            fillcolor= 'white',
            line = dict(
                    color='#000000'
            ),
            marker=dict(
                    color='#000000'
            )
        )
        data = [trace0, trace1]
        layout = go.Layout(
            yaxis=dict(
                title='Decrease residual deployment capacity ring (%)',
                zeroline=False,
                dtick=5, # steps of 5 on y ax
                tickfont=dict(
                    # family='Old Standard TT, serif',
                    size=26, # size of y ax numbers
                ),
                titlefont = dict(
                        # family='Arial, sans-serif',
                        size=35
                    ),
                showticklabels=True
            ),
            xaxis=dict(
                    zeroline=False,
                    # titlefont=dict(
                    #     # family='Arial, sans-serif',
                    #     size=25,
                    # ),
                    tickfont=dict(
                        # family='Old Standard TT, serif',
                        size=35, # size of x labels
                    ),
                    showticklabels=True
                ),
            legend=dict(
                    font=dict(
                        # family='sans-serif',
                        size=35,
                    ),
                    bordercolor='black',
                    borderwidth=1
                ),
            
            boxmode='group'
        )
        
        fig = go.Figure(data=data, layout=layout)
        plot(fig, image = 'png', image_filename = 'testbox', image_height=1100, image_width=1500) # in w/h pixels)
        
        # plot(data, image = 'png', image_filename = 'testbox', image_height=600, image_width=800) # in w/h pixels
        
        
    def plot_pp_vv_distance_ratio_old(self, saveFig=True):
        """ Plot pp and vv distance ratio
        show asymmetry
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        
        # init figure
        f1 = plt.figure(num=3, figsize=(7.6, 5))
        ax1 = f1.add_subplot(111)
        _initaxis([ax1])
        ax1.set_xlabel('Patient number', fontsize=14)
        ax1.set_ylabel('Asymmetry ratio PP/VV distance', fontsize=14)
        plt.ylim(0.6,1.5)
        ax1.plot([0,15],[1,1], ls='--', color='dimgrey')
        fillstyles = ('full', 'left', 'bottom', 'none')
        colors = ('#ffffcc', '#a1dab4', '#41b6c4', '#2c7fb8', '#253494') 
        # html hex codes from http://colorbrewer2.org/#type=sequential&scheme=YlGnBu&n=5
        legend = ['discharge', '1month', '6months', '12months']#, '24months'] first adjust excel
        
        # read data
        colStart = [1, 6] # B, G
        rowStart = 85 # pp/vv ratio
        patients = self.patients
        xrange = range(1,1+len(patients[:16])) # first 15 patients
        
        for i, patient in enumerate(patients):
            if patient == 'LSPEAS_023':
                break
            sheet = wb.get_sheet_by_name(patient)
            # read R1
            R1ratio = sheet.rows[rowStart][colStart[0]:colStart[0]+4] # +4 is read until 12M
            R1ratio = [obj.value for obj in R1ratio]
            # read R2
            R2ratio = sheet.rows[rowStart][colStart[1]:colStart[1]+4] # +4 is read until 12M
            R2ratio = [obj.value for obj in R2ratio]
            
            # plot R1
            for j in range(len(R1ratio)):
                if i == 0: # legend only once
                    plt.plot(xrange[i], R1ratio[j], marker='D', fillstyle=fillstyles[0], 
                         color=colors[j], ls='None', mec='k', mew=1, label=legend[j]) # each timepoint a specific fill
                else:
                    plt.plot(xrange[i], R1ratio[j], marker='D', fillstyle=fillstyles[0], 
                         color=colors[j], ls='None', mec='k', mew=1) # each timepoint a specific fill
        
            plt.xticks(xrange)
            ax1.legend(loc='upper right', numpoints=1, fontsize=12)
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 'plot_pp_vv_distance_ratio.png'), 
            papertype='a0', dpi=300)
    
    
    def plot_pp_vv_distance_ratio(self, patients=None, preop=False, ylim=[0.6,1.5],saveFig=True):
        """ Plot pp and vv distance ratio
        show asymmetry change over time
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        workbook_vars = self.workbook_variables
        
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        wbvars = openpyxl.load_workbook(os.path.join(exceldir, workbook_vars), data_only=True)
        
        # init figure
        f1 = plt.figure(num=4, figsize=(14.55, 6.9)) # 11.6,4.6
        if preop:
            xlabels = ['Pre','D', '1M', '6M', '12M', '24M']
        else:
            xlabels = ['D', '1M', '6M', '12M', '24M']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        
        # init axis
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, (6.1/5.2)]) 
        # ax1 = f1.add_subplot(1,2,1)
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = 14)
        # ax2 = f1.add_subplot(1,2,2)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = 14)
        
        ax1.set_ylabel('Asymmetry ratio PP/VV R1', fontsize=15)
        ax2.set_ylabel('Asymmetry ratio PP/VV R2', fontsize=15)
        ax1.set_ylim(ylim)
        ax2.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax2.set_xlim([0.8, len(xlabels)+1.1]) # longer for legend
        # plot horizontal line at 1
        ax1.plot([0,max(xrange)],[1,1], ls='--', color='dimgrey')
        ax2.plot([0,max(xrange)],[1,1], ls='--', color='dimgrey')
        
        # lines and colors; 12-class Paired
        colors = itertools.cycle(['#a6cee3','#1f78b4','#b2df8a','#33a02c',
        '#fb9a99','#e31a1c','#fdbf6f','#ff7f00','#cab2d6','#6a3d9a','#ffff99','#b15928'])
        # lStyles = ['-', '--']
        # mStyles = itertools.cycle(['^', '^'])#'D', 's', '+'])
        # marker = 'o'
        markers = ['D', 'o', '^', 's', '*'] # for device size
        lw = 1
        
        # read data
        colStart = [1, 6] # B, G
        rowStart = 100 # 101 = ratio distance PP/VV outer bundle, row 102 excel
        if patients is None:
            patients = self.patients
        # loop through patient sheets
        for i, patient in enumerate(patients):
            sheet = wb.get_sheet_by_name(patient)
            # read R1/R2
            ratioR1 = sheet.rows[rowStart][colStart[0]:colStart[0]+5] # +4 is read until 12M
            ratioR2 = sheet.rows[rowStart][colStart[1]:colStart[1]+5] # +4 is read until 12M
            ratioR1 = [obj.value for obj in ratioR1]
            ratioR2 = [obj.value for obj in ratioR2]

            # read preop applied oversize
            sheet_preop = wbvars.get_sheet_by_name(self.workbook_variables_presheet)
            rowPre = 8 # row 9 in excel
            for j in range(18): # read sheet column with patients
                pt = sheet_preop.rows[rowPre+j][1]
                pt = pt.value
                if pt == patient[-3:]:
                    ratiopreR1 = sheet_preop.rows[rowPre+j][11] # 13 = col N
                    ratiopreR1 = ratiopreR1.value
                    ratiopreR2 = sheet_preop.rows[rowPre+j][12] # 14 = col O
                    ratiopreR2 = ratiopreR2.value
                    devicesize = sheet_preop.rows[rowPre+j][4].value # 4 = col E
                    break
            # plot
            ls = '-'
            color = next(colors)
            # if i > 11: # through 12 colors
            #     marker = next(mStyles)
            if devicesize == 25.5:
                marker = markers[0]
                olb = 'OLB25'
            elif devicesize == 28:
                marker = markers[1]
                olb = 'OLB28'
            elif devicesize == 30.5:
                marker = markers[2]
                olb = 'OLB30'
            elif devicesize == 32:
                marker = markers[3]
                olb = 'OLB32'
            else:
                marker = markers[4]
                olb = 'OLB34'
            if patient == 'LSPEAS_004': # FEVAR
                color = 'k'
                ls = ':'
                olb = 'OLB32'
            elif patient == 'LSPEAS_023': # endurant
                color = 'k'
                ls = '-.'
                olb = 'BODY28'
                
            # when scans are not scored in excel do not plot '#DIV/0!'
            ratioR1 = [el if not isinstance(el, str) else None for el in ratioR1]
            ratioR2 = [el if not isinstance(el, str) else None for el in ratioR2]
            
            alpha = 0.9
            if preop:
                xaxis = xrange[1:]
            else:
                xaxis = xrange
            # plot postop ratio
            ax1.plot(xaxis, ratioR1, ls=ls, lw=lw, marker=marker, color=color, 
            # label='%i: ID %s' % (i+1, patient[-3:]), alpha=alpha)
            label='%i: %s' % (i+1, olb), alpha=alpha)
            ax2.plot(xaxis, ratioR2, ls=ls, lw=lw, marker=marker, color=color, 
            label='%i: %s' % (i+1, olb), alpha=alpha)
            if preop:
                # plot preop ratio
                if not isinstance(ratiopreR1, str): # not yet scored in excel so '#DIV/0!'
                    ax1.plot(xrange[:2], [ratiopreR1,ratioR1[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
                if not isinstance(ratiopreR2, str):
                    ax2.plot(xrange[:2], [ratiopreR2,ratioR2[0]], ls=ls, lw=lw, marker=marker, color=color, alpha=alpha)
        
        ax2.legend(loc='upper right', fontsize=8, numpoints=1, title='Patients')
        _initaxis([ax1, ax2])
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_pp_vv_distance_ratio.png'), papertype='a0', dpi=300)
       
    
        
    def plot_ellipse_pp_vv():
        """
        """
        from stentseg.utils.fitting import sample_ellipse
        import numpy as np
        #todo: wip
        x0, y0 = 0, 0
        r1 = 20 # pp; x
        r2 = 18 # vv; y
        phi = 0
        e = [x0, y0, r1, r2, phi]
        # define line over axis
        dxax1 = np.cos(phi)*r1 
        dyax1 = np.sin(phi)*r1
        dxax2 = np.cos(phi+0.5*np.pi)*r2 
        dyax2 = np.sin(phi+0.5*np.pi)*r2
        # p1ax1, p2ax1 
        r1ax = np.array((x0+dxax1, y0+dyax1), (x0-dxax1, y0-dyax1)) # r1
        # p1ax2, p2ax2 
        r2ax = np.array((x0+dxax2, y0+dyax2), (x0-dxax2, y0-dyax2)) # r2
        
        
        ppax = [[x0-r1*np.cos(phi), x0+r1*np.cos(phi) ], [y0-r1*np.sin(phi), y0+r1*np.sin(phi)]  ] # [x1, x2]  [y1, y2]
        # vvax =  
        
        pp = sample_ellipse(e, N=32) # sample ellipse to get a PointSet r1,r2
        
        f1 = plt.figure(num=4, figsize=(7.6, 5))
        plt.cla()
        ax1 = f1.add_subplot(111)
        _initaxis([ax1])
        ax1.set_xlabel('PP', fontsize=14)
        ax1.set_ylabel('VV', fontsize=14)
        
        ax1.plot(pp[:,0], pp[:,1])
        ax1.axis('equal')
        ax1.plot(r1ax)
        # vv.plot(np.array([p1ax1, p2ax1]), lc='w', lw=2) # major axis
        # vv.plot(np.array([p1ax2, p2ax2]), lc='w', lw=2) # minor axis
        
        ax1.plot(ppax)
        ax1.plot([0, 4], [0, 4], linestyle='-', color='g')


if __name__ == '__main__':
    
    patients =['LSPEAS_001', 'LSPEAS_002',	'LSPEAS_003', 'LSPEAS_005',	
                'LSPEAS_008', 'LSPEAS_009',	'LSPEAS_011', 'LSPEAS_015',	'LSPEAS_017',	
                'LSPEAS_018', 'LSPEAS_019',	'LSPEAS_020', 'LSPEAS_021',	'LSPEAS_022',
                'LSPEAS_025']#, 'LSPEAS_004', 'LSPEAS_023']#, 'LSPEAS_024'] 
    
    # create class object for excel analysis
    foo = ExcelAnalysis() # excel locations initialized in class
    foo.plot_pp_vv_distance_ratio(patients=patients, ylim=[0.6,1.5], saveFig=True)
    # foo.plot_pp_vv_deployment(ring=12, saveFig=False)
    # foo.plot_ring_deployment(patients=patients, ylim=[0, 30], ylim_mm=[20,33.5], saveFig=True)
    # foo.change_in_rdc_D_12()
    
    