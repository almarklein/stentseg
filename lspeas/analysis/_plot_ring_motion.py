""" LSPEAS: Script to analyze and plot motion during the cardiac cycle from
excel pulsatility_expansion or from the dynamic models
Copyright 2018-2019, Maaike A. Koenrades
"""

from lspeas.analysis.utils_analysis import readRingExcel, _initaxis, cols2num, read_deforms
from lspeas.analysis._get_ring_dynamics import print_stats_var_over_time_
from lspeas.analysis._plot_ring_dynamics import print_stats_var_mean_over_time_
import openpyxl
from openpyxl.utils import column_index_from_string
from stentseg.utils.datahandling import select_dir
import sys, os
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np 
import itertools
from matplotlib import gridspec
import visvis as vv
from stentseg.utils.datahandling import select_dir, loadvol, loadmodel
from stentseg.utils.picker import pick3d
from stentseg.utils.visualization import DrawModelAxes, show_ctvolume, plot_points
from stentseg.utils import PointSet, _utils_GUI, visualization
from stentseg.motion.displacement import _calculateAmplitude
from lspeas.utils.vis import showModel3d
from lspeas.utils.normality_statistics import normality_check
from lspeas.analysis._ringPulsatility import point_to_point_pulsatility
import scipy
from scipy import io

class MotionAnalysis():
    """ Create graphs from excel data or from the segmented models
    """
    
    exceldir = select_dir(r'C:\Users\Maaike\SURFdrive\UTdrive\LSPEAS\Analysis', 
                    r'D:\Profiles\koenradesma\SURFdrive\UTdrive\LSPEAS\Analysis')
    dirsaveIm =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
    # basedir = select_dir(r'D:\LSPEAS\LSPEAS_ssdf', 
    #                  r'F:\LSPEAS_ssdf_backup',r'G:\LSPEAS_ssdf_backup')
    
    def __init__(self):
        self.exceldir =  MotionAnalysis.exceldir
        self.dirsaveIm = MotionAnalysis.dirsaveIm
        # self.basedir = MotionAnalysis.basedir
        self.workbook_stent = 'LSPEAS_pulsatility_expansion_avgreg_subp_v2.1.xlsx'
        self.patients =['LSPEAS_001', 
                        'LSPEAS_002',	
                        'LSPEAS_003', 
                        'LSPEAS_005',	
                        'LSPEAS_008', 
                        'LSPEAS_009',	
                        'LSPEAS_011', 
                        'LSPEAS_015',	
                        'LSPEAS_017',	
                        'LSPEAS_018',
                        'LSPEAS_019', 
                        'LSPEAS_020', 
                        'LSPEAS_021', 
                        'LSPEAS_022', 
                        'LSPEAS_025', 
                        #'LSPEAS_023', 
                        #'LSPEAS_024', 
                        #'LSPEAS_004'
                        ]
        
        self.relPosDispNodes = [] # [pavg,pdisplacement]
        self.points_plotted = []
        self.analysis = {}
        self.modelssdf = [] # for s of models (spine point analysis)
        self.storevars = {}
        
        self.fontsize1 = 17 # 14
        self.fontsize2 = 17 # 15
        self.fontsize3 = 13 # legend title
        self.fontsize4 = 10 # legend contents
        
        # self.colorsdirections =  [              # from chevas paper colors
        #                     '#a6cee3', # 1 *
        #                     '#fb9a99', # 2 *
        #                     # '#33a02c', # 3
        #                     '#fdbf6f', # 4 *
        #                     '#1f78b4', # 5 *
        #                     # '#e31a1c', # 6
        #                     # '#b2df8a', # 7 
        #                     # '#cab2d6', # 8 
        #                     # '#ff7f00', # 9
        #                     ]
        
        # self.colorsdirections = ['#D02D2E', '#D02D2E', 'blue', 'blue']
        self.colorsdirections = [              # from phantom paper in vivo plots
                            '#d73027', # 1 red
                            '#fc8d59', # orange
                            '#91bfdb', # blue
                            '#4575b4' # 5
                            ]
    
    def _init_fig_plot_displacement(self, ylim, ylimRel):
        """ use to initialize the displacement plots
        """
        # init figure
        self.f1 = plt.figure(figsize=(12.5, 9.1)) # 11.6, 5.8 or 4.62
        # xlabels = ['0', '10', '20', '30', '40', '50', '60', '70','80','90']
        xlabels = ['0', '', '20', '', '40', '', '60', '','80','']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        fontsize1 = self.fontsize1
        fontsize2 = self.fontsize2
        
        # init axis
        factor = 1
        gs = gridspec.GridSpec(2, 3, width_ratios=[factor, factor, factor]) 
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax3 = plt.subplot(gs[2])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax4 = plt.subplot(gs[3])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        ax1.set_ylabel('Relative position x (mm)', fontsize=fontsize2)
        ax1.set_ylim(ylim)
        # ax1.set_yticks(np.arange(ylim[0], ylim[1], 0.2))
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax2.set_ylabel('Relative position y (mm)', fontsize=fontsize2)
        ax2.set_ylim(ylim)
        # ax2.set_yticks(np.arange(ylim[0], ylim[1], 0.2))
        ax2.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax3.set_ylabel('Relative position z (mm)', fontsize=fontsize2)
        ax3.set_ylim(ylim)
        # ax3.set_yticks(np.arange(ylim[0], ylim[1], 0.2))
        ax3.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax3.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax4.set_ylabel('Displacement 3D (mm)', fontsize=fontsize2) # relative pos wrt avgreg pos
        ax4.set_ylim(ylimRel)
        # ax4.set_yticks(np.arange(ylimRel[0], ylimRel[1], 0.2))
        ax4.set_xlim([0.8, len(xlabels)+0.2]) # len(xlabels)*factor+0.7
        ax4.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        self.ax1, self.ax2, self.ax3, self.ax4 = ax1, ax2, ax3, ax4
        self.xrange = xrange
    
    def plot_displacement(self, patient='LSPEAS_001', ctcodes=['discharge', '24months'],
                        analysis=['ant'], ring='R1', 
                        ylim=[-0.5, 0.5], ylimRel=[0,0.8], saveFig=False):
        """
        Plot relative displacement with respect to position at avgreg and 
        all positions at each phase - for 1 patient
        Rows is row as in excel sheet; 
        analysis='ant' or 'post' or 'left' or 'right';
        ctcodes= [discharge, 1month, 6months, 12months, 24months]
        """
        # init figure
        self._init_fig_plot_displacement(ylim, ylimRel)
        ax1, ax2, ax3, ax4 = self.ax1, self.ax2, self.ax3, self.ax4
        xrange = self.xrange
        
        # plot init
        lw = 1
        # ls = '-'
        # marker = 'o'
        alpha = 0.9
        
        colors = create_iter_colors(type=2)
        colors3d = create_iter_colors(type=3)
        markers = create_iter_markers(type=3)
        lstyles = create_iter_ls(type=3)
        
        # read workbook
        sheetname = 'Output_'+ patient[-3:]
        workbook_stent = os.path.join(self.exceldir, self.workbook_stent)
        # read sheet
        wb = openpyxl.load_workbook(workbook_stent, data_only=True)
        sheet = wb.get_sheet_by_name(sheetname)  
        
        ctnames = []
        for ctcode in ctcodes:
            ctname = ctcode_printname(ctcode)
            ctnames.append(ctname)
            color1 = next(colors)
            color3d = next(colors3d)
            #marker = next(markers)
            ls = next(lstyles)
            out = readRingExcel(sheet, patient,ctcode, ring=ring, motion=True, nphases=10) 
            # out = ant post left right
            #todo: built in check in readRingExcel for ant post
            for a in analysis:
                shortname = a # e.g. 'ant'
                if a == 'ant':
                    pavg, pdisplacement = out[0], out[1]
                elif a == 'post':
                    pavg, pdisplacement = out[2], out[3]
                elif a == 'left':
                    pavg, pdisplacement = out[4], out[5]
                elif a == 'right':
                    pavg, pdisplacement = out[6], out[7]
                
                self.relPosDispNodes.append([pavg,pdisplacement])
                
                # get vector magnitudes of rel displacement for each phase
                vec_rel_displ_magn = np.linalg.norm(pdisplacement, axis=1)
                
                # plot
                marker = next(markers)
                
                # plot x
                ax1.plot(xrange, pdisplacement[:,0], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s-%s' % (ctname,shortname), alpha=alpha)
                # plot y
                ax2.plot(xrange, pdisplacement[:,1], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s-%s' % (ctname,shortname), alpha=alpha)
                # plot z
                ax3.plot(xrange, pdisplacement[:,2], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s-%s' % (ctname,shortname), alpha=alpha)
                # plot 3d vector magnitude
                ax4.plot(xrange, vec_rel_displ_magn, ls=ls, lw=lw, marker=marker, color=color3d, 
                        label='%s-%s' % (ctname,shortname), alpha=alpha)
            
        ax1.legend(loc='upper right', fontsize=self.fontsize4, numpoints=2, title='Analysis:')
        ax2.legend(loc='upper right', fontsize=self.fontsize4, numpoints=2, title='Analysis:')
        ax3.legend(loc='upper right', fontsize=self.fontsize4, numpoints=2, title='Analysis:')
        ax4.legend(loc='upper right', fontsize=self.fontsize4, numpoints=2, title='Analysis:')
        _initaxis([ax1,ax2,ax3,ax4])
    
        if saveFig:
            savename = 'plot_displacement_{}_{}_{}.png'.format(patient[8:],ctnames,analysis)
            savename = savename.replace('[', '')
            savename = savename.replace(']', '')
            savename = savename.replace('\'', '')
            savename = savename.replace(', ', '_')
            plt.savefig(os.path.join(self.dirsaveIm, savename), papertype='a0', dpi=600) 
    
    
    def plot_displacement_selected_points(self, colortype=2, colortype3d=3,
                    markertype=3, linetype=3,
                    ylim=[-0.5, 0.5], ylimRel=[0,0.8], saveFig=False):
        """ plot displacement for point that were selected in 
        show_displacement_3d_select_points
        """
        # init figure
        self._init_fig_plot_displacement(ylim, ylimRel)
        ax1, ax2, ax3, ax4 = self.ax1, self.ax2, self.ax3, self.ax4
        xrange = self.xrange
        
        # plot init
        lw = 1
        alpha = 0.9
        
        colors = create_iter_colors(type=colortype)
        colors3d = create_iter_colors(type=colortype3d)
        markers = create_iter_markers(type=markertype)
        lstyles = create_iter_ls(type=linetype)
        
        # loop through stored node analysis
        for key in self.analysis.keys(): # per scan
            # get order for plot since dict has no order
            order = []
            key2s = []
            for key2 in self.analysis[key].keys(): # per analyzed point
                [i, selectn1,pdisplacement] = self.analysis[key][key2]
                order.append(i)
                key2s.append(key2)
            for j in sorted(order): # 1,2,3..
                index = order.index(j)
                key2 = key2s[index]
                [i, selectn1,pdisplacement] = self.analysis[key][key2]
                color1 = next(colors)
                color3d = next(colors3d)
                ls = next(lstyles)
                
                # get vector magnitudes of rel displacement for each phase
                vec_rel_displ_magn = np.linalg.norm(pdisplacement, axis=1)
                
                # plot
                marker = next(markers)
                
                # plot x
                ax1.plot(xrange, pdisplacement[:,0], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s' % (key2), alpha=alpha)
                # plot y
                ax2.plot(xrange, pdisplacement[:,1], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s' % (key2), alpha=alpha)
                # plot z
                ax3.plot(xrange, pdisplacement[:,2], ls=ls, lw=lw, marker=marker, color=color1, 
                    label='%s' % (key2), alpha=alpha)
                # plot 3d vector magnitude
                ax4.plot(xrange, vec_rel_displ_magn, ls=ls, lw=lw, marker=marker, color=color3d, 
                        label='%s' % (key2), alpha=alpha)
        
        #ax1.legend(loc='upper right', fontsize=10, numpoints=2, title='Analysis:')
        #ax2.legend(loc='upper right', fontsize=10, numpoints=2, title='Analysis:')
        ax3.legend(loc='upper right', fontsize=self.fontsize3, numpoints=2, title='Legend:')
        ax4.legend(loc='upper right', fontsize=self.fontsize4, numpoints=2, title='Analysis:')
        _initaxis([ax1,ax2,ax3,ax4])
    
        if saveFig:
            savename = 'plot_displacement.png'
            savename = savename.replace('[', '')
            savename = savename.replace(']', '')
            savename = savename.replace('\'', '')
            savename = savename.replace(', ', '_')
            plt.savefig(os.path.join(self.dirsaveIm, savename), papertype='a0', dpi=600)
        
    
    def show_displacement_3d(self,patient='LSPEAS_001', ctcode='discharge',
                        analysis=['R1ant', 'R1post'], 
                        showVol='MIP', showModelavgreg=True, locationcheck=True, **kwargs):
        """ Show the displacement of a ring point or points in 3d space with 
        respect to the position at mid cardiac cycle = avgreg
        """
        # load vol for patient and show 3d
        self.f, self.a1, self.label, self.s = showModel3d(self.basedir, patient, ctcode, 
                cropname='ring', showVol=showVol,showmodel=showModelavgreg, **kwargs)
        # read workbook for point positions
        sheetname = 'Output_'+ patient[-3:]
        workbook_stent = os.path.join(self.exceldir, self.workbook_stent)
        # read sheet
        wb = openpyxl.load_workbook(workbook_stent, data_only=True)
        sheet = wb.get_sheet_by_name(sheetname) 
        # ring 1
        out = readRingExcel(sheet, patient,ctcode, ring='R1', motion=True, 
                nphases=10, locationcheck=locationcheck) # ant post left right
        # ring 2 (or distal zenith points R1 for z stent)
        out2 = readRingExcel(sheet, patient,ctcode, ring='R2', motion=True, 
                nphases=10, locationcheck=locationcheck) # ant post left right
        #todo: built in check in readRingExcel for ant post
        # show for given points by analysis and ring
        for a in analysis:
            if a == 'R1ant':
                pavg, pdisplacement = out[0], out[1]
            elif a == 'R1post':
                pavg, pdisplacement = out[2], out[3]
            elif a == 'R1left':
                pavg, pdisplacement = out[4], out[5]
            elif a == 'R1right':
                pavg, pdisplacement = out[6], out[7]
            elif a == 'R2ant':
                pavg, pdisplacement = out2[0], out2[1]
            elif a == 'R2post':
                pavg, pdisplacement = out2[2], out2[3]
            elif a == 'R2left':
                pavg, pdisplacement = out2[4], out2[5]
            elif a == 'R2right':
                pavg, pdisplacement = out2[6], out2[7]
            
            pphases = pdisplacement + pavg
            
            # show these point positions
            points = plot_points(pphases, mc='y', ax=self.a1)
            self.points_plotted.append(points)
            
            #show the points at mid cycle = avgreg
            plot_points(pavg, mc='b', ax=self.a1)
        
        ctname = ctcode_printname(ctcode)
        savename = 'displacement3d_{}_{}_{}.png'.format(patient[8:],ctname,analysis)
        savename = savename.replace('[', '')
        savename = savename.replace(']', '')
        savename = savename.replace('\'', '')
        savename = savename.replace(', ', '_')
        print(savename)
    
    def show_displacement_3d_select_points(self,patient='LSPEAS_001', ctcode='discharge',
                        showVol='MIP', showModelavgreg=True, basedir=None, cropname='ring',
                        analysis=['Anaconda R1ant', 'Anaconda R1post'], **kwargs):
        """ Same as for show_displacement_3d but user can select which nodes from 3d image
        The position and deforms are stored
        analysis = labels for points selected (in order) to store in self for 
        later plotting with legend
        """
        # load vol for patient and show 3d
        if basedir is None: # use default from init
            basedir = self.basedir
        self.f, self.a1, self.label, self.s = showModel3d(basedir, patient, ctcode, 
                cropname=cropname, showVol=showVol,showmodel=showModelavgreg,
                graphname='model', **kwargs)
        
        # combine models if multiple were stored in ssdf
        cnt = 0
        for key in dir(self.s):
            if key.startswith('model'):
                if cnt == 0: # first get graph
                    graph = self.s[key]
                else:
                    graph.add_nodes_from(self.s[key].nodes(data=True)) # also attributes
                    graph.add_edges_from(self.s[key].edges(data=True))
                cnt += 1
        
        # create clickable nodes
        node_points = _utils_GUI.interactive_node_points(graph, scale=0.7)
        selected_nodes = list()
        
        # Initialize label
        t0 = vv.Label(self.a1, '\b{Node nr|location}: ', fontSize=11, color='w')
        t0.position = 0.1, 25, 0.5, 20  # x (frac w), y, w (frac), h
        t0.bgcolor = None
        t0.visible = True
        
        # create dict for this scan for storing displacement vectors of selected nodes
        self.analysis[patient+ctcode] = {}
        
        def on_key(event, node_points):
            if event.key == vv.KEY_DOWN:
                # hide nodes
                for node_point in node_points:
                    node_point.visible = False
                    t0.visible = False
            if event.key == vv.KEY_UP:
                # show nodes
                for node_point in node_points:
                    node_point.visible = True
            if event.key == vv.KEY_ESCAPE:
                # get deforms of nodes selected and plot displacement
                for i, n in enumerate(selected_nodes):
                    # get node
                    selectn1 = n.node # position avgreg mid cycle
                    # get deforms of node
                    n1Deforms = graph.node[selectn1]['deforms']
                    pphases = n1Deforms + selectn1
                    # show these point positions (relative to avgreg)
                    points = plot_points(pphases, mc='y', ax=self.a1)
                    self.points_plotted.append(points)
                    #show the point at mid cycle = avgreg
                    plot_points(selectn1, mc='b', ax=self.a1)
                    self.analysis[patient+ctcode][analysis[i]] = [i, selectn1,n1Deforms]
                    # self.analysis[patient+ctcode+analysis[i]] = [selectn1,n1Deforms]
        
        # Bind event handlers
        self.f.eventKeyDown.Bind(lambda event: on_key(event, node_points) )
        # bind callback functions to node points
        _utils_GUI.node_points_callbacks(node_points, selected_nodes, t0=t0) 
        
        # Print user instructions
        print('')
        print('Esc = finish to plot displacement of selected nodes')
        print('x = axis invisible/visible')

    
    def loadssdf(self,patient='LSPEAS_001', ctcode='discharge',
                        basedir=None, cropname='ring'):
        """ load ssdf to store in self
        """
        if basedir is None:
            basedir = self.basedir
        s = loadmodel(basedir, patient, ctcode, cropname, modelname='modelavgreg')
        self.modelssdf.append(s)
    
    def amplitude_of_node_points_models(self, graphname='modelspine'):
        """ to calculate the displacement of the node points for the models
        in the ssdfs loaded to self with loadssdf, e.g. spine points for error estimation
        """ 
        from stentseg.motion.displacement import _calculateAmplitude
        from lspeas.utils import normality_statistics
        from lspeas.utils.normality_statistics import normality_check
        
        xamplitudes = []
        yamplitudes = []
        zamplitudes = []
        xyzamplitudes = []
        
        for s in self.modelssdf:
            for key in dir(s):
                if key.startswith(graphname):
                    graph = s[key]
                    for i, n in enumerate(sorted(graph.nodes())):
                        relpositions = graph.node[n]['deforms']
                        dmax, p1, p2 = _calculateAmplitude(relpositions, dim = 'x')
                        xamplitudes.append(dmax)
                        dmax, p1, p2 = _calculateAmplitude(relpositions, dim = 'y')
                        yamplitudes.append(dmax)
                        dmax, p1, p2 = _calculateAmplitude(relpositions, dim = 'z')
                        zamplitudes.append(dmax)
                        dmax, p1, p2 = _calculateAmplitude(relpositions, dim = 'xyz')
                        xyzamplitudes.append(dmax)
                    print(i)
        
        # check normality
        W, pValue, normality = normality_check(xamplitudes, alpha=0.05, showhist=True)
        print('Amplitude distribution in x of tested graphnodes gives a pValue of {}'.format(pValue))
        
        # get stats
        xamplitudemean = np.mean(xamplitudes)
        yamplitudemean = np.mean(yamplitudes)
        zamplitudemean = np.mean(zamplitudes)
        xyzamplitudemean = np.mean(xyzamplitudes)
        
        xamplitudestd = np.std(xamplitudes)
        yamplitudestd = np.std(yamplitudes)
        zamplitudestd = np.std(zamplitudes)
        xyzamplitudestd = np.std(xyzamplitudes)
        
        xamplitudemin, xamplitudemax = np.min(xamplitudes), np.max(xamplitudes)
        yamplitudemin, yamplitudemax = np.min(yamplitudes), np.max(yamplitudes)
        zamplitudemin, zamplitudemax = np.min(zamplitudes), np.max(zamplitudes)
        xyzamplitudemin, xyzamplitudemax = np.min(xyzamplitudes), np.max(xyzamplitudes)
        
        xstats = [xamplitudemean, xamplitudestd, xamplitudemin, xamplitudemax]
        ystats = [yamplitudemean, yamplitudestd, yamplitudemin, yamplitudemax]
        zstats = [zamplitudemean, zamplitudestd, zamplitudemin, zamplitudemax]
        xyzstats = [xyzamplitudemean, xyzamplitudestd, xyzamplitudemin, xyzamplitudemax]
        
        return xstats, ystats, zstats, xyzstats
    

    def get_R1R2distance_change_peaks_valleys(self, patients=None, storemat=False):
        """ Read excel pulsatility and expansion to obtain distance change between ring1
        and ring2 during the cardiac cycle at the peaks and valleys, i.e. ant post left right
        """
        if patients == None:
            patients = self.patients
            
        antDistOverTime_pts = [] # mm; 5 values for each patient
        antDistPOverTime_pts = [] # %
        postDistOverTime_pts = []
        postDistPOverTime_pts = []
        leftDistOverTime_pts = [] 
        leftDistPOverTime_pts = [] 
        rightDistOverTime_pts = []
        rightDistPOverTime_pts = []
        
        # read workbook
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        
        # read sheets
        for patient in patients:
            sheetname = 'Output_'+ patient[-3:]
            sheet = wb.get_sheet_by_name(sheetname)  
            # read peaks valleys for each ring
            rowsStart = [18,31,55,68]
            timepoints = ['discharge', '1M', '6M', '12M', '24M']
            # init to collect distance change peaks valleys R1 R2 for all timepoints
            antDistOverTime = [] # mm
            antDistPOverTime = [] # %
            postDistOverTime = []
            postDistPOverTime = []
            leftDistOverTime = [] 
            leftDistPOverTime = [] 
            rightDistOverTime = []
            rightDistPOverTime = []
            
            for i in range(len(timepoints)):
                time = i
                # R1
                colStart = 'B'
                positionAnt, deformsAnt, poscycleAnt = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[0], colStart=colStart, nphases=10 ) # 1x3, 10x3, and 10x3
                positionPost, deformsPost, poscyclePost = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[1], colStart=colStart, nphases=10 )
                positionLeft, deformsLeft, poscycleLeft = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[2], colStart=colStart, nphases=10 )
                positionRight, deformsRight, poscycleRight = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[3], colStart=colStart, nphases=10 )
                # check and reorder ant post left right
                P = [positionAnt, positionPost, positionLeft, positionRight]
                D = [deformsAnt,  deformsPost,  deformsLeft,  deformsRight]
                indexorder = orderlocation(positionAnt, positionPost, positionLeft, positionRight) # index A, P, L, R
                R1positionAnt, R1deformsAnt = P[indexorder[0]], D[indexorder[0]]
                R1positionPost, R1deformsPost = P[indexorder[1]], D[indexorder[1]]
                R1positionLeft, R1deformsLeft = P[indexorder[2]], D[indexorder[2]]
                R1positionRight, R1deformsRight = P[indexorder[3]], D[indexorder[3]]
                
                # R2
                colStart = 'V'
                positionAnt, deformsAnt, poscycleAnt = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[0], colStart=colStart, nphases=10 ) # 1x3, 10x3, and 10x3
                positionPost, deformsPost, poscyclePost = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[1], colStart=colStart, nphases=10 )
                positionLeft, deformsLeft, poscycleLeft = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[2], colStart=colStart, nphases=10 )
                positionRight, deformsRight, poscycleRight = readPosDeformsOverCycle(
                    sheet, time=time, rowStart=rowsStart[3], colStart=colStart, nphases=10 )
                # check and reorder ant post left right
                P = [positionAnt, positionPost, positionLeft, positionRight]
                D = [deformsAnt,  deformsPost,  deformsLeft,  deformsRight]
                indexorder = orderlocation(positionAnt, positionPost, positionLeft, positionRight) # index A, P, L, R
                R2positionAnt, R2deformsAnt = P[indexorder[0]], D[indexorder[0]]
                R2positionPost, R2deformsPost = P[indexorder[1]], D[indexorder[1]]
                R2positionLeft, R2deformsLeft = P[indexorder[2]], D[indexorder[2]]
                R2positionRight, R2deformsRight = P[indexorder[3]], D[indexorder[3]]
                
                # Get distance between R1 R2 ant post left right
                #handle missing scans
                try:
                    output = point_to_point_pulsatility(R1positionAnt, R1deformsAnt, R2positionAnt, R2deformsAnt)
                    distancechangeAnt = output[5][0] # index 5 is pulsatility in [mm, %]
                    distancechangePAnt = output[5][1]
                    output = point_to_point_pulsatility(R1positionPost, R1deformsPost, R2positionPost, R2deformsPost)
                    distancechangePost = output[5][0]
                    distancechangePPost = output[5][1]
                    output = point_to_point_pulsatility(R1positionLeft, R1deformsLeft, R2positionLeft, R2deformsLeft)
                    distancechangeLeft = output[5][0]
                    distancechangePLeft = output[5][1]
                    output = point_to_point_pulsatility(R1positionRight, R1deformsRight, R2positionRight, R2deformsRight)
                    distancechangeRight = output[5][0]
                    distancechangePRight = output[5][1]
                    
                    antDistOverTime.append(distancechangeAnt)
                    antDistPOverTime.append(distancechangePAnt)
                    postDistOverTime.append(distancechangePost)
                    postDistPOverTime.append(distancechangePPost)
                    leftDistOverTime.append(distancechangeLeft) 
                    leftDistPOverTime.append(distancechangePLeft) 
                    rightDistOverTime.append(distancechangeRight)
                    rightDistPOverTime.append(distancechangePRight)
                    
                except ValueError:
                    antDistOverTime.append(np.nan)
                    antDistPOverTime.append(np.nan)
                    postDistOverTime.append(np.nan)
                    postDistPOverTime.append(np.nan)
                    leftDistOverTime.append(np.nan) 
                    leftDistPOverTime.append(np.nan) 
                    rightDistOverTime.append(np.nan)
                    rightDistPOverTime.append(np.nan)
            
            # collect for pts
            antDistOverTime_pts.append(np.asarray(antDistOverTime))
            antDistPOverTime_pts.append(np.asarray(antDistPOverTime))
            postDistOverTime_pts.append(np.asarray(postDistOverTime))
            postDistPOverTime_pts.append(np.asarray(postDistPOverTime))
            leftDistOverTime_pts.append(np.asarray(leftDistOverTime))
            leftDistPOverTime_pts.append(np.asarray(leftDistPOverTime)) 
            rightDistOverTime_pts.append(np.asarray(rightDistOverTime))
            rightDistPOverTime_pts.append(np.asarray(rightDistPOverTime))
            
        # Store to .mat
        if storemat:
            # mm
            self.store_var_to_mat(np.asarray(antDistOverTime_pts), varname='distchangebetweenR1R2_pts_ant')
            self.store_var_to_mat(np.asarray(postDistOverTime_pts), varname='distchangebetweenR1R2_pts_post')
            self.store_var_to_mat(np.asarray(leftDistOverTime_pts), varname='distchangebetweenR1R2_pts_left')
            self.store_var_to_mat(np.asarray(rightDistOverTime_pts), varname='distchangebetweenR1R2_pts_right')
            # %
            self.store_var_to_mat(np.asarray(antDistPOverTime_pts), varname='distchangePbetweenR1R2_pts_ant')
            self.store_var_to_mat(np.asarray(postDistPOverTime_pts), varname='distchangePbetweenR1R2_pts_post')
            self.store_var_to_mat(np.asarray(leftDistPOverTime_pts), varname='distchangePbetweenR1R2_pts_left')
            self.store_var_to_mat(np.asarray(rightDistPOverTime_pts), varname='distchangePbetweenR1R2_pts_right')
    
    def plot_pulsatility_during_cycle(self, patients=['LSPEAS_001'], analysis=['AP'], 
                time = 'discharge', ring='R1', ylim=[20, 35], ylimRel=[-3,3], saveFig=False):
        """ plot change in diametric distance during the cardiac cycle at
        a certain time point (CT scan)
        * time = discharge, 1M, 6M, 12M, 24M
        * analysis = AP, LR, LARP, and/or RALP in list (perhaps add max, mean)
        """
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        
        self.distsAll = {}
        self.distsRelAll = {}
        
        # init figure
        self.f1 = plt.figure(figsize=(11.5, 4.2)) # 11.8, 5.8
        xlabels = ['0', '10', '20', '30', '40', '50', '60', '70','80','90']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        fontsize1 = 16
        fontsize2 = 16
        
        # init axis
        factor = 1.33 # 1.36
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, factor]) # plot right wider
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        ax1.set_ylabel('Distance (mm)', fontsize=fontsize2) # absolute distances
        ax1.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax2.set_ylabel('Relative distance change (mm)', fontsize=fontsize2) # relative dist from avgreg
        ax2.set_ylim(ylimRel)
        ax2.set_xlim([0.8, len(xlabels)*factor+0.2]) # xlim margins 0.2; # longer for legend
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        # plot init
        lw = 1
        alpha = 0.9
        ls = '-'
        markers = ['D', '^', 'o', 's']
        colorsdirections = self.colorsdirections
        
        if patients == None:
            patients = self.patients
        
        if ring == 'R1':
            colStart = 'B'
        elif ring == 'R2':
            colStart = 'V'
        
        for patient in patients:
            # read sheet
            sheet = wb.get_sheet_by_name(patient)
            for a in analysis:
                if a == 'AP':
                    dists, distsRel = readDistancesOverCycle(sheet,rowStart=50, colStart=colStart, time=time)
                elif a == 'LR':
                    dists, distsRel = readDistancesOverCycle(sheet,rowStart=50, colStart=colStart, colOffset=5, time=time)
                elif a == 'LARP':
                    dists, distsRel = readDistancesOverCycle(sheet,rowStart=50, colStart=colStart, colOffset=10, time=time)
                elif a == 'RALP':
                    dists, distsRel = readDistancesOverCycle(sheet,rowStart=50, colStart=colStart, colOffset=15, time=time)
                self.distsAll['{}_{}'.format(patient, a)] = dists
                self.distsRelAll['{}_{}'.format(patient, a)] = distsRel
                
                # plotting
                # color1 = color_per_patient(patient)
                colorindex = ['AP', 'LR', 'LARP', 'RALP']
                color1 = colorsdirections[colorindex.index(a)]
                marker = markers[colorindex.index(a)]
                
                ptlegend = a
                ax1.plot(xrange, dists, ls=ls, lw=lw, marker=marker, color=color1, 
                        label=ptlegend, alpha=alpha)
                ax2.plot(xrange, distsRel, ls=ls, lw=lw, marker=marker, color=color1, 
                        label=ptlegend, alpha=alpha)
                
        ax2.legend(loc='upper right', fontsize=self.fontsize4, numpoints=1, title='Legend:')
        _initaxis([ax1,ax2], axsize=fontsize1)
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_distances_over_cycle_{}.png'.format(analysis)), papertype='a0', dpi=600)
        
    
    def plot_pulsatility_line_per_patient_or_mean(self, patients=None, ylim=[0, 2], 
                    ylim_perc=[0,5], plottype='max', analysis='pulsatility', 
                    storemat=False, saveFig=False, showlegend=True):
        """ Plot maximum pulsatility rings for individul patients (plottype = max) or 
        for all 4 directions using the mean of the patients (plottype = directionsmean)
        plot in absolute change in mm and as percentage
        * Option to store variables as .mat for SPSS
        * Option to read other values in same excel table in sheets by analysis
        analysis = pulsatility, compliance or bloodpressure
        """
        
        exceldir = self.exceldir
        workbook_stent = self.workbook_stent
        wb = openpyxl.load_workbook(os.path.join(exceldir, workbook_stent), data_only=True)
        
        # init figure
        f1 = plt.figure(figsize=(10, 9.2)) # 11.6 to 10 for legend
        xlabels = ['D', '1M', '6M', '12M', '24M']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        
        # init axis
        ax1 = f1.add_subplot(2,2,1)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax2 = f1.add_subplot(2,2,2)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax3 = f1.add_subplot(2,2,3)
        plt.xticks(xrange, xlabels, fontsize = 14)
        ax4 = f1.add_subplot(2,2,4)
        plt.xticks(xrange, xlabels, fontsize = 14)
        
        yname2 = 'mm'
        if analysis == 'pulsatility':
            yname = 'Pulsatility'
        elif analysis == 'compliance':
            yname = 'Compliance'
            yname2 = '%' # percentage change in diametric distance per 100 mmHg
        else:
            yname = analysis
        
        ax1.set_ylabel('{} R1 ({})'.format(yname, yname2), fontsize=15) #Radial distension?
        ax2.set_ylabel('{} R2 ({})'.format(yname, yname2), fontsize=15)
        ax1.set_ylim(ylim)
        ax2.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax2.set_xlim([0.8, len(xlabels)+0.2])
        # plots in perc change cycle
        ax3.set_ylabel('{} R1 (%)'.format(yname), fontsize=15) # mean distance pp vv
        ax4.set_ylabel('{} R2 (%)'.format(yname), fontsize=15) # mean distance pp vv
        ax3.set_ylim(ylim_perc)
        ax4.set_ylim(ylim_perc)
        ax3.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax4.set_xlim([0.8, len(xlabels)+0.2])
        
        # lines and colors; 12-class Paired (ring deployment paper colors)
        colors = itertools.cycle(['#a6cee3','#1f78b4','#b2df8a','#33a02c',
        '#fb9a99','#e31a1c','#fdbf6f','#ff7f00','#cab2d6','#6a3d9a','#ffff99','#b15928'])
        markers = ['D', 'o', '^', 's', '*']
        lw = 1
        
        def readValuesOverTimePerPatient(sheet, colStart, rowStart):
            """ read per patient sheet pulsatility or minimum or compliance...
            for R1 or R2 over time from D to 24M
            * tuple with 4 arrays for pp, vv, midmid1, midmid2 with 5 values over time 
            """
            R1pp = sheet.rows[rowStart[0]][colStart:colStart+9:2] # AX, AZ, BB etc.
            R1vv = sheet.rows[rowStart[1]][colStart:colStart+9:2]
            R1LAR = sheet.rows[rowStart[2]][colStart:colStart+9:2]
            R1RAL = sheet.rows[rowStart[3]][colStart:colStart+9:2]
            
            R1pp = [obj.value for obj in R1pp]
            R1vv = [obj.value for obj in R1vv]
            R1LAR = [obj.value for obj in R1LAR]
            R1RAL = [obj.value for obj in R1RAL]
            
            # handle unscored measures 'NA'
            R1pp = [el if not isinstance(el, str) else np.nan for el in R1pp]
            R1vv = [el if not isinstance(el, str) else np.nan for el in R1vv]
            R1LAR = [el if not isinstance(el, str) else np.nan for el in R1LAR]
            R1RAL = [el if not isinstance(el, str) else np.nan for el in R1RAL]
            
            return np.asarray(R1pp), np.asarray(R1vv), np.asarray(R1LAR), np.asarray(R1RAL)
        
        # init to collect all patients
        R1allP_pts = []
        R1allPperc_pts = []
        R2allP_pts = []
        R2allPperc_pts = []
        
        R1maxP_pts = []
        R1maxPperc_pts = []
        R2maxP_pts = []
        R2maxPperc_pts = []
        
        R1maxPloc_pts = []
        R1maxPpercloc_pts = []
        R2maxPloc_pts = []
        R2maxPpercloc_pts = []
        
        R1meanP_pts = []
        R1meanPperc_pts = [] 
        R2meanP_pts = []
        R2meanPperc_pts = []
        
        # read data
        colsStart = ['AX', 'AY'] # R1, R2
        colsStart = [column_index_from_string(col)-1 for col in colsStart] 
        
        if patients is None:
            patients = self.patients
        # loop through patient sheets
        for i, patient in enumerate(patients):
            rowsStartP = [84, 90, 96, 102] # 84 = row 85 excel; peak peak, valley valley, LARP, RALP pulsatility
            if analysis == 'compliance':
                rowsStartP = [row+3 for row in rowsStartP]
            
            # 19 and 25 rotated placement; switch to use anatomical groups
            if patient == 'LSPEAS_019' or patient == 'LSPEAS_025':
                rowsStartP = [rowsStartP[i] for i in [1,0,3,2]]
            
            rowsStartMin = [row-2 for row in rowsStartP] # minimum diameter; ignore percentage outcome for complicance
            
            if analysis == 'bloodpressure':
                rowsStartP = [80, 81, 80, 81]
            
            sheet = wb.get_sheet_by_name(patient)
            
            # read R1/R2
            if patient == 'LSPEAS_023' or patient == 'LSPEAS_024':
                continue
            else:
                # pulsatility (P) R1
                R1ppP, R1vvP, R1LARP, R1RALP = readValuesOverTimePerPatient(sheet, colsStart[0], rowsStartP)
                R1ppMin, R1vvMin, R1LARMin, R1RALMin = readValuesOverTimePerPatient(sheet, colsStart[0], rowsStartMin)
                R1ppPperc = R1ppP / R1ppMin *100
                R1vvPperc = R1vvP / R1vvMin *100
                R1LARPperc = R1LARP / R1LARMin *100
                R1RALPperc = R1RALP / R1RALMin *100
                R1allP = np.vstack((R1ppP, R1vvP, R1LARP, R1RALP)) # np.stack((..), axis=0)
                R1allPperc = np.vstack((R1ppPperc, R1vvPperc, R1LARPperc, R1RALPperc))
                # max pulsatility R1
                R1maxP = np.nanmax(R1allP, axis=0) # 5 values over time
                R1maxPperc = np.nanmax(R1allPperc, axis=0) # 5 values over time
                # max pulsatility location (0=peak peak, 1=valley valley, 2=LARP, 3=RALP)
                R1maxPloc = [np.where(R1allP[:,i] == R1maxP[i]) for i in range(len(R1maxP))]
                R1maxPpercloc = [np.where(R1allPperc[:,i] == R1maxPperc[i]) for i in range(len(R1maxPperc))]
                # mean pulsatility R1
                R1meanP = np.nanmean(R1allP, axis=0) # 5 values over time
                R1meanPperc = np.nanmean(R1allPperc, axis=0) # 5 values over time
                
                # pulsatility (P) R2
                R2ppP, R2vvP, R2LARP, R2RALP = readValuesOverTimePerPatient(sheet, colsStart[1], rowsStartP)
                R2ppMin, R2vvMin, R2LARMin, R2RALMin = readValuesOverTimePerPatient(sheet, colsStart[1], rowsStartMin)
                R2ppPperc = R2ppP / R2ppMin *100
                R2vvPperc = R2vvP / R2vvMin *100
                R2LARPperc = R2LARP / R2LARMin *100
                R2RALPperc = R2RALP / R2RALMin *100
                R2allP = np.vstack((R2ppP, R2vvP, R2LARP, R2RALP))
                R2allPperc = np.vstack((R2ppPperc, R2vvPperc, R2LARPperc, R2RALPperc))
                # max pulsatility R2
                R2maxP = np.nanmax(R2allP, axis=0) # 5 values over time
                R2maxPperc = np.nanmax(R2allPperc, axis=0) # 5 values over time
                # max pulsatility location (0=peak peak, 1=valley valley, 2=LARP, 3=RALP)
                R2maxPloc = [np.where(R2allP[:,i] == R2maxP[i]) for i in range(len(R2maxP))]
                R2maxPpercloc = [np.where(R2allPperc[:,i] == R2maxPperc[i]) for i in range(len(R2maxPperc))]
                # mean pulsatility R2
                R2meanP = np.nanmean(R2allP, axis=0) # 5 values over time
                R2meanPperc = np.nanmean(R2allPperc, axis=0) # 5 values over time
                
                # collect max for max mean
                R1maxP_pts.append(R1maxP)
                R1maxPperc_pts.append(R1maxPperc)
                R2maxP_pts.append(R2maxP)
                R2maxPperc_pts.append(R2maxPperc)
                
                # collect location of max for location mean
                R1maxPloc_pts.append(R1maxPloc)
                R1maxPpercloc_pts.append(R1maxPpercloc)
                R2maxPloc_pts.append(R2maxPloc)
                R2maxPpercloc_pts.append(R2maxPpercloc)
                
                # collect mean for directions mean
                R1meanP_pts.append(R1meanP)
                R1meanPperc_pts.append(R1meanPperc)
                R2meanP_pts.append(R2meanP)
                R2meanPperc_pts.append(R2meanPperc)
                
                # collect for patient mean per direction
                R1allP_pts.append(R1allP)
                R1allPperc_pts.append(R1allPperc)
                R2allP_pts.append(R2allP)
                R2allPperc_pts.append(R2allPperc)
                
            devicesize = get_devicesize(patient)
            
            # plot
            ls = '-'
            ls2 = '--'
            alpha = 0.9
            color = next(colors)
            if devicesize == 25.5:
                marker = markers[0]
                olb = 'OLB25' 
            elif devicesize == 28:
                marker = markers[1]
                olb = 'OLB28' 
            elif devicesize == 30.5:
                marker = markers[2]
                olb = 'OLB30'
            elif devicesize == 32:
                marker = markers[3]
                olb = 'OLB32'
            else:
                marker = markers[4]
                olb = 'OLB34'
            if patient == 'LSPEAS_004': # FEVAR
                color = 'k'
                ls = ':'
                olb = 'OLB32'
            elif patient == 'LSPEAS_023': # endurant
                color = 'k'
                ls = '-.'
                olb = 'body28'
            
            # plot 
            if plottype == 'max':
                # mm pulsatility
                ax1.plot(xrange, R1maxP, ls=ls, lw=lw, marker=marker, color=color, 
                        label='%s: %s' % (patient[-2:], olb), alpha=alpha)
                ax2.plot(xrange, R2maxP, ls=ls, lw=lw, marker=marker, color=color,
                        label='%s: %s' % (patient[-2:], olb), alpha=alpha)
                
                # in %
                ax3.plot(xrange, R1maxPperc, ls=ls, lw=lw, marker=marker, color=color, 
                        label='%s: %s' % (patient[-2:], olb), alpha=alpha)
                ax4.plot(xrange, R2maxPperc, ls=ls, lw=lw, marker=marker, color=color, 
                        label='%s: %s' % (patient[-2:], olb), alpha=alpha)
            
        # Get pt mean of max pulsatility
        R1maxP_mean = np.nanmean(R1maxP_pts, axis=0) # 1x5 array
        aR1maxP_pts = np.asarray(R1maxP_pts) # 15x5 store as mat and use in spss
        R1maxPperc_mean = np.nanmean(R1maxPperc_pts, axis=0) # 1x5 array
        aR1maxPperc_pts = np.asarray(R1maxPperc_pts) # 15x5 store as mat and use in spss
        
        R2maxP_mean = np.nanmean(R2maxP_pts, axis=0) # 1x5 array
        aR2maxP_pts = np.asarray(R2maxP_pts) # 15x5 store as mat and use in spss
        R2maxPperc_mean = np.nanmean(R2maxPperc_pts, axis=0) # 1x5 array
        aR2maxPperc_pts = np.asarray(R2maxPperc_pts) # 15x5 store as mat and use in spss

        # # check normality
        # for time in range(len(R1maxP)):
        #     W, pValue, normality = normality_check(aR1maxP_pts[:,time], alpha=0.05, showhist=False)
        #     print('Pulsatility distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
        # print('')
        
        if plottype == 'max':
            # plot mean for max pulsatility
            ax1.plot(xrange, R1maxP_mean, ls=ls2, lw=2, marker='p', color='k', 
                    label='Mean', alpha=alpha)
            ax2.plot(xrange, R2maxP_mean, ls=ls2, lw=2, marker='p', color='k', 
                    label='Mean', alpha=alpha)
            ax3.plot(xrange, R1maxPperc_mean, ls=ls2, lw=2, marker='p', color='k', 
                    label='Mean', alpha=alpha)
            ax4.plot(xrange, R2maxPperc_mean, ls=ls2, lw=2, marker='p', color='k', 
                    label='Mean', alpha=alpha)
        
        
        # Get mean per direction
        # -- R1 mm pulsatility
        aR1allP_pts = np.asarray(R1allP_pts) # 15x4x5 array
        # pp
        aR1ppP_pts = aR1allP_pts[:,0,:] # 15x5 array to store as mat and use in spss?
        R1ppP_mean = np.nanmean(aR1ppP_pts, axis=0) # 1x5 array
        # vv
        aR1vvP_pts = aR1allP_pts[:,1,:] # to store as mat and use in spss?
        R1vvP_mean = np.nanmean(aR1vvP_pts, axis=0) 
        # LARP
        aR1LAP_pts = aR1allP_pts[:,2,:] # to store as mat and use in spss?
        R1LAP_mean = np.nanmean(aR1LAP_pts, axis=0) 
        # RALP
        aR1RAP_pts = aR1allP_pts[:,3,:] # to store as mat and use in spss?
        R1RAP_mean = np.nanmean(aR1RAP_pts, axis=0) 
        # check normality
        for i in range(aR1allP_pts.shape[1]):
            for time in range(aR1allP_pts.shape[2]):
                W, pValue, normality = normality_check(aR1allP_pts[:,i,time], alpha=0.05, showhist=False)
                print('Pulsatility distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
            
        # R1 % pulsatility
        aR1allPperc_pts = np.asarray(R1allPperc_pts)
        # pp
        aR1ppPperc_pts = aR1allPperc_pts[:,0,:] # to store as mat and use in spss?
        R1ppPperc_mean = np.nanmean(aR1ppPperc_pts, axis=0) 
        # vv
        aR1vvPperc_pts = aR1allPperc_pts[:,1,:] # to store as mat and use in spss?
        R1vvPperc_mean = np.nanmean(aR1vvPperc_pts, axis=0) 
        # LARP
        aR1LAPperc_pts = aR1allPperc_pts[:,2,:] # to store as mat and use in spss?
        R1LAPperc_mean = np.nanmean(aR1LAPperc_pts, axis=0) 
        # RALP
        aR1RAPperc_pts = aR1allPperc_pts[:,3,:] # to store as mat and use in spss?
        R1RAPperc_mean = np.nanmean(aR1RAPperc_pts, axis=0) 
        # check normality
        for i in range(aR1allP_pts.shape[1]):
            for time in range(aR1allP_pts.shape[2]):
                W, pValue, normality = normality_check(aR1allPperc_pts[:,i,time], alpha=0.05, showhist=False)
                print('Pulsatility distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
        
        # -- R2 mm pulsatility
        aR2allP_pts = np.asarray(R2allP_pts) # 15x4x5 array
        # pp
        aR2ppP_pts = aR2allP_pts[:,0,:] # 15x5 array to store as mat and use in spss?
        R2ppP_mean = np.nanmean(aR2ppP_pts, axis=0) # 1x5 array
        # vv
        aR2vvP_pts = aR2allP_pts[:,1,:] # to store as mat and use in spss?
        R2vvP_mean = np.nanmean(aR2vvP_pts, axis=0) 
        # LARP
        aR2LAP_pts = aR2allP_pts[:,2,:] # to store as mat and use in spss?
        R2LAP_mean = np.nanmean(aR2LAP_pts, axis=0) 
        # RALP
        aR2RAP_pts = aR2allP_pts[:,3,:] # to store as mat and use in spss?
        R2RAP_mean = np.nanmean(aR2RAP_pts, axis=0)
        # check normality
        for i in range(aR2allP_pts.shape[1]):
            for time in range(aR2allP_pts.shape[2]):
                W, pValue, normality = normality_check(aR2allP_pts[:,i,time], alpha=0.05, showhist=False)
                print('Pulsatility distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
        
        # R2 % pulsatility
        aR2allPperc_pts = np.asarray(R2allPperc_pts)
        # pp
        aR2ppPperc_pts = aR2allPperc_pts[:,0,:] # to store as mat and use in spss?
        R2ppPperc_mean = np.nanmean(aR2ppPperc_pts, axis=0) 
        # vv
        aR2vvPperc_pts = aR2allPperc_pts[:,1,:] # to store as mat and use in spss?
        R2vvPperc_mean = np.nanmean(aR2vvPperc_pts, axis=0) 
        # LARP
        aR2LAPperc_pts = aR2allPperc_pts[:,2,:] # to store as mat and use in spss?
        R2LAPperc_mean = np.nanmean(aR2LAPperc_pts, axis=0) 
        # RALP
        aR2RAPperc_pts = aR2allPperc_pts[:,3,:] # to store as mat and use in spss?
        R2RAPperc_mean = np.nanmean(aR2RAPperc_pts, axis=0)
        # check normality
        for i in range(aR2allP_pts.shape[1]):
            for time in range(aR2allP_pts.shape[2]):
                W, pValue, normality = normality_check(aR2allPperc_pts[:,i,time], alpha=0.05, showhist=False)
                print('Pulsatility distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
        
        # get ratio directions by dividing max by min values - to export to .mat for spss
        # R1
        ratioR1ppvvP_pts =     (np.maximum(aR1ppP_pts, aR1vvP_pts) / 
                                np.minimum(aR1ppP_pts, aR1vvP_pts) ) # el wise comparison
        ratioR1ppvvPperc_pts = (np.maximum(aR1ppPperc_pts, aR1vvPperc_pts) / 
                                np.minimum(aR1ppPperc_pts, aR1vvPperc_pts) ) # el wise comparison
        ratioR1LARAP_pts =     (np.maximum(aR1LAP_pts, aR1RAP_pts) / 
                                np.minimum(aR1LAP_pts, aR1RAP_pts) ) # el wise comparison
        ratioR1LARAPperc_pts = (np.maximum(aR1LAPperc_pts, aR1RAPperc_pts) / 
                                np.minimum(aR1LAPperc_pts, aR1RAPperc_pts) ) # el wise comparison
        # R2
        ratioR2ppvvP_pts =     (np.maximum(aR2ppP_pts, aR2vvP_pts) / 
                                np.minimum(aR2ppP_pts, aR2vvP_pts) )  # el wise comparison
        ratioR2ppvvPperc_pts = (np.maximum(aR2ppPperc_pts, aR2vvPperc_pts) / 
                                np.minimum(aR2ppPperc_pts, aR2vvPperc_pts) ) # el wise comparison
        ratioR2LARAP_pts =     (np.maximum(aR2LAP_pts, aR2RAP_pts) / 
                                np.minimum(aR2LAP_pts, aR2RAP_pts) ) # el wise comparison
        ratioR2LARAPperc_pts = (np.maximum(aR2LAPperc_pts, aR2RAPperc_pts) / 
                                np.minimum(aR2LAPperc_pts, aR2RAPperc_pts) ) # el wise comparison
        
        # plotting
        colorsdirections = self.colorsdirections
        
        if plottype == 'directionsmean':
            markers = ['D', '^', 'o', 's']
            capsize = 8
            lw2 = 1.1
            # R1 mm pulsatility
            ax1.plot(xrange, R1ppP_mean, ls=ls, lw=lw, marker=markers[0], color=colorsdirections[0], 
                    label='AP', alpha=alpha)
            ax1.errorbar(xrange, R1ppP_mean, yerr=np.nanstd(aR1ppP_pts, axis=0), fmt=None,
                         ecolor=colorsdirections[0], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax1.plot(xrange, R1vvP_mean, ls=ls, lw=lw, marker=markers[1], color=colorsdirections[1], 
                    label='LR', alpha=alpha)
            ax1.errorbar(xrange, R1vvP_mean, yerr=np.nanstd(aR1vvP_pts, axis=0), fmt=None,
                         ecolor=colorsdirections[1], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax1.plot(xrange, R1LAP_mean, ls=ls2, lw=lw, marker=markers[2], color=colorsdirections[2], 
                    label='LARP', alpha=alpha)
            ax1.errorbar(xrange, R1LAP_mean, yerr=np.nanstd(aR1LAP_pts, axis=0), fmt=None,
                         ecolor=colorsdirections[2], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax1.plot(xrange, R1RAP_mean, ls=ls2, lw=lw, marker=markers[3], color=colorsdirections[3], 
                    label='RALP', alpha=alpha)
            ax1.errorbar(xrange, R1RAP_mean, yerr=np.nanstd(aR1RAP_pts, axis=0), fmt=None,
                         ecolor=colorsdirections[3], capsize=capsize, elinewidth=lw2, capthick=lw2)
            
            # R2 mm pulsatility
            ax2.plot(xrange, R2ppP_mean, ls=ls, lw=lw, marker=markers[0], color=colorsdirections[0], 
                label='AP', alpha=alpha)
            ax2.errorbar(xrange, R2ppP_mean, yerr=np.nanstd(aR2ppP_pts, axis=0), fmt=None,
                ecolor=colorsdirections[0], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax2.plot(xrange, R2vvP_mean, ls=ls, lw=lw, marker=markers[1], color=colorsdirections[1], 
                label='LR', alpha=alpha)
            ax2.errorbar(xrange, R2vvP_mean, yerr=np.nanstd(aR2vvP_pts, axis=0), fmt=None,
                ecolor=colorsdirections[1], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax2.plot(xrange, R2LAP_mean, ls=ls2, lw=lw, marker=markers[2], color=colorsdirections[2], 
                label='LARP', alpha=alpha)
            ax2.errorbar(xrange, R2LAP_mean, yerr=np.nanstd(aR2LAP_pts, axis=0), fmt=None,
                ecolor=colorsdirections[2], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax2.plot(xrange, R2RAP_mean, ls=ls2, lw=lw, marker=markers[3], color=colorsdirections[3], 
                label='RALP', alpha=alpha)
            ax2.errorbar(xrange, R2RAP_mean, yerr=np.nanstd(aR2RAP_pts, axis=0), fmt=None,
                ecolor=colorsdirections[3], capsize=capsize, elinewidth=lw2, capthick=lw2)

            # R1 % pulsatility
            ax3.plot(xrange, R1ppPperc_mean, ls=ls, lw=lw, marker=markers[0], color=colorsdirections[0], 
                label='AP', alpha=alpha)
            ax3.errorbar(xrange, R1ppPperc_mean, yerr=np.nanstd(aR1ppPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[0], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax3.plot(xrange, R1vvPperc_mean, ls=ls, lw=lw, marker=markers[1], color=colorsdirections[1], 
                label='LR', alpha=alpha)
            ax3.errorbar(xrange, R1vvPperc_mean, yerr=np.nanstd(aR1vvPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[1], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax3.plot(xrange, R1LAPperc_mean, ls=ls2, lw=lw, marker=markers[2], color=colorsdirections[2], 
                label='LARP', alpha=alpha)
            ax3.errorbar(xrange, R1LAPperc_mean, yerr=np.nanstd(aR1LAPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[2], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax3.plot(xrange, R1RAPperc_mean, ls=ls2, lw=lw, marker=markers[3], color=colorsdirections[3], 
                label='RALP', alpha=alpha)
            ax3.errorbar(xrange, R1RAPperc_mean, yerr=np.nanstd(aR1RAPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[3], capsize=capsize, elinewidth=lw2, capthick=lw2)
            
            # R2 % pulsatility
            ax4.plot(xrange, R2ppPperc_mean, ls=ls, lw=lw, marker=markers[0], color=colorsdirections[0], 
                label='AP', alpha=alpha)
            ax4.errorbar(xrange, R2ppPperc_mean, yerr=np.nanstd(aR2ppPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[0], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax4.plot(xrange, R2vvPperc_mean, ls=ls, lw=lw, marker=markers[1], color=colorsdirections[1], 
                label='LR', alpha=alpha)
            ax4.errorbar(xrange, R2vvPperc_mean, yerr=np.nanstd(aR2vvPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[1], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax4.plot(xrange, R2LAPperc_mean, ls=ls2, lw=lw, marker=markers[2], color=colorsdirections[2], 
                label='LARP', alpha=alpha)
            ax4.errorbar(xrange, R2LAPperc_mean, yerr=np.nanstd(aR2LAPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[2], capsize=capsize, elinewidth=lw2, capthick=lw2)
            ax4.plot(xrange, R2RAPperc_mean, ls=ls2, lw=lw, marker=markers[3], color=colorsdirections[3], 
                label='RALP', alpha=alpha)
            ax4.errorbar(xrange, R2RAPperc_mean, yerr=np.nanstd(aR2RAPperc_pts, axis=0), fmt=None,
                ecolor=colorsdirections[3], capsize=capsize, elinewidth=lw2, capthick=lw2)
        
        if showlegend:
            if plottype == 'max':
                # ax1.legend(loc='lower right', fontsize=self.fontsize3, numpoints=1, title='Patients')
                ax4.legend(loc='upper right', fontsize=self.fontsize4, numpoints=1, title='Patients')
            elif plottype == 'directionsmean':
                ax4.legend(loc='upper right', fontsize=self.fontsize4, numpoints=1, title='Legend')
            
        _initaxis([ax1, ax2, ax3, ax4])
        
        # Store to .mat
        if storemat:
            if analysis =='bloodpressure':
                self.store_var_to_mat(aR1ppP_pts, varname='BP_systolic_pts')
                self.store_var_to_mat(aR1vvP_pts, varname='BP_diastolic_pts')
            elif analysis == 'pulsatility':
                fname = 'P'
                # max
                self.store_var_to_mat(aR1maxP_pts, varname='R1max{}_pts'.format(fname))
                self.store_var_to_mat(aR2maxP_pts, varname='R2max{}_pts'.format(fname))
                self.store_var_to_mat(aR1maxPperc_pts, varname='R1max{}perc_pts'.format(fname))
                self.store_var_to_mat(aR2maxPperc_pts, varname='R2max{}perc_pts'.format(fname))
                #mean over directions
                self.store_var_to_mat(np.asarray(R1meanP_pts), varname='R1mean{}_pts'.format(fname))
                self.store_var_to_mat(np.asarray(R2meanP_pts), varname='R2mean{}_pts'.format(fname))
                self.store_var_to_mat(np.asarray(R1meanPperc_pts), varname='R1mean{}perc_pts'.format(fname))
                self.store_var_to_mat(np.asarray(R2meanPperc_pts), varname='R2mean{}perc_pts'.format(fname))
                # per direction
                names =    ['R1pp{}_pts'.format(fname), 'R1vv{}_pts'.format(fname), 
                            'R1LA{}_pts'.format(fname), 'R1RA{}_pts'.format(fname)]
                for i, var in enumerate([aR1ppP_pts, aR1vvP_pts, aR1LAP_pts, aR1RAP_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                names =    ['R2pp{}_pts'.format(fname), 'R2vv{}_pts'.format(fname), 
                            'R2LA{}_pts'.format(fname), 'R2RA{}_pts'.format(fname)]
                for i, var in enumerate([aR2ppP_pts, aR2vvP_pts, aR2LAP_pts, aR2RAP_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                names =    ['R1pp{}perc_pts'.format(fname), 'R1vv{}perc_pts'.format(fname), 
                            'R1LA{}perc_pts'.format(fname), 'R1RA{}perc_pts'.format(fname)]
                for i, var in enumerate([aR1ppPperc_pts, aR1vvPperc_pts, aR1LAPperc_pts, aR1RAPperc_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                names =    ['R2pp{}perc_pts'.format(fname), 'R2vv{}perc_pts'.format(fname), 
                            'R2LA{}perc_pts'.format(fname), 'R2RA{}perc_pts'.format(fname)]
                for i, var in enumerate([aR2ppPperc_pts, aR2vvPperc_pts, aR2LAPperc_pts, aR2RAPperc_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                # ratio directions pulsatility
                self.store_var_to_mat(ratioR1ppvvP_pts, varname='ratioR1ppvv{}_pts'.format(fname))
                self.store_var_to_mat(ratioR2ppvvP_pts, varname='ratioR2ppvv{}_pts'.format(fname))
                self.store_var_to_mat(ratioR1LARAP_pts, varname='ratioR1LARA{}_pts'.format(fname))
                self.store_var_to_mat(ratioR2LARAP_pts, varname='ratioR2LARA{}_pts'.format(fname))
            elif analysis == 'compliance': # ignore percentage output
                fname= 'Com'
                # max
                self.store_var_to_mat(aR1maxP_pts, varname='R1max{}_pts'.format(fname))
                self.store_var_to_mat(aR2maxP_pts, varname='R2max{}_pts'.format(fname))
                #mean over directions
                self.store_var_to_mat(np.asarray(R1meanP_pts), varname='R1mean{}_pts'.format(fname))
                self.store_var_to_mat(np.asarray(R2meanP_pts), varname='R2mean{}_pts'.format(fname))
                # per direction
                names =       ['R1pp{}_pts'.format(fname), 'R1vv{}_pts'.format(fname), 
                            'R1LA{}_pts'.format(fname), 'R1RA{}_pts'.format(fname)]
                for i, var in enumerate([aR1ppP_pts, aR1vvP_pts, aR1LAP_pts, aR1RAP_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                names =       ['R2pp{}_pts'.format(fname), 'R2vv{}_pts'.format(fname), 
                            'R2LA{}_pts'.format(fname), 'R2RA{}_pts'.format(fname)]
                for i, var in enumerate([aR2ppP_pts, aR2vvP_pts, aR2LAP_pts, aR2RAP_pts]):
                    self.store_var_to_mat(var, varname=names[i])
                # ratio directions pulsatility
                self.store_var_to_mat(ratioR1ppvvP_pts, varname='ratioR1ppvv{}_pts'.format(fname))
                self.store_var_to_mat(ratioR2ppvvP_pts, varname='ratioR2ppvv{}_pts'.format(fname))
                self.store_var_to_mat(ratioR1LARAP_pts, varname='ratioR1LARA{}_pts'.format(fname))
                self.store_var_to_mat(ratioR2LARAP_pts, varname='ratioR2LARA{}_pts'.format(fname))
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_ring_{}_{}.png'.format(analysis, plottype)), papertype='a0', dpi=600)
    
    
    def store_var_to_mat(self, variable, varname=None, storematdir=None, printstats=True):
        """ Save as .mat to easy copy to spss
        also store var to self
        """
        if storematdir is None:
            storematdir = os.path.join(self.dirsaveIm, 'python_to_mat_output')
        if varname is None:
            varname = 'variable_from_python'
        storemat = os.path.join(storematdir, varname+'.mat')
        
        storevar = dict()
        # check if variable has multiple vars in list
        if isinstance(variable, list):
            for i, var in enumerate(variable):
                name = varname+'{}'.format(i)
                storevar[name] = var
                self.storevars[name] = var
                print_stats_var_over_time_(var, varname=name, dec=1, showvar=False)
                if 'ratio' in name: # not normal distributed
                    print_stats_var_mean_over_time_(var, varname=name, dec=dec, showvar=False, median=True)
                else:
                    print_stats_var_mean_over_time_(var, varname=name, dec=dec, showvar=False)
        else:
            storevar[varname] = variable
            self.storevars[varname] = variable
            print_stats_var_over_time_(variable, varname=varname, dec=1, showvar=False)
            if 'ratio' in varname: # not normal distributed
                print_stats_var_mean_over_time_(variable, varname=varname, dec=1, showvar=False, median=True)
            else:
                print_stats_var_mean_over_time_(variable, varname=varname, dec=1, showvar=False)
        
        storevar['workbook_stent'] = self.workbook_stent
        storevar['patients'] = self.patients
        
        io.savemat(storemat,storevar)
        print()
        print('variable {} was stored as.mat to {}'.format(varname, storemat))
        print()


def readPosDeformsOverCycle(sheet, time=0, rowStart=18, colStart='B', nphases=10 ):
    """ Read deforms during phases in x,y,z in a scan
    colStart is column for R1 = B or R2 = V; rowStart is start of deforms in excel
    * array with point coordinates avgreg 1x3, deforms 10x3, pointpositions 10x3
    """
    rowStart = rowStart-1
    colindex = time # 0,1,2,3,4 for 'Discharge', '1M' ...
    colStart = column_index_from_string(colStart)+colindex*4-1 # B, F, J... 4 cols
    # read deforms
    deforms = read_deforms(sheet,rowStart,colStart,10)
    deforms = np.asarray(deforms) #10x3
    # location in avgreg/mid cycle
    pos = sheet.rows[rowStart-1][colStart:colStart+3] #3tuple
    pos = [obj.value for obj in pos] #3tuple x y z
    pos = np.asarray(pos) #1x3 mid cycle
    # locations during cycle
    # handle unscored scans
    try:
        poscycle = pos+deforms
        poscycle = np.asarray(poscycle) #10x3
    except TypeError: # create arrays with Nans
        pos = np.ones_like(pos)*np.nan
        deforms = np.ones_like(deforms)*np.nan
        poscycle = deforms 
    
    return pos, deforms, poscycle

def orderlocation(positionAnt, positionPost, positionLeft, positionRight):
    """
    origin of vol is right ant superior
    """ 
    A = np.asarray([positionAnt, positionPost, positionLeft, positionRight])
    # find ant
    miny = np.where(A[:,1] == np.min(A[:,1]) ) # y=smallest for anterior
    indexAnt = miny[0] # first el in tuple to get index int
    # find post
    maxy = np.where(A[:,1] == np.max(A[:,1]) )
    indexPost = maxy[0]
    # find left
    maxx = np.where(A[:,0] == np.max(A[:,0]) ) # x=greatest for left
    indexLeft = maxx[0]
    # find right
    minx = np.where(A[:,0] == np.min(A[:,0]) )
    indexRight = minx[0]
    
    try:
        indexorder = [int(indexAnt), int(indexPost), int(indexLeft), int(indexRight)]
    except: #handle missing scans, nan values
        indexorder = [0,1,2,3]
    
    return indexorder

def readDistancesOverCycle(sheet,rowStart=50, colStart='B', colOffset=0, time='discharge', nphases=10):
    """ read distances over all phases cardiac cycle in excel (pulsatility).
    """
    # read distances
    timepoints = ['discharge', '1M', '6M', '12M', '24M']
    colindex = timepoints.index(time)
    rowStart = rowStart-1
    colStart = column_index_from_string(colStart)+colOffset+colindex-1 
    dists = sheet.columns[colStart][rowStart:rowStart+nphases] # 10 rows, 10 phases 
    dists = [obj.value for obj in dists]
    # handle unscored measures 'NA'
    dists = [el if not isinstance(el, str) else np.nan for el in dists]
    # convert to array
    dists = np.asarray(dists)
    # get distance at mid heart cycle
    avgdist = np.nanmean(dists) 
    # relative distances from avgreg
    distsRel = dists - avgdist
    
    return dists, distsRel

def create_iter_colors(type=1):
    if type == 1:
        colors = itertools.cycle([  
                                '#a6cee3', # 1
                                '#1f78b4', # 2 
                                '#b2df8a', # 3
                                '#33a02c', # 4 
                                '#fb9a99', # 5
                                '#e31a1c', # 6
                                '#fdbf6f', # 7 
                                '#ff7f00', # 8 
                                '#cab2d6', # 9
                                '#6a3d9a', # 10
                                '#ffff99', # 11
                                '#b15928'])# 12

    elif type == 2:
        colors =  itertools.cycle([
                                    '#d73027', # 1 red
                                    '#91bfdb', # blue
                                    '#fee090', # yellow
                                    #'#4575b4' # 5
                                    ])
    elif type == 3:
        colors =  itertools.cycle([
                                    '#fc8d59', # orange
                                    '#4575b4', # 5 dark blue
                                    'k'
                                    ])
    elif type == 4:
        colors =  itertools.cycle([
                                    '#d73027', # 1 red
                                    '#91bfdb', # blue
                                    'k', # black
                                    '#fc8d59', # orange
                                    ])
    
    return colors


def create_iter_markers(type=1):
    if type == 1:
        markers = itertools.cycle([
                                    'o', 'o', 'o', 'o',
                                    '^', '^', '^', '^',
                                    's', 's', 's', 's',
                                    'd', 'd', 'd', 'd'  ])
    if type == 2:
        markers = itertools.cycle([
                                    'o', 'o', 'o',
                                    '^', '^', '^',
                                    's', 's', 's'  ])
    
    
    if type == 3:
        markers = itertools.cycle(['o', '^'])
    
    if type == 4:
        markers = itertools.cycle(['o', '^', 's','d'])
    
    if type == 5:
        markers = itertools.cycle(['o', 'o', '.', 's'])
    
    return markers


def create_iter_ls(type=1):
    if type == 1:
        lstyles = itertools.cycle(['-'])
    
    elif type == 2:
        lstyles = itertools.cycle(['-', '-', '-', '--','--','--'])
    
    elif type == 3:
        lstyles = itertools.cycle(['-', '--']) #-.
    
    elif type == 4:
        lstyles = itertools.cycle(['-', '--', '-.'])

    return lstyles

def ctcode_printname(ctcode):
    name = ''
    if ctcode == 'discharge':
        name = 'D'
    elif ctcode == '1month':
        name = '1M'
    elif ctcode == '6months':
        name = '6M'
    elif ctcode == '12months':
        name = '12M'
    elif ctcode == '24months':
        name = '24M'
    return name
    
def get_devicesize(ptcode):
    """ get device size per patient (ptcode)
    """
    if ptcode == 'LSPEAS_001':
        devicesize = 25.5 
    elif ptcode in ['LSPEAS_005','LSPEAS_008','LSPEAS_011','LSPEAS_015','LSPEAS_022']:
        devicesize = 28
    elif ptcode in ['LSPEAS_002','LSPEAS_003','LSPEAS_009','LSPEAS_018','LSPEAS_019','LSPEAS_021']:
        devicesize = 30.5
    elif ptcode == 'LSPEAS_017':
        devicesize = 32
    elif ptcode in ['LSPEAS_020','LSPEAS_025']:
        devicesize = 34
    
    return devicesize

if __name__ == '__main__':
    
    foo = MotionAnalysis()
    
    # foo.plot_displacement(patient='LSPEAS_020', ctcodes=['discharge','24months'], 
    #     analysis=['post', 'right'], ring='R1', ylim=[-0.65, 0.75], ylimRel=[0,1], saveFig=False)
    
    # foo.show_displacement_3d('LSPEAS_020', '24months',
    #                         analysis=['R1post', 'R1right'], 
    #                         clim=(0,2600),showVol='MIP', showModelavgreg=True)
    # 
    # foo.show_displacement_3d('LSPEAS_023', 'discharge',
    #                         analysis=['R1ant', 'R2ant'], 
    #                         clim=(0,2600),showVol='MIP', showModelavgreg=True, locationcheck=False)
    
    # For phantom validation manuscript
    if False:
        # Select points from 3d image
        # step 1 select and visualize 3d
        foo.show_displacement_3d_select_points(patient='lspeasf_C_01', ctcode='d', cropname='ring',
                analysis=['Ring right', 'Spine'], #labels for plot legend 'Ring posterior', 'Spine',
                showVol='MIP', basedir=None)
        # step 2 plot 2d
        foo.plot_displacement_selected_points(colortype=4, colortype3d=3,
                markertype=5, linetype=4,
                ylim=[-0.7, 0.85], ylimRel=[0,0.8], saveFig=False)
    # For phantom validation manuscript
    if False:
        # Get displacement amplitude of spinepoints
        # load models to analyze
        foo.loadssdf(patient='lspeasf_c_01', ctcode='d', basedir=None, cropname='ring')
        foo.loadssdf(patient='lspeas_020', ctcode='discharge', basedir=None, cropname='ring')
        foo.loadssdf(patient='lspeas_023', ctcode='discharge', basedir=None, cropname='ring')
        foo.loadssdf(patient='chevas_09', ctcode='12months', basedir=None, cropname='prox')
        # analyze nodepoint amplitudes of loaded models
        xstats, ystats, zstats, xyzstats = foo.amplitude_of_node_points_models(graphname='modelspine')
        print(xstats)
        print(ystats)
        print(zstats)
        #print(xyzstats)
    
    
    ## Ring motion manuscript
    # ---- storemat True to print stats over time ----
    
    ## plot and store pulsatility
    # foo.plot_pulsatility_line_per_patient_or_mean(patients=None, ylim=[0, 1], ylim_perc=[0,4], 
    #                 plottype='max', analysis='pulsatility', storemat=True, saveFig=False, showlegend=False)
    # foo.plot_pulsatility_line_per_patient_or_mean(patients=None, ylim=[0, 1], ylim_perc=[0,4], 
    #                 plottype='directionsmean', analysis='pulsatility', saveFig=True)
    
    ## store blood pressure
    # foo.plot_pulsatility_line_per_patient_or_mean(patients=None, ylim=[0, 200], ylim_perc=[0,4], 
    #                 plottype='directionsmean', analysis='bloodpressure', storemat=True, saveFig=False)
    
    ## store compliance
    # foo.plot_pulsatility_line_per_patient_or_mean(patients=None, ylim=[0, 9], ylim_perc=[0,4], 
    #                 plottype='max', analysis='compliance', storemat=True, saveFig=False)
    # foo.plot_pulsatility_line_per_patient_or_mean(patients=None, ylim=[0, 9], ylim_perc=[0,4], 
    #                 plottype='directionsmean', analysis='compliance', saveFig=False)
    
    ## plot pulsatility during the cycle
    foo.plot_pulsatility_during_cycle(patients=['LSPEAS_002'], analysis=['AP', 'LR', 'LARP', 'RALP'], 
            time = '12M', ring= 'R1', ylim=[24, 32], ylimRel=[-1,1], saveFig=True)
    
    ## get and store distance between R1 R2 at ant post left right
    # foo.get_R1R2distance_change_peaks_valleys(patients=None, storemat=True)
    
    
    
    
    
    

