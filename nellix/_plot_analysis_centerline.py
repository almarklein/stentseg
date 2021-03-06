""" Module to create plots from automated analysis of motion on the centerlines
Reads Excel output per patient
Created November 2017
Copyright 2017-2019, Maaike A. Koenrades
"""
from stentseg.utils import PointSet
import openpyxl
from stentseg.utils.datahandling import select_dir
import sys, os
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import itertools
from matplotlib import gridspec
from lspeas.analysis.utils_analysis import _initaxis
from lspeas.utils.normality_statistics import normality_check, independent_samples_ttest

class ExcelAnalysisNellix():
    """ Create graphs from excel data
    """
    
    # exceldir = select_dir(r'F:\Nellix_chevas\CT_SSDF\SSDF')
    exceldir = select_dir(r'F:\Nellix_chevas\CT_SSDF\SSDF_automated',
                            r'D:\Nellix_chevas_BACKUP\CT_SSDF\SSDF_automated')
    dirsaveIm =  select_dir(r'C:\Users\Maaike\Desktop','D:\Profiles\koenradesma\Desktop')
    
    def __init__(self):
        self.exceldir =  ExcelAnalysisNellix.exceldir
        self.dirsaveIm = ExcelAnalysisNellix.dirsaveIm
        self.workbook_analysis = 'ChevasStoreOutput'
        self.patients =['chevas_01', 
                        'chevas_02',	
                        'chevas_03', 
                        'chevas_04',	
                        'chevas_05', 
                        'chevas_06',	
                        'chevas_07', 
                        'chevas_08',
                        'chevas_10',
                        'chevas_09', # reintervention of pt 1
                        'chevas_11'  # reintervention of pt 7
                        ]
    
        self.distsAll = [] # distances between all stents that were analyzed
        self.distsRelAll = [] # relative from avgreg distance
        self.posAll = []
        self.relPosAll = []
        self.fontsize1 = 16 # 14
        self.fontsize2 = 16 # 15
        self.fontsize3 = 10.3
    
    def get_angle_change(self, patients=None, analysis='ChimNel', chimneys=['LRA', 'RRA', 'SMA'], angletype='pointdeflection'):
        """ Read angle change for the chimneys or for the angle between prox chimney and nellix
        or for angle between dist chimney and vessel (end-stent angle)
        Analysis: 'ChimNel' or 'Chim' or 'ChimVessel'
        chimneys: ['LRA', 'RRA', 'SMA'] or ['LRA'] get single chimney
        angletype: 'pointdeflection'    --> point with greatest angle change
                   'peakangle'          --> max diff between peak angle over phases
        """
        
        if patients == None:
            patients = self.patients
        
        self.angleChange = []
        self.angleMin = []
        self.angleMax = []
        self.locationOnChimney = [] # location on chimney as percentage distance from prox / length chimney
        self.locationChange = [] # location change of peak angle 
        self.lengthchimneys = []
        
        # read workbooks
        for patient in patients:
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            # read sheet
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            sheetnames = wb.get_sheet_names()
            
            if analysis == 'Chim': # chimney angle
                for a in chimneys:
                    # set row for type of angle for chim angle analysis
                    if angletype == 'pointdeflection':
                        row1 = 3
                        # see which sheetname, not known if NelL or NelR
                        for sheetname in sheetnames:
                            if sheetname.startswith('Ang_'+a):
                                sheet = wb.get_sheet_by_name(sheetname)
                                # read change
                                angchange = readMaxChange(sheet, row=row1, colStart=1)
                                self.angleChange.append(angchange)
                                # where was point of max deflection on ccl?
                                pointlocation, distfromproxendchimney, lengthchimney = readLocationPointDeflection(sheet, row=9, colStart=1)
                                self.locationOnChimney.append(pointlocation) # percentage of chimney length
                                self.lengthchimneys.append(lengthchimney)
                                break # next a
                    elif angletype == 'peakangle':
                        row1 = 16
                        # see which sheetname, not known if NelL or NelR
                        for sheetname in sheetnames:
                            if sheetname.startswith('Ang_'+a):
                                sheet = wb.get_sheet_by_name(sheetname)
                                # read change
                                angchange = readMaxChange(sheet, row=row1, colStart=1)
                                self.angleChange.append(angchange)
                                angmin, angmax = readMinMax(sheet, row=row1+1, colStart=1, correctorientation=True)
                                self.angleMin.append(angmin)
                                self.angleMax.append(angmax)
                                # how did location of peakangle change during cycle?
                                locationchange = readLocationChange(sheet, row=19, colStart=1, nphases=10)
                                self.locationChange.append(locationchange)
                                # where was point of peak angle at mid cycle?
                                pointlocation, distfromproxendchimney, lengthchimney = readLocationPointDeflection(sheet, row=24, colStart=1)
                                self.locationOnChimney.append(pointlocation) # percentage of chimney length
                                self.lengthchimneys.append(lengthchimney)
                                break # next a
                            
            elif analysis == 'ChimNel': # chimney-to-Nellix angle
                for a in chimneys:
                    # see which sheetname, not known if NelL or NelR
                    for sheetname in sheetnames:
                        if sheetname.startswith('Ang_'+a+'_Nel'):
                            sheet = wb.get_sheet_by_name(sheetname)
                            # read change
                            angchange = readMaxChange(sheet, row=4, colStart=1)
                            self.angleChange.append(angchange)
                            angmin, angmax = readMinMax(sheet, row=5, colStart=1, correctorientation=True)
                            self.angleMin.append(angmin)
                            self.angleMax.append(angmax)
                            break # next a
            
            elif analysis == 'ChimVessel': # chimney-to-Nellix angle
                for a in chimneys:
                    # see which sheetname, not known if NelL or NelR
                    for sheetname in sheetnames:
                        if sheetname.startswith('Ang_'+a+'_Vessel'):
                            sheet = wb.get_sheet_by_name(sheetname)
                            # read change
                            angchange = readMaxChange(sheet, row=4, colStart=1)
                            self.angleChange.append(angchange)
                            angmin, angmax = readMinMax(sheet, row=5, colStart=1, correctorientation=True)
                            self.angleMin.append(angmin)
                            self.angleMax.append(angmax)
                            break # next a
            
            
        # check normality anglechange
        W, pValue, normality = normality_check(self.angleChange, alpha=0.05, showhist=False)
        print('')
        print('AngleChange distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
        print('')
        print('Average maximum angle change: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.angleChange),
                                            np.std(self.angleChange),
                                            np.min(self.angleChange),
                                            np.max(self.angleChange)
                                            ))
        # location and min max peak angle
        if analysis == 'Chim':
            print('')
            print('Average location on chimney as percentage: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.mean(self.locationOnChimney),
                                                np.std(self.locationOnChimney),
                                                np.min(self.locationOnChimney),
                                                np.max(self.locationOnChimney)
                                                )) # for peakangle this is at mid cardiac cycle
            print('Average length of chimney stents (mm): {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.mean(self.lengthchimneys),
                                                np.std(self.lengthchimneys),
                                                np.min(self.lengthchimneys),
                                                np.max(self.lengthchimneys)
                                                ))
            if angletype == 'peakangle':
                print('Average location change of peak angle: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                    np.mean(self.locationChange),
                                                    np.std(self.locationChange),
                                                    np.min(self.locationChange),
                                                    np.max(self.locationChange)
                                                    ))
                print('')
                print('Average minimum Peak angle cycle: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                    np.mean(self.angleMin),
                                                    np.std(self.angleMin),
                                                    np.min(self.angleMin),
                                                    np.max(self.angleMin)
                                                    ))
                print('')
                print('Average maximum Peak angle cycle: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                    np.mean(self.angleMax),
                                                    np.std(self.angleMax),
                                                    np.min(self.angleMax),
                                                    np.max(self.angleMax)
                                                    ))
        else:
            print('')
            print('Average minimum Vector angle cycle: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.mean(self.angleMin),
                                                np.std(self.angleMin),
                                                np.min(self.angleMin),
                                                np.max(self.angleMin)
                                                ))
            print('')
            print('Average maximum Vector angle cycle: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                                np.mean(self.angleMax),
                                                np.std(self.angleMax),
                                                np.min(self.angleMax),
                                                np.max(self.angleMax)
                                                ))
    
    def get_distance_change(self, patients=None, analysis='NelNel', chimneys=['LRA', 'RRA', 'SMA']):
        """ Read distance change between nellix-nellix and nellix-chimneys for all patients
        Analysis: 'NelNel' or 'ChimNel'
        Chimneys: ['LRA', 'RRA', 'SMA'] or ['LRA'] get single chimney; only used with ChimNel analysis
        """
        
        if patients == None:
            patients = self.patients
        
        self.distanceChange = []
        
        # read workbooks
        for patient in patients:
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            # read sheet
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            sheetnames = wb.get_sheet_names()
            
            if analysis == 'NelNel':
                sheetname = 'Dist_NelL_NelR'
                sheet = wb.get_sheet_by_name(sheetname)
                # read distance change
                distchange = readMaxChange(sheet, row=6, colStart=1)
                self.distanceChange.append(distchange)
            elif analysis == 'ChimNel':
                for a in chimneys:
                    # see which sheetname, not known if NelL or NelR
                    for sheetname in sheetnames:
                        if sheetname.startswith('Dist_'+a):
                            sheet = wb.get_sheet_by_name(sheetname)
                            # read distance change
                            distchange = readMaxChange(sheet, row=6, colStart=1)
                            self.distanceChange.append(distchange)
                            break # next a
        
        # check normality
        try:
            W, pValue, normality = normality_check(self.distanceChange, alpha=0.05, showhist=False)
            print('DistanceChange distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
        except ValueError:
            pass
        print('Average maximum distance change: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.distanceChange),
                                            np.std(self.distanceChange),
                                            np.min(self.distanceChange),
                                            np.max(self.distanceChange)
                                            ))
    
    def get_segment_displacement(self, patients=None, analysis=['NelL'],location='prox'):
        """ Read segment mean displacement in x,y,z direction from excels for all patients
        Location: prox or dist
        Analysis:  [NelL,NelR] or [NelL] or [NelR] or [SMA] or [LRA] or [RRA] or 
        [vRRA] [vLRA] [vSMA] for vessel portion distal to stent
        """
        if patients == None:
            patients = self.patients
            
        self.segmentDisplacementX = []
        self.segmentDisplacementY = []
        self.segmentDisplacementZ = []
        self.segmentDisplacement3d = []
        
        # read workbooks
        for patient in patients:
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            # read sheet
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            
            for a in analysis:
                sheetname = 'Motion_'+a+location
                # see if patient has this stent analysis
                try:
                    sheet = wb.get_sheet_by_name(sheetname)
                except KeyError:
                    continue # next analysis
                
                # read segment displ
                displX, displY, displZ, displ3d = readSegmentDisplacement(sheet, rows=[4,5,6,7], colStart=1)
                self.segmentDisplacement3d.append(displ3d)
                self.segmentDisplacementX.append(displX)
                self.segmentDisplacementY.append(displY)
                self.segmentDisplacementZ.append(displZ)
        
        # check normality
        try: # does not work if less than 3 values
            W, pValue, normality = normality_check(self.segmentDisplacementX, alpha=0.05, showhist=False)
            print('DisplacementX distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            W, pValue, normality = normality_check(self.segmentDisplacementY, alpha=0.05, showhist=False)
            print('DisplacementY distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            W, pValue, normality = normality_check(self.segmentDisplacementZ, alpha=0.05, showhist=False)
            print('DisplacementZ distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            W, pValue, normality = normality_check(self.segmentDisplacement3d, alpha=0.05, showhist=False)
            print('Displacement3d distribution normal:{} (pValue of {:.3f})'.format(normality, pValue))
            print('')
        except ValueError:
            pass
        print('Average segment displacement in x: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.segmentDisplacementX),
                                            np.std(self.segmentDisplacementX),
                                            np.min(self.segmentDisplacementX),
                                            np.max(self.segmentDisplacementX)
                                            ))
        print('')
        print('Average segment displacement in y: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.segmentDisplacementY),
                                            np.std(self.segmentDisplacementY),
                                            np.min(self.segmentDisplacementY),
                                            np.max(self.segmentDisplacementY)
                                            ))
        
        print('')
        print('Average segment displacement in z: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.segmentDisplacementZ),
                                            np.std(self.segmentDisplacementZ),
                                            np.min(self.segmentDisplacementZ),
                                            np.max(self.segmentDisplacementZ)
                                            ))
        print('')
        print('Average segment displacement in 3d: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
                                            np.mean(self.segmentDisplacement3d),
                                            np.std(self.segmentDisplacement3d),
                                            np.min(self.segmentDisplacement3d),
                                            np.max(self.segmentDisplacement3d)
                                            ))
        
    
    def plot_displacement(self, patients=None, analysis=['NelLprox'], rows=[8,9], 
                ylim=[-0.5, 0.5], saveFig=False):
        """ Plot relative displacement with respect to position at avgreg and 
        all positions at each phase. 
        Rows is row as in excel sheet; analysis='NelRprox' or 'LRAprox' or 'LRAdist'
        """
        # init figure
        self.f1 = plt.figure(figsize=(11.8, 4.562)) # 11.6, 5.8
        xlabels = ['0', '', '20', '', '40', '', '60', '','80','']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        xrange2 = np.asarray([1,2,3,4,5,6,7,8,9,10])
        xrange2 = xrange2 + 0.5
        fontsize1 = self.fontsize1
        fontsize2 = self.fontsize2
        fontsize3 = self.fontsize3
        
        # init axis
        factor = 1
        gs = gridspec.GridSpec(1, 3, width_ratios=[1, factor, factor]) 
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax3 = plt.subplot(gs[2])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        ax1.set_ylabel('Relative position (mm)', fontsize=fontsize2)
        ax1.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        ax1.set_title('x',fontsize=fontsize2+1)
        
        # ax2.set_ylabel('Displacement 3D (mm)', fontsize=fontsize2) # relative pos wrt avgreg
        # ax2.set_ylim(ylimRel)
        # ax2.set_xlim([0.8, len(xlabels)*factor+0.7]) # xlim margins 0.2
        # ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        # ax2.set_ylabel('Relative position x/y/z (mm)', fontsize=fontsize2)
        ax2.set_ylim(ylim)
        ax2.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        ax2.set_title('y',fontsize=fontsize2+1)
        
        # ax3.set_ylabel('Relative position x/y/z (mm)', fontsize=fontsize2)
        ax3.set_ylim(ylim)
        ax3.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax3.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        ax3.set_title('z',fontsize=fontsize2+1)
        
        # plot init
        lw = 1
        # ls = '-'
        # marker = 'o'
        alpha = 0.9
        
        lstyles = create_iter_ls(type=3)
        
        if patients == None:
            patients = self.patients
            
        for patient in patients:
            # color = color_per_patient(patient)
            ls = next(lstyles)
            # read workbook
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            # read sheet
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            sheetnames = wb.get_sheet_names()
            for a in analysis:
                sheetname = 'Motion_'+a
                # see if patient has this stent analysis
                try:
                    sheet = wb.get_sheet_by_name(sheetname)
                except KeyError:
                    continue # next analysis
            
                shortname = sheetname[7:] # e.g. 'RRAprox'
                
                pavg, pphases, pdisplacement = readPositionPhases(sheet,rows=rows,colStart=1)
                self.relPosAll.append(pdisplacement)
                
                # # get vector magnitudes between phases
                # vec_between_phases = []
                # for i, p in enumerate(pphases[:-1]):
                #     vec_between_phases.append(p-pphases[i+1])
                # # add vector between last and first phase
                # vec_last_to_first = pphases[-1] - pphases[0]
                # vec_between_phases.append(vec_last_to_first)
                # # to array for magnitude
                # vec_between_phases = np.asarray(vec_between_phases)
                # vec_between_phases_magn = np.linalg.norm(vec_between_phases, axis=1) # vector magnitude for each phase
                
                marker = marker_per_chimney(shortname)
                color = color_per_stent(shortname)
                
                # plot legend
                if patient == 'chevas_09':
                    ptlegend = '01**'+':'+shortname # reintervention of pt 1
                elif patient == 'chevas_10':
                    ptlegend = '09*'+': '+shortname
                elif patient == 'chevas_11':
                    ptlegend = '07**'+':'+shortname # reintervention of pt 7
                elif patient == 'chevas_09_thin':
                    ptlegend = '01**-thin'+':'+shortname # reintervention of pt 1
                else:
                    ptlegend = patient[7:]+':  '+shortname
                
                # plot
                for i, ax in enumerate([ax1, ax2, ax3]): # for x,y,z
                    ax.plot(xrange, pdisplacement[:,i], ls=ls, lw=lw, marker=marker, color=color, 
                            label=ptlegend, alpha=alpha)
            
        ax3.legend(loc='upper right', fontsize=10, numpoints=2, title='Legend:')
        _initaxis([ax1,ax2,ax3])
    
        if saveFig:
            savename = 'plot_displacement_{}_{}.png'.format(patient[7:],analysis)
            savename = savename.replace('[', '')
            savename = savename.replace(']', '')
            savename = savename.replace('\'', '')
            savename = savename.replace(', ', '_')
            plt.savefig(os.path.join(self.dirsaveIm, savename), papertype='a0', dpi=600) 
        
    def plot_distances_between_points(self, patients=None, analysis=['NelNel'], rows=[9,5], 
                ylim=[0, 32], ylimRel=[-0.5,0.5], saveFig=False):
        """
        Plot relative distance change with respect to distance at avgreg and 
        absolute distance at each phase. 
        Rows is row as in excel sheet; analysis='NelNel' or 'ChimNel'
        """
        # init figure
        self.f1 = plt.figure(figsize=(11.8, 5.8)) # 11.6, 4.62
        xlabels = ['0', '10', '20', '30', '40', '50', '60', '70','80','90']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        fontsize1 = self.fontsize1
        fontsize2 = self.fontsize2
        fontsize3 = self.fontsize3
        
        # init axis
        factor = 1.36 # 1.33
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, factor]) # plot right wider
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        ax1.set_ylabel('Distance (mm)', fontsize=fontsize2) # absolute distances
        ax1.set_ylim(ylim)
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        ax2.set_ylabel('Relative distance change (mm)', fontsize=fontsize2) # relative dist from avgreg
        ax2.set_ylim(ylimRel)
        ax2.set_xlim([0.8, len(xlabels)*factor+0.2]) # xlim margins 0.2; # longer for legend
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        # plot init
        lw = 1
        # ls = '-'
        # marker = 'o'
        alpha = 0.9
        
        # colors = create_iter_colors()
        # markers = create_iter_markers()
        lstyles = create_iter_ls(type=1)
        
        if patients == None:
            patients = self.patients
        
        for patient in patients:
            # read workbook
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            # read sheet
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            sheetnames = wb.get_sheet_names()
            for a in analysis:
                for sheetname in sheetnames:
                    sheet = None
                    if a == 'ChimNel':
                        if (sheetname.startswith('Dist_RRA') or
                                sheetname.startswith('Dist_LRA') or
                                sheetname.startswith('Dist_SMA')):
                            sheet = wb.get_sheet_by_name(sheetname)  
                            shortname = 'Nel-'+sheetname[5:8] # e.g. 'RRA_Nel' to Nel-RRA
                    elif a == 'NelNel':
                        if sheetname.startswith('Dist_Nel'):
                            sheet = wb.get_sheet_by_name(sheetname)
                            shortname = 'Nel-Nel'
                    if not sheet is None:        
                        dists, distsRel = readDistancesBetweenPoints(sheet,rows=rows,colStart=1)
                        self.distsAll.append(dists)
                        self.distsRelAll.append(distsRel)
                        
                        color1 = color_per_patient(patient)
                        # color1 = color_per_stent(shortname) # for 09 vs 09 thin
                        marker = marker_per_chimney(shortname)
                        ls = next(lstyles)
                        # plot
                        if patient == 'chevas_09':
                            ptlegend = '01**'+':'+shortname # reintervention of pt 1
                            ls = '--'
                        elif patient == 'chevas_10':
                            ptlegend = '09*'+': '+shortname
                            ls = '--'
                        elif patient == 'chevas_11':
                            ptlegend = '07**'+':'+shortname # reintervention of pt 7
                            ls = '--'
                        elif patient == 'chevas_09_thin':
                            ptlegend = '01**-thin'+':'+shortname # reintervention of pt 1
                            ls = '--'
                        else:
                            ptlegend = patient[7:]+':  '+shortname
                        
                        ax1.plot(xrange, dists, ls=ls, lw=lw, marker=marker, color=color1, 
                                label=ptlegend, alpha=alpha)
                        ax2.plot(xrange, distsRel, ls=ls, lw=lw, marker=marker, color=color1, 
                                label=ptlegend, alpha=alpha)
                
        ax2.legend(loc='upper right', fontsize=fontsize3, numpoints=2, title='Legend:')
        _initaxis([ax1,ax2], axsize=fontsize1)
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_distances_between_points_{}.png'.format(analysis)), papertype='a0', dpi=600)


    def plot_angles_chimney(self, patients=None, analysis='AngChim', 
                ylim=[0, 47], ylimRel=[-2,2], saveFig=False):
        """ Plot Angulation or Tortuosity of chimney. 
        analysis: 'AngChim' or AngChimNel or 'Tort'
        Angulation AngChim is pointdeflection angle
        Angulation AngChimNel is vector angle during cycle
        Angulation AngChimVessel is vector angle between stent and distal vessel (end-stent angle)
        """
        # init figure
        self.f2 = plt.figure(figsize=(11.8, 5.8)) # 11.6, 4.62
        xlabels = ['0', '10', '20', '30', '40', '50', '60', '70','80','90']
        xrange = range(1,1+len(xlabels)) # not start from x=0 to play with margin xlim
        fontsize1 = self.fontsize1
        fontsize2 = self.fontsize2
        fontsize3 = self.fontsize3
        
        # init axis
        factor = 1.36
        gs = gridspec.GridSpec(1, 2, width_ratios=[1, factor]) # plot right wider
        ax1 = plt.subplot(gs[0])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        ax2 = plt.subplot(gs[1])
        plt.xticks(xrange, xlabels, fontsize = fontsize1)
        
        # init to collect angles/tortuosity
        self.angsAll = []
        self.angsRelAll = []
        self.tortsAll = []
        self.tortsRelAll = []
        
        if 'Ang' in analysis:
            ax1.set_ylabel('Angulation ($\degree$)', fontsize=fontsize2) # absolute
        elif 'Tort' in analysis:
            ax1.set_ylabel('Tortuosity (AU)', fontsize=fontsize2) # absolute
            yticks = 0.02
        if analysis=='AngChim':
            yticks = 5
        elif analysis=='AngChimNel':
            yticks = 10
        elif analysis == 'AngChimVessel':
            yticks = 10
        ax1.set_ylim(ylim)
        ax1.set_yticks(np.arange(ylim[0], ylim[1], yticks))
        ax1.set_xlim([0.8, len(xlabels)+0.2]) # xlim margins 0.2
        ax1.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        
        if 'Ang' in analysis:
            ax2.set_ylabel('Relative angulation change ($\degree$)', fontsize=fontsize2) # relative dist from avgreg
        elif 'Tort' in analysis:
            ax2.set_ylabel('Relative tortuosity change (AU)', fontsize=fontsize2) # relative dist from avgreg
        ax2.set_ylim(ylimRel)
        ax2.set_xlim([0.8, len(xlabels)*factor+0.2]) # xlim margins 0.2; # longer for legend
        ax2.set_xlabel('Phase in cardiac cycle', fontsize=fontsize2)
        yticksRel = 0.5
        ax2.set_yticks(np.arange(ylimRel[0], ylimRel[1], yticksRel))
        
        # plot init
        lw = 1
        # ls = '-'
        # marker = 'o'
        alpha = 0.9
        
        lstyles = create_iter_ls(type=1)
        
        def plot_per_chimney(sheetname, analysis, patient, xrange, y, yrel):
            """ To plot per chimney available for patient
            """        
            if analysis =='AngChimNel':
                shortname = 'Nel-'+sheetname[4:7] # e.g. 'RRA_Nel' to Nel-RRA
            elif analysis == 'AngChimVessel':
                shortname = sheetname[4:7] # e.g. LRA
            else:
                shortname = sheetname[-3:] # LRA or RRA or SMA
            
            color1 = color_per_patient(patient)
            # color1 = color_per_stent(shortname) # for 09 vs 09 thin
            marker = marker_per_chimney(shortname)
            ls = next(lstyles)
            # plot
            if patient == 'chevas_09':
                ptlegend = '01**'+':'+shortname # reintervention of pt 1
                ls = '--'
            elif patient == 'chevas_10':
                ptlegend = '09*'+': '+shortname
                ls = '--'
            elif patient == 'chevas_11':
                ptlegend = '07**'+':'+shortname # reintervention of pt 7
                ls = '--'
            elif patient == 'chevas_09_thin':
                ptlegend = '01**-thin'+':'+shortname # reintervention of pt 1
                ls = '--'
            else:
                ptlegend = patient[7:]+':  '+shortname
            
            ax1.plot(xrange, y, ls=ls, lw=lw, marker=marker, color=color1, 
                    label=ptlegend, alpha=alpha)
            ax2.plot(xrange, yrel, ls=ls, lw=lw, marker=marker, color=color1, 
                    label=ptlegend, alpha=alpha)

        if patients == None:
            patients = self.patients
        
        for patient in patients:
            # read workbooks and sheets
            workbook_stent = os.path.join(self.exceldir,patient, self.workbook_analysis+patient[7:]+'.xlsx')
            wb = openpyxl.load_workbook(workbook_stent, data_only=True)
            sheetnames = wb.get_sheet_names()
            if analysis=='AngChim':
                for sheetname in ['Ang_RRA', 'Ang_LRA', 'Ang_SMA']:
                    if sheetname in sheetnames:
                        sheet = wb.get_sheet_by_name(sheetname)
                        angs, angsRel = readAnglesChimney(sheet,rows=[11,12],colStart=1, analysis=analysis)
                        self.angsAll.append(angs)
                        self.angsRelAll.append(angsRel)
                        # plot
                        plot_per_chimney(sheetname, analysis, patient, xrange, angs, angsRel)
            elif analysis == 'AngChimNel':
                for sheetname in ['Ang_RRA_Nel', 'Ang_LRA_Nel', 'Ang_SMA_Nel']:
                    if sheetname in sheetnames:
                        sheet = wb.get_sheet_by_name(sheetname)
                        angs, angsRel = readAnglesChimney(sheet,rows=[8,6],colStart=1, analysis=analysis)
                        self.angsAll.append(angs)
                        self.angsRelAll.append(angsRel)
                        # plot
                        plot_per_chimney(sheetname, analysis, patient, xrange, angs, angsRel)
            elif analysis == 'AngChimVessel':
                for sheetname in ['Ang_RRA_Vessel', 'Ang_LRA_Vessel', 'Ang_SMA_Vessel']:
                    if sheetname in sheetnames:
                        sheet = wb.get_sheet_by_name(sheetname)
                        angs, angsRel = readAnglesChimney(sheet,rows=[8,6],colStart=1, analysis=analysis)
                        self.angsAll.append(angs)
                        self.angsRelAll.append(angsRel)
                        # plot
                        plot_per_chimney(sheetname, analysis, patient, xrange, angs, angsRel)
            elif analysis=='Tort': 
                for sheetname in ['Tort_RRA', 'Tort_LRA', 'Tort_SMA']:
                    if sheetname in sheetnames:
                        torts, tortsRel = readAnglesChimney(sheet,rows=[3,7],colStart=1, analysis=analysis)
                        self.tortsAll.append(torts)
                        self.tortsRelAll.append(tortsRel)
                        # plot
                        plot_per_chimney(sheetname, analysis, patient, xrange, torts, tortsRel)
            
        ax2.legend(loc='upper right', fontsize=fontsize3, numpoints=2, title='Legend:')
        _initaxis([ax1,ax2])
        
        if saveFig:
            plt.savefig(os.path.join(self.dirsaveIm, 
            'plot_angles_chimney{}.png'.format(analysis)), papertype='a0', dpi=600)
        

def readLocationPointDeflection(sheet, row=9, rowlength=10, colStart=1):
    """ Read location of point with max deflection, obtain as percentage of chimney length
    """
    distfromproxendchimney = sheet.rows[row-1][colStart].value
    lengthchimney = sheet.rows[rowlength-1][colStart].value
    pointlocation = 100*distfromproxendchimney/lengthchimney
    
    return pointlocation, distfromproxendchimney, lengthchimney
    
def readLocationChange(sheet, row=19, colStart=1, nphases=10):
    """ Read locations of peak angle and obtain change in location
    """
    distances = sheet.rows[row-1][colStart:colStart+nphases]
    distances = [obj.value for obj in distances]
    locationchange = max(distances) - min(distances)
    
    return locationchange

def readMaxChange(sheet, row=6, colStart=1):
    """ Read the maximum distance change during cardiac cycle, e.g. distance or angle
    rows as excel
    """
    distchange = sheet.rows[row-1][colStart].value
    
    return distchange

def readMinMax(sheet, row=17, colStart=1, correctorientation=True):
    """ Read the min and max value during cardiac cycle, e.g. for angle
    rows as excel
    angles defined such that 0 = straight, 60 = sharp, 110 = severe
    """
    if correctorientation:
        maxvalue = 180 - sheet.rows[row-1][colStart].value # B
        minvalue = 180 - sheet.rows[row-1][colStart+1].value # C
    else:
        minvalue = sheet.rows[row-1][colStart].value # B
        maxvalue = sheet.rows[row-1][colStart+1].value # C
    
    return minvalue, maxvalue
    
def readSegmentDisplacement(sheet, rows=[4,5,6,7], colStart=1):
    """  Read the segment mean displacement amplitude
    rows as excel
    """
    rowStart3d= rows[0]-1
    rowStartx= rows[1]-1
    rowStarty= rows[2]-1
    rowStartz= rows[3]-1
    # read mean displacement amplitude
    displ3d = sheet.rows[rowStart3d][colStart].value
    displX = sheet.rows[rowStartx][colStart].value
    displY = sheet.rows[rowStarty][colStart].value
    displZ = sheet.rows[rowStartz][colStart].value
    
    return displX, displY, displZ, displ3d

def readPositionPhases(sheet,rows=[8, 9],colStart=1, nphases=10):
    """ Read position of the CoM point at avgreg and at each phase in cycle
    rows as excel rows
    """ 
    rowStart1 = rows[0]-1
    rowStart2 = rows[1]-1
    # read mean position avgreg
    obj = sheet.rows[rowStart2][colStart]
    pavg = obj.value.strip('()')
    pavg = pavg.split(',')
    pavg = float(pavg[0]), float(pavg[1]), float(pavg[2]) # x,y,z
    pavg = np.asarray(pavg)
    
    def convert_str_tuple(coord):
        return float(coord[0]), float(coord[1]), float(coord[2])
    
    objs = sheet.rows[rowStart1][colStart:colStart+nphases]
    pphases = [obj.value.strip('()') for obj in objs]
    pphases = [coord.split(',') for coord in pphases]
    pphases = [convert_str_tuple(coord) for coord in pphases] 
    pphases = np.asarray(pphases)
    
    pdisplacement = pphases - pavg
    
    return pavg, pphases, pdisplacement

def readDistancesBetweenPoints(sheet,rows=[9,5],colStart=1, nphases=10):
    """ read distances over all phases cardiac cycle in excel
    """
    # read distances
    rowStart1 = rows[0]-1
    rowStart2 = rows[1]-1
    dists = sheet.rows[rowStart1][colStart:colStart+nphases] # B to K, 10 phases 
    dists = [obj.value for obj in dists]
    dists = np.asarray(dists)
    obj = sheet.rows[rowStart2][colStart] # avgreg distance
    avgdist = obj.value # get distance at mid heart cycle
    # relative distances from avgreg
    distsRel = dists - avgdist
    
    return dists, distsRel


def readAnglesChimney(sheet,rows=[11, 12],colStart=1, analysis='AngChim'):
    """ read angles over all phases cardiac cycle in excel
    analysis: 'AngChim' or AngChimNel or AngChimVessel or 'Tort'
    """
    # read distances
    rowStart1 = rows[1]-1 # angles
    rowStart2 = rows[0]-1 # ang mid cycle
    
    # colStart = 1 # B
    # angles / tort
    angs = sheet.rows[rowStart1][colStart:colStart+10] # B to K, 10 phases 
    if 'AngChim' in analysis: # for AngChim, AngChimNel and AngChimVessel
        angs = [180-obj.value for obj in angs] # so that 0 is straight: 70 is scherpere hoek dan 40
        angs = np.asarray(angs)
        # get angle mid cycle
        obj = sheet.rows[rowStart2][colStart]
        avgang = 180-obj.value
    elif analysis == 'Tort':
        angs = [obj.value for obj in angs] 
        angs = np.asarray(angs)
        # get tort mid cycle
        obj = sheet.rows[rowStart2][colStart]
        avgang = obj.value
    
    # relative angles from avgreg angle (or tort)
    angsRel = angs - avgang # change with respect to mid cycle angle
    
    return angs, angsRel

def create_iter_colors(type=1):
    if type == 1:
        colors = itertools.cycle([  
                                '#a6cee3', # 1
                                '#fb9a99', # 2 
                                '#33a02c', # 3
                                '#fdbf6f', # 4 
                                '#1f78b4', # 5
                                '#e31a1c', # 6
                                '#b2df8a', # 7 
                                '#cab2d6', # 8 
                                '#ff7f00', # 9
                                '#6a3d9a', # 10
                                '#ffff99', # 11
                                '#b15928'])# 12
    
    elif type == 2:
        colors =  itertools.cycle([
                                    '#d73027', # 1
                                    '#fc8d59',
                                    #'#fee090', # yellow
                                    '#91bfdb',
                                    '#4575b4' # 5
                                    ])
    
    return colors
    
def color_per_patient(patient):
    colors = [
                                '#a6cee3', # 1
                                '#fb9a99', # 2 
                                '#33a02c', # 3
                                '#fdbf6f', # 4 
                                '#1f78b4', # 5
                                '#e31a1c', # 6
                                '#b2df8a', # 7 
                                '#cab2d6', # 8 
                                '#ff7f00', # 9
                                ]
    if patient == 'chevas_01':
        color = colors[0] 
    elif patient == 'chevas_02':
        color = colors[1] 
    elif patient == 'chevas_03':
        color = colors[2] 
    elif patient == 'chevas_04':
        color = colors[3] 
    elif patient == 'chevas_05':
        color = colors[4] 
    elif patient == 'chevas_06':
        color = colors[5] 
    elif patient == 'chevas_07':
        color = colors[6]
    elif patient == 'chevas_08':
        color = colors[7] 
    elif patient == 'chevas_09':
        color = colors[0] # pt 1
    elif patient == 'chevas_10':
        color = colors[-1] # color 9th pt
    elif patient == 'chevas_11':
        color = colors[6] # pt 7
    elif patient == 'chevas_09_thin':
        color = colors[0] # pt 1
    
    return color
    
def create_iter_markers(type=1):
    if type == 1:
        markers = itertools.cycle([
                                    'o', 'o', 'o', 'o',
                                    '^', '^', '^', '^',
                                    's', 's', 's', 's',
                                    'd', 'd', 'd', 'd'  ])
    if type == 2:
        markers = itertools.cycle([
                                    'o', 'o', 'o',
                                    '^', '^', '^',
                                    's', 's', 's'  ])
    
    
    if type == 3:
        markers = itertools.cycle(['o', '^', 's','d'])
    
    return markers

def marker_per_chimney(shortname):
    if 'RRA' in shortname:
        marker = 'o'
    elif 'LRA' in shortname:
        marker = 's'
    elif 'SMA' in shortname:
        marker = '^'
    elif 'NelL' in shortname:
        marker = '^'
    elif 'NelR' in shortname:
        marker = 'd'
    else:
        marker = '^'
    return marker

def color_per_stent(shortname):
    colors = [
                                '#a6cee3', # 1
                                '#fb9a99', # 2 
                                '#33a02c', # 3
                                '#fdbf6f', # 4 
                                '#1f78b4', # 5
                                '#e31a1c', # 6
                                '#b2df8a', # 7 
                                '#cab2d6', # 8 
                                '#ff7f00', # 9
                                ]
    if 'RRA' in shortname:
        color = colors[0]
    elif 'LRA' in shortname:
        color = colors[1]
    elif 'SMA' in shortname:
        color = colors[3]
    elif 'NelL' in shortname:
        color = colors[0]
    elif 'NelR' in shortname:
        color = colors[1]
    elif 'Nel-Nel' in shortname:
        color = colors[4]
    
    return color

def create_iter_ls(type=1):
    if type == 1:
        lstyles = itertools.cycle(['-'])
    
    elif type == 2:
        lstyles = itertools.cycle(['-', '-', '-', '-', '--','--','--','--'])
    
    elif type == 3:
        lstyles = itertools.cycle(['-', '--']) #, '.-'])

    return lstyles

if __name__ == '__main__':
    
    foo = ExcelAnalysisNellix()
    
    patients = None # None = all in self.patients
    # patients=['chevas_09', 'chevas_09_thin' ]
    
    # Plots
    # ==========================================
    # Distances
    # foo.plot_distances_between_points(patients=patients,analysis=['NelNel'], ylim=[8,16], saveFig=False) 
    # foo.plot_distances_between_points(patients=patients,analysis=['ChimNel'], ylim=[5,32], saveFig=False)
    
    # Angles pointdeflection chimney
    # foo.plot_angles_chimney(patients=patients,analysis='AngChim', ylim=[0, 35.01], ylimRel=[-3,3.01], saveFig=True)
    
    # Angles vectors chimney nellix
    # foo.plot_angles_chimney(patients=patients,analysis='AngChimNel', ylim=[100, 170.01], ylimRel=[-3,3.01], saveFig=False)
    
    # Angles vectors chimney to vessel transition (end-stent angle)
    # foo.plot_angles_chimney(patients=patients,analysis='AngChimVessel', ylim=[0, 85.01], ylimRel=[-3,3.01], saveFig=True)
    
    # # Tortuosity
    # foo.plot_angles_chimney(patients=patients,analysis='Tort', ylim=[0.99, 1.11], ylimRel=[-0.01,0.01], saveFig=False)
    
    # # Displacement
    # foo.plot_displacement(patients=patients, analysis=['NelRprox','NelLprox' ], 
    #             rows=[8,9], ylim=[-0.51, 0.5], saveFig=True)
    # foo.plot_displacement(patients=patients, analysis=['LRAprox', 'RRAprox', 'SMAprox' ], 
    #             rows=[8,9], ylim=[-0.51, 0.5], saveFig=True)
    # foo.plot_displacement(patients=patients, analysis=['LRAdist', 'RRAdist', 'SMAdist' ], 
    #             rows=[8,9], ylim=[-0.51, 0.5], saveFig=True)
    # foo.plot_displacement(patients=patients, analysis=['vLRAprox', 'vRRAprox', 'vSMAprox' ], 
    #             rows=[8,9], ylim=[-0.51, 0.5], saveFig=True)
    
    # ==========================================
    
    # Get statistics
    # ==========================================
    ## Get displacement of centerline segment prox or dist
     
    # foo.get_segment_displacement(patients=patients, analysis=['NelL', 'NelR'], location='prox')
    # foo.get_segment_displacement(patients=patients, analysis=['LRA','RRA', 'SMA'], location='dist')
    # print(len(foo.segmentDisplacementX)) # verify number of chimneys/nellix stents
    # outcomeDispl = foo.segmentDisplacement3d
    # foo.get_segment_displacement(patients=patients, analysis=['LRA','RRA','SMA'], location='prox')
    # if False:
    #     t, p = independent_samples_ttest(foo.segmentDisplacement3d, outcomeDispl)
    
    ## Get displacement of centerline segment vessel distal to stent (location is always prox)
    
    # foo.get_segment_displacement(patients=patients, analysis=['vLRA'], location='prox')
    # foo.get_segment_displacement(patients=patients, analysis=['vLRA','vRRA', 'vSMA'], location='prox')
    # print(len(foo.segmentDisplacementX)) # verify number of chimneys/nellix stents
    
    ## Get distance change between Nellix stents or between Nellix and chimney ends
    
    # foo.get_distance_change(patients=patients, analysis='ChimNel', chimneys=['LRA', 'RRA', 'SMA'])
    # foo.get_distance_change(patients=patients, analysis='NelNel')
    # print(len(foo.distanceChange)) # verify number of chimneys/nellix stents
    # if False:
    #     outcomeNel = foo.distanceChange
    #     t, p = independent_samples_ttest(foo.distanceChange, outcomeNel)
    
    ## Get chimney angle change
    
    # foo.get_angle_change(patients=patients, analysis='Chim', chimneys=['LRA', 'RRA', 'SMA'], angletype = 'pointdeflection') # pointdeflection or peakangle 
    # print(len(foo.angleChange))
    # outcomeChim = foo.angleChange
    # outcomeChimMin = foo.angleMin
    # outcomeChimMax = foo.angleMax
    # foo.get_angle_change(patients=patients, analysis='Chim', chimneys=['SMA'], angletype = 'pointdeflection')
    # print(len(foo.angleChange)) # verify number of chimneys
    # if True:
    #     t, p = independent_samples_ttest(foo.angleChange, outcomeChim)
    # if True:
    #     t, p = independent_samples_ttest(foo.angleMin, outcomeChimMin)
    #     t, p = independent_samples_ttest(foo.angleMax, outcomeChimMax)
    
    ## Get chimney-Nellix vector angle change
    # 
    # foo.get_angle_change(patients=patients, analysis='ChimNel', chimneys=['LRA','RRA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # outcomeChimNel = foo.angleChange
    # outcomeChimNelMin = foo.angleMin
    # outcomeChimNelMax = foo.angleMax
    # foo.get_angle_change(patients=patients, analysis='ChimNel', chimneys=['SMA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # if True:
    #     t, p = independent_samples_ttest(foo.angleChange, outcomeChimNel)
    # if True:
    #     t, p = independent_samples_ttest(foo.angleMin, outcomeChimNelMin)
    #     t, p = independent_samples_ttest(foo.angleMax, outcomeChimNelMax)
    
    ## Get chimney-vessel vector angle change
    
    # foo.get_angle_change(patients=patients, analysis='ChimVessel', chimneys=['LRA', 'RRA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # outcomeChimVessel = foo.angleChange
    # outcomeChimVesselMin = foo.angleMin
    # outcomeChimVesselMax = foo.angleMax
    # foo.get_angle_change(patients=patients, analysis='ChimVessel', chimneys=['SMA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # if True:
    #     t, p = independent_samples_ttest(foo.angleChange, outcomeChimVessel)
    # if True:
    #     t, p = independent_samples_ttest(foo.angleMin, outcomeChimVesselMin)
    #     t, p = independent_samples_ttest(foo.angleMax, outcomeChimVesselMax)
    
    ## Compare chimney-Nellix vector angle change with end-stent angle change
    
    # foo.get_angle_change(patients=None, analysis='ChimNel', chimneys=['LRA','RRA','SMA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # outcomeChimNel = foo.angleChange
    # foo.get_angle_change(patients=None, analysis='ChimVessel', chimneys=['LRA', 'RRA', 'SMA'])
    # print(len(foo.angleChange)) # verify number of chimneys
    # if True:
    #     t, p = independent_samples_ttest(foo.angleChange, outcomeChimNel)
    
    # ==========================================
    
    
    ## Compare length of chimney stents between single, double and triple for angle
    # 
    # foo.get_angle_change(patients=None, analysis='Chim', chimneys=['LRA', 'RRA','SMA'], angletype = 'peakangle')
    # lensingles = [foo.lengthchimneys[i] for i in [0,3,4,5,9] ]
    # lendoubles = [foo.lengthchimneys[i] for i in [1,2,10,11,12,13] ]
    # lentriples = [foo.lengthchimneys[i] for i in [6,7,8,14,15,16,17,18,19] ]
    # lendoublestriples = [foo.lengthchimneys[i] for i in [1,2,10,11,12,13,6,7,8,14,15,16,17,18,19] ]
    # 
    # print('Average length single chimneys: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
    #                                     np.mean(lensingles),
    #                                     np.std(lensingles),
    #                                     np.min(lensingles),
    #                                     np.max(lensingles)
    #                                     ))
    # print('Average length double chimneys: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
    #                                     np.mean(lendoubles),
    #                                     np.std(lendoubles),
    #                                     np.min(lendoubles),
    #                                     np.max(lendoubles)
    #                                     ))
    # print('Average length triple chimneys: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
    #                                     np.mean(lentriples),
    #                                     np.std(lentriples),
    #                                     np.min(lentriples),
    #                                     np.max(lentriples)
    #                                     ))
    # print('Average length double and triple chimneys: {:.1f} ± {:.1f} ({:.1f}-{:.1f})'.format(
    #                                     np.mean(lendoublestriples),
    #                                     np.std(lendoublestriples),
    #                                     np.min(lendoublestriples),
    #                                     np.max(lendoublestriples)
    #                                     ))
    # if True:
    #     t, p = independent_samples_ttest(lensingles, lendoublestriples)
    
  
    # ==========================================
    
    # Compare 09 with 09-thin
    
    # ## angles
    # patients=['chevas_09']
    # 
    # foo.get_angle_change(patients=patients, analysis='ChimNel', chimneys=['LRA', 'RRA','SMA'], angletype = 'pointdeflection')
    # p09 = np.asarray(foo.angleChange)
    # 
    # patients2=['chevas_09_thin']
    # foo.get_angle_change(patients=patients2, analysis='ChimNel', chimneys=['LRA', 'RRA','SMA'], angletype = 'pointdeflection')
    # p09thin = np.asarray(foo.angleChange)
    # 
    # print(p09)
    # print()
    # print(p09thin)
    # print()
    # print(p09-p09thin)
    # print()
    # meandiff = np.mean(abs(p09-p09thin))
    # print(meandiff)
    # print(np.std(abs(p09-p09thin)))
    # print()
    # print(np.max(abs(p09-p09thin)))
    
    # ## distances
    # 
    # patients=['chevas_09']
    # 
    # foo.get_distance_change(patients=patients, analysis='ChimNel', chimneys=['LRA', 'RRA', 'SMA'])
    # p09 = np.asarray(foo.distanceChange)
    # 
    # patients2=['chevas_09_thin']
    # foo.get_distance_change(patients=patients2, analysis='ChimNel', chimneys=['LRA', 'RRA', 'SMA'])
    # p09thin = np.asarray(foo.distanceChange)
    # 
    # print(p09)
    # print()
    # print(p09thin)
    # print()
    # print(p09-p09thin)
    # print()
    # meandiff = np.mean(abs(p09-p09thin))
    # print(meandiff)
    # print(np.std(abs(p09-p09thin)))
    # print()
    # print(np.max(abs(p09-p09thin)))
    
    # ## displacement
    # 
    # patients=['chevas_09']
    # foo.get_segment_displacement(patients=patients, analysis=['LRA','RRA', 'SMA'], location='prox')
    # d3d = foo.segmentDisplacement3d
    # foo.get_segment_displacement(patients=patients, analysis=['LRA','RRA', 'SMA'], location='dist')
    # p09 = np.asarray(d3d+foo.segmentDisplacement3d)
    # 
    # patients2=['chevas_09_thin']
    # foo.get_segment_displacement(patients=patients2, analysis=['LRA','RRA', 'SMA'], location='prox')
    # d3d = foo.segmentDisplacement3d
    # foo.get_segment_displacement(patients=patients2, analysis=['LRA','RRA', 'SMA'], location='dist')
    # p09thin = np.asarray(d3d+foo.segmentDisplacement3d)
    # 
    # print(p09)
    # print()
    # print(p09thin)
    # print()
    # print(p09-p09thin)
    # print()
    # meandiff = np.mean(abs(p09-p09thin))
    # print(meandiff)
    # print(np.std(abs(p09-p09thin)))
    # print()
    # print(np.max(abs(p09-p09thin)))
    # print()
    # print(np.max(abs(p09-p09thin)))

